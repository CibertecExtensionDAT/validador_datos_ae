import io
import re
import streamlit as st
import pandas as pd
import os
import PyPDF2
import zipfile
from io import BytesIO
from openpyxl import Workbook
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.colors import HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from zipfile import ZipFile
from datetime import datetime
from tempfile import NamedTemporaryFile
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from PIL import Image, ImageDraw, ImageFont

st.set_page_config(
    page_title="Validador de Archivos",
    page_icon="üìä",
    layout="wide"
)

if "paso_actual" not in st.session_state:
    st.session_state.paso_actual = 0
if "nombre_colegio" not in st.session_state:
    st.session_state.nombre_colegio = ""
if "comparador_archivo_base" not in st.session_state:
    st.session_state.comparador_archivo_base = None
if "comparador_archivo_revisar" not in st.session_state:
    st.session_state.comparador_archivo_revisar = None
if "comparador_resultados" not in st.session_state:
    st.session_state.comparador_resultados = None
if "comparador_comparacion_realizada" not in st.session_state:
    st.session_state.comparador_comparacion_realizada = False
if "comparador_reset_counter" not in st.session_state:
    st.session_state.comparador_reset_counter = 0
if "archivo1_df" not in st.session_state:
    st.session_state.archivo1_df = None
if "archivo2_df" not in st.session_state:
    st.session_state.archivo2_df = None
if "archivo2_1p3p_df" not in st.session_state:
    st.session_state.archivo2_1p3p_df = None
if "archivo2_4p5s_df" not in st.session_state:
    st.session_state.archivo2_4p5s_df = None
if "archivo1_bytes" not in st.session_state:
    st.session_state.archivo1_bytes = None
if "archivo2_bytes" not in st.session_state:
    st.session_state.archivo2_bytes = None
if "archivo1_fila_cabecera" not in st.session_state:
    st.session_state.archivo1_fila_cabecera = None
if "archivo2_1p3p_fila_cabecera" not in st.session_state:
    st.session_state.archivo2_1p3p_fila_cabecera = None
if "archivo2_4p5s_fila_cabecera" not in st.session_state:
    st.session_state.archivo2_4p5s_fila_cabecera = None
if "cursos_equivalentes" not in st.session_state:
    st.session_state.cursos_equivalentes = [
    "ADOBE ILLUSTRATOR",
    "ADOBE INDESING",
    "ADOBE PHOTOSHOP PROFICIENT",
    "COACHING PERSONAL Y VOCACIONAL",
    "CODE FOR KIDS",
    "DE LA IDEA AL EMPRENDIMIENTO",
    "DESARROLLO DE APLICACIONES M√ìVILES",
    "DESARROLLO WEB",
    "DISE√ëO CREATIVO CON INTELIGENCIA ARTIFICIAL (IA)",
    "DISE√ëO WEB",
    "EDICI√ìN DE AUDIO",
    "EDICI√ìN DE VIDEO",
    "EXCEL EXPERT SPECIALIST",
    "EXCEL INTERMEDIATE SPECIALIST",
    "EXCEL PROFICIENT SPECIALIST",
    "EXPLORANDO LA INTELIGENCIA ARTIFICIAL (IA)",
    "FINANZAS PERSONALES",
    "GESTI√ìN DE DATA CON GOOGLE SHEETS & LOCKER STUDIO",
    "GESTI√ìN DE DATA CON MS EXCEL Y POWER BI",
    "GESTI√ìN EMPRESARIAL",
    "GOOGLE DOCS AVANZADO",
    "GOOGLE DOCS B√ÅSICO",
    "GOOGLE SHEETS AVANZADO",
    "GOOGLE SHEETS B√ÅSICO",
    "GOOGLE SHEETS INTERMEDIO",
    "GOOGLE SLIDES PRESENTACIONES DE IMPACTO",
    "HABILIDADES BLANDAS",
    "INNOVACI√ìN DIGITAL CON INTELIGENCIA ARTIFICIAL (IA)",
    "LEARNING FOR BEGINNERS 1",
    "LEARNING FOR BEGINNERS 2",
    "MARKETING DIGITAL",
    "MARKETING PERSONAL",
    "PRESENTACIONES DE IMPACTO",
    "PROGRAMACI√ìN VISUAL KODU PLANET I",
    "PROGRAMACI√ìN VISUAL KODU PLANET II",
    "PROGRAMACI√ìN VISUAL KODU PLANET III",
    "ROBLOX FOR TEENS",
    "ROB√ìTICA",
    "SCRATCH",
    "WORD EXPERT SPECIALIST",
    "WORD PROFICIENT SPECIALIST"
]
if "tab3_archivo_procesado" not in st.session_state:
    st.session_state.tab3_archivo_procesado = False
if "tab3_df_reporte" not in st.session_state:
    st.session_state.tab3_df_reporte = None
if "tab3_nombre_colegio" not in st.session_state:
    st.session_state.tab3_nombre_colegio = ""
if "tab3_tipo_archivo" not in st.session_state:
    st.session_state.tab3_tipo_archivo = ""
if "tab3_reset_counter" not in st.session_state:
    st.session_state.tab3_reset_counter = 0
if 'tipo_certificado_seleccionado' not in st.session_state:
    st.session_state.tipo_certificado_seleccionado = None
if 'usar_marca_agua_seleccionado' not in st.session_state:
    st.session_state.usar_marca_agua_seleccionado = False
if 'fecha_certificado_seleccionada' not in st.session_state:
    st.session_state.fecha_certificado_seleccionada = datetime.now().date()
if 'tab4_reset_counter' not in st.session_state:
    st.session_state.tab4_reset_counter = 0
if 'df_procesado' not in st.session_state:
    st.session_state.df_procesado = None
if 'grupos' not in st.session_state:
    st.session_state.grupos = None
if 'plantillas' not in st.session_state:
    st.session_state.plantillas = None
if 'certificados_generados' not in st.session_state:
    st.session_state.certificados_generados = False
if 'zip_buffer' not in st.session_state:
    st.session_state.zip_buffer = None
if 'tab5_reset_counter' not in st.session_state:
    st.session_state.tab5_reset_counter = 0

COLUMNAS_ARCHIVO1 = [
    "NRO.", "PATERNO", "MATERNO", "NOMBRES", "NACIMIENTO (DD/MM/YYYY)", "SEXO (M/F)",
    "GRADO", "SECCI√ìN", "CORREO INSTITUCIONAL", "NEURODIVERSIDAD (S√ç/NO)", "DNI"
]

COLUMNAS_TAB03 = [
    "NRO.", "PATERNO", "MATERNO", "NOMBRE", "GRADO", "SECCI√ìN", "CURSO", "NOTA LABORATORIO", "¬øASISTI√ì?", "P1 4PTOS.", "P2 4PTOS.", "P3 4PTOS.", "P4 4PTOS.", "P5 4PTOS.", "NOTA EVALUADOR", "NOTA FINAL", "OBSERVADOS", "ESTATUS", "NUMERACI√ìN"
]

COLUMNAS_ARCHIVO2_1P3P = [
    "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", "GRADO", "SECCI√ìN", "NOTA VIGESIMAL 100%"
]

COLUMNAS_ARCHIVO2_4P5S = [
    "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", "GRADO", "SECCI√ìN", "NOTA VIGESIMAL 25%"
]

COLUMNAS_EVALUADOR = [
    "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", "GRADO", "SECCI√ìN", 
    "NOTA VIGESIMAL 25%", "NOTAS VIGESIMALES 75%", "PROMEDIO", "OBSERVADOS"
]

COLUMNAS_ARCHIVO_PDF_1P3P = [
    'NRO.', 'PATERNO', 'MATERNO', 'NOMBRES', 'CURSO', 'GRADO', 'SECCI√ìN', 'NOTA VIGESIMAL 100%'
]

COLUMNAS_ARCHIVO_PDF_4P5S = [
    'NRO.', 'PATERNO', 'MATERNO', 'NOMBRES', 'CURSO', 'GRADO', 'SECCI√ìN', 'NOTA VIGESIMAL 25%'
]

SEXO_VALIDO = ["M", "F"]
SECCIONES_VALIDAS = ["A", "B", "C", "D", "E", "F", "G", "U", "UNICO", "UNICA", "√öNICO", "√öNICA", "√önico", "√önica"]
GRADOS_VALIDOS = ["1P", "2P", "3P", "4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"]
MAPEO_GRADOS = {
    "1": "1P", "2": "2P", "3": "3P", "4": "4P", "5": "5P", "6": "6P",
    "7": "1S", "8": "2S", "9": "3S", "10": "4S", "11": "5S"
}
MAPEO_SECCIONES = {
    "UNICO": "U",
    "UNICA": "U",
    "√öNICO": "U", 
    "√öNICA": "U", 
    "√önico": "U", 
    "√önica": "U"
}
LISTA_COLEGIOS = [
    "COLEGIO ANDINO",
    "COLEGIO ANNIES SCHOOL",
    "COLEGIO ANTARES",
    "COLEGIO ATENEO",
    "COLEGIO BARBARA D'ACHILLE",
    "COLEGIO BUENAS NUEVAS",
    "COLEGIO CEPEBAN",
    "COLEGIO CERVANTES SCHOOL",
    "COLEGIO CRISTIANO PIONERO",
    "COLEGIO DIVINA MISERICORDIA",
    "COLEGIO DIVINO MAESTRO DE PRO",
    "COLEGIO DIVINO NI√ëO JES√öS",
    "COLEGIO DON BOSCO",
    "COLEGIO DORA MAYER",
    "COLEGIO EL CARMELO",
    "COLEGIO EL SEMBRADOR",
    "COLEGIO ELIM",
    "COLEGIO GIORDANO BRUNO",
    "COLEGIO GRACIAS JESUS",
    "COLEGIO INGENIER√çA",
    "COLEGIO INTEGRA COLLEGE",
    "COLEGIO JES√öS DEL PROGRESO",
    "COLEGIO JES√öS EL NAZARENO",
    "COLEGIO JES√öS MAR√çA",
    "COLEGIO JORDAN DE JESUS",
    "COLEGIO JOSE MARIA ARGUEDAS",
    "COLEGIO JOSEPH NOVAK",
    "COLEGIO JUAN ENRIQUE NEWMAN",
    "COLEGIO LISOFT",
    "COLEGIO LISSON",
    "COLEGIO LORIS MALAGUZZI",
    "COLEGIO LOS ROSALES DE SANTA ROSA",
    "COLEGIO LUZ CASANOVA",
    "COLEGIO MAESTRO REDENTOR",
    "COLEGIO MAGISTER",
    "COLEGIO MAKARENKO",
    "COLEGIO MARCELINO CHAMPAGNAT",
    "COLEGIO MARIA DE LAS MERCEDES (MIRAFLORES)",
    "COLEGIO MARIA DE LAS MERCEDES (LA VICTORIA)",
    "COLEGIO MARIA RAFOLS",
    "COLEGIO MARISTAS",
    "COLEGIO MASHAL SCHOOL",
    "COLEGIO MELVIN JONES",
    "COLEGIO MI AMIGO JESUS",
    "COLEGIO MIGUEL ANGEL",
    "COLEGIO MUNDIAL",
    "COLEGIO MY HOME AND SCHOOL",
    "COLEGIO NUESTRA SE√ëORA DEL CARMEN DE PALAO",
    "COLEGIO NUESTRA SE√ëORA DEL BUEN CONSEJO",
    "COLEGIO PARROQUIAL SAN JOSE",
    "COLEGIO PERUANO JAPON√âS LA VICTORIA",
    "COLEGIO PLAY SCHOOL",
    "COLEGIO REDIMER JESUS",
    "COLEGIO SAN ANTONIO DE PADUA",
    "COLEGIO SAN ANTONIO MAR√çA CLARET",
    "COLEGIO SAN CHARBEL",
    "COLEGIO SAN FRANCISCO DE BORJA",
    "COLEGIO SAN GERMAN",
    "COLEGIO SAN IGNACIO DE LOYOLA (MIRAFLORES)",
    "COLEGIO SAN IGNACIO DE LOYOLA (VILLA)",
    "COLEGIO SAN JUAN DE BARRANCO",
    "COLEGIO SAN MARTINCITO",
    "COLEGIO SANTA ANA DE LOS JARDINES",
    "COLEGIO SANTA ANA INGENIERIA",
    "COLEGIO SANTA ANA (LIMA)",
    "COLEGIO SANTA ANA (TACNA)",
    "COLEGIO SANTA ANGELA MERICI",
    "COLEGIO SANTA ANITA",
    "COLEGIO SANTA MARIA DE SURCO",
    "COLEGIO SHONA",
    "COLEGIO SANTA ROSA DE LIMA",
    "COLEGIO TRAVESURAS TRAZOS Y COLORES",
    "COLEGIO VIRGEN DEL ROSARIO YUNGAY"
]

CONFIG_INSIGNIAS = {
            'IDENTIFICADOR': {
                'font_size_inicial': 60,
                'max_width': 765,
                'max_height': 90,
                'min_font_size': 30
            },
            'CURSO': {
                'font_size_inicial': 60,
                'max_width': 765,
                'max_height': 200,
                'min_font_size': 25
            },
            'A√ëO': {
                'font_size_inicial': 65,
                'max_width': 400,
                'max_height': 80,
                'min_font_size': 40
            }
        }

def detectar_cabecera_automatica(df: pd.DataFrame, columnas_objetivo: list):
    """Detecta autom√°ticamente la fila de cabecera"""
    max_filas, max_cols = min(15, len(df)), min(25, len(df.columns))
    subset = df.iloc[:max_filas, :max_cols]
    columnas_objetivo_norm = [c.strip().lower() for c in columnas_objetivo]

    for idx in range(max_filas):
        fila = subset.iloc[idx].astype(str).str.strip().str.lower().tolist()
        if all(col in fila for col in columnas_objetivo_norm):
            return idx
    return None

def crear_identificador(df, col_paterno, col_materno, col_nombres):
    """Crea columna identificador"""
    return (
        df[col_nombres].astype(str).str.strip() + " " +
        df[col_paterno].astype(str).str.strip() + " " +
        df[col_materno].astype(str).str.strip()
    )

def normalizar_enie(texto):
    """
    Normaliza texto a may√∫sculas preservando TODOS los acentos (tildes y √ë)
    """
    if pd.isna(texto):
        return ""
    
    texto = str(texto).strip().upper()
    texto = ' '.join(texto.split())
    
    return texto

def limpiar_filas_vacias(df, columnas_clave=None):
    """
    Args:
        df: DataFrame a limpiar
        columnas_clave: Lista de nombres de columnas para verificar (default: primeras 4)
    
    Returns:
        DataFrame limpio sin filas completamente vac√≠as (evita s√≥lo los Nro o N¬∞)
    """
    if columnas_clave is None:
        columnas_clave = df.columns[1:4].tolist()
    
    total_original = len(df)
    
    df_limpio = df.dropna(subset=columnas_clave, how='all').copy()
    
    if df_limpio.empty:
        st.warning(f"‚ö†Ô∏è La hoja est√° vac√≠a o no contiene datos v√°lidos (se ignorar√°)")
        return df_limpio

    mask = df_limpio[columnas_clave].apply(
        lambda x: x.astype(str).str.strip().ne('')
    ).any(axis=1)
    df_limpio = df_limpio[mask].reset_index(drop=True)
    
    filas_eliminadas = total_original - len(df_limpio)
    if filas_eliminadas > 0:
        st.info(f"üßπ Se eliminaron {filas_eliminadas} filas vac√≠as (quedaron {len(df_limpio)} registros)")
    
    return df_limpio

def homologar_dataframe(df):
    """
    Homologa un DataFrame completo:
    - Todas las columnas: Convierte a may√∫sculas y quita espacios
    - Columnas PATERNO, MATERNO, NOMBRES: Adem√°s quita acentos y mantiene la √ë
    """

    if df.empty:
        return df

    columnas_nombres = ["PATERNO", "MATERNO", "NOMBRES"]
    filas_vacias = df[df[columnas_nombres].isnull().any(axis=1)]

    if not filas_vacias.empty:
        st.error("‚ùå Se detectaron campos vac√≠os en nombres o apellidos (Archivo 1 - N√≥mina)")
        st.dataframe(filas_vacias, use_container_width=True)
        st.stop()

    for col in df.columns:
        if col.upper() in columnas_nombres:
            df[col] = df[col].apply(normalizar_enie)
            df[col] = df[col].str.replace(r'\s+', ' ', regex=True).str.strip()
        else:
            df[col] = (
                df[col].astype(str)
                .str.strip()
                .str.upper()
                .str.replace(r'\s+', ' ', regex=True)
                .str.strip()
            )
    
    return df

def convertir_numericas_a_entero(df, columnas=None):
    """
    Convierte valores num√©ricos flotantes a enteros (1.0 ‚Üí "1")
    Funciona incluso en columnas mixtas (1.0, "2P", 3.0)
    
    Args:
        df: DataFrame a procesar
        columnas: Lista de columnas a convertir
    
    Returns:
        DataFrame con columnas convertidas
    """

    if df.empty:
        return df

    if columnas is None:
        columnas = df.select_dtypes(include=['float64', 'float32']).columns.tolist()
    
    for col in columnas:
        if col not in df.columns:
            continue
        
        def convertir_valor(val):
            """Convierte un valor individual"""
            if pd.isna(val):
                return val
            
            val_str = str(val).strip()
            
            if any(c.isalpha() for c in val_str):
                return val_str
            
            try:
                val_num = float(val)
                if val_num % 1 == 0:
                    return str(int(val_num))
                else:
                    return val_str
            except (ValueError, TypeError):
                return val_str
        
        df[col] = df[col].apply(convertir_valor)
    
    return df

def validar_y_mapear_grados(df, col_grado="GRADO", tipo_validacion="todos"):
    """
    Valida y mapea los grados. Convierte n√∫meros 1-11 a formato est√°ndar (1P-6P, 1S-5S).
    Retorna DataFrame procesado y lista de errores.
    
    Args:
        df: DataFrame a validar
        col_grado: Nombre de la columna de grado
        tipo_validacion: Tipo de validaci√≥n a aplicar:
            - "todos": Valida todos los grados (1P-6P, 1S-5S) - Para Archivo 1
            - "1p3p": Solo valida 1P, 2P, 3P - Para hoja 1P-3P del Archivo 2
            - "4p5s": Solo valida 4P-6P, 1S-5S - Para hoja 4P-5S del Archivo 2
    """
    errores = []
    df[col_grado] = df[col_grado].astype(str).str.strip().str.upper()
    
    if tipo_validacion == "1p3p":
        mapeo_grados = {
            "1": "1P", "2": "2P", "3": "3P"
        }
        grados_validos = ["1P", "2P", "3P"]
    elif tipo_validacion == "4p5s":
        mapeo_grados = {
            "4": "4P", "5": "5P", "6": "6P",
            "7": "1S", "8": "2S", "9": "3S", "10": "4S", "11": "5S"
        }
        grados_validos = ["4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"]
    else:  # "todos"
        mapeo_grados = MAPEO_GRADOS
        grados_validos = GRADOS_VALIDOS
    
    df[col_grado] = df[col_grado].replace(mapeo_grados)
    grados_invalidos = df.loc[~df[col_grado].isin(grados_validos)]

    if len(grados_invalidos) > 0:
        for idx, row in grados_invalidos.iterrows():
            errores.append(f"Fila {idx + 2}: Grado inv√°lido '{row[col_grado]}'")
    
    return df, errores

def inferir_sexo_por_nombre(nombre):
    """
    Infiere el sexo bas√°ndose en el nombre.
    Retorna 'M' o 'F' seg√∫n terminaciones comunes en espa√±ol.
    """
    if pd.isna(nombre) or str(nombre).strip() == "":
        return "M"
    
    nombre = str(nombre).strip().upper()
    primer_nombre = nombre.split()[0] if nombre else ""
    
    terminaciones_femeninas = ['A', 'IA', 'INA', 'ELA', 'ANA', 'LIA', 'RIA', 'TA', 'DA']
    nombres_femeninos = ['MARIA', 'CARMEN', 'ROSA', 'LUZ', 'SOL', 'MERCEDES', 'BEATRIZ', 'INES', 'ISABEL']
    
    if primer_nombre in nombres_femeninos:
        return "F"
    
    for term in terminaciones_femeninas:
        if primer_nombre.endswith(term):
            return "F"
    
    return "M"

def validar_sexo(df, col_sexo="SEXO (M/F)"):
    """
    Valida que el sexo sea M o F.
    Si est√° vac√≠o, infiere el sexo seg√∫n el nombre del alumno.
    Retorna lista de errores (ahora solo para casos que no se puedan resolver).
    """
    errores = []
    df[col_sexo] = df[col_sexo].astype(str).str.strip().str.upper()
    mask_vacios_invalidos = ~df[col_sexo].isin(SEXO_VALIDO)
    
    if mask_vacios_invalidos.any():
        for idx in df[mask_vacios_invalidos].index:
            nombre = df.loc[idx, "NOMBRES"] if "NOMBRES" in df.columns else ""
            sexo_inferido = inferir_sexo_por_nombre(nombre)
            df.loc[idx, col_sexo] = sexo_inferido
            identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
            errores.append(f"INFO - Fila {idx + 2}: Sexo vac√≠o/inv√°lido, se asign√≥ '{sexo_inferido}' seg√∫n nombre - {identificador}")
    return errores

def validar_secciones(df, col_seccion="SECCI√ìN"):
    """
    Valida que las secciones sean v√°lidas (A-G, U, UNICO, UNICA y estas dos √∫ltimas reemplazarlas por U).
    Retorna lista de errores.
    """
    errores = []
    df[col_seccion] = df[col_seccion].astype(str).str.strip().str.upper()
    df[col_seccion] = df[col_seccion].replace(MAPEO_SECCIONES)

    secciones_invalidas = df.loc[~df[col_seccion].isin(SECCIONES_VALIDAS)]

    if len(secciones_invalidas) > 0:
        for idx, row in secciones_invalidas.iterrows():
            errores.append(f"Fila {idx + 2}: Secci√≥n inv√°lida '{row[col_seccion]}' (debe ser A-G o U)")
    
    return errores

def validar_neurodiversidad(df, col_neuro="NEURODIVERSIDAD (S√ç/NO)"):
    """
    Valida que neurodiversidad sea S√≠ o No.
    Retorna lista de errores.
    """
    errores = []
    df[col_neuro] = df[col_neuro].astype(str).str.strip().str.upper()
    
    mapeo_neuro = {
        "SI": "S√ç", "S": "S√ç", "YES": "S√ç", "Y": "S√ç",
        "N": "NO", "NOT": "NO"
    }
    df[col_neuro] = df[col_neuro].replace(mapeo_neuro)
    neuros_invalidas = df.loc[~df[col_neuro].isin(["S√ç", "NO"])]

    if len(neuros_invalidas) > 0:
        for idx, row in neuros_invalidas.iterrows():
            identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
            errores.append(f"Fila {idx + 2}: Neurodiversidad inv√°lida '{row[col_neuro]}' - {identificador}")
    
    return errores

def validar_fecha_nacimiento(df, col_fecha="NACIMIENTO (DD/MM/YYYY)"):
    """
    Valida y formatea fechas al formato DD/MM/YYYY.
    Retorna lista de errores y modifica el DataFrame.
    """
    errores = []
    
    for idx, row in df.iterrows():
        fecha_original = str(row[col_fecha]).strip()
        identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
        fecha_parseada = pd.to_datetime(fecha_original, errors="coerce", dayfirst=True)
        
        if pd.isna(fecha_parseada):
            errores.append(f"Fila {idx + 2}: Fecha inv√°lida '{fecha_original}' - {identificador}")
        else:
            fecha_formateada = fecha_parseada.strftime("%d/%m/%Y")
            df.at[idx, col_fecha] = fecha_formateada
    
    return errores

def validar_dni(df, col_dni="DNI"):
    """
    Valida que el DNI tenga exactamente 8 d√≠gitos.
    Retorna lista de errores.
    """
    errores = []
    df[col_dni] = df[col_dni].astype(str).str.strip()
    
    for idx, row in df.iterrows():
        dni = row[col_dni]
        identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
        
        if not (dni.isdigit() and len(dni) == 8):
            errores.append(f"Fila {idx + 2}: DNI inv√°lido '{dni}' (debe ser 8 d√≠gitos) - {identificador}")
    
    return errores

def validar_correo(df, col_correo="CORREO INSTITUCIONAL"):
    """
    Valida formato b√°sico de correo electr√≥nico.
    Retorna lista de errores.
    """
    errores = []
    
    for idx, row in df.iterrows():
        correo = str(row[col_correo]).strip().lower()
        identificador = crear_identificador(df.loc[[idx]], "PATERNO", "MATERNO", "NOMBRES").iloc[0]
        
        if "@" not in correo or "." not in correo.split("@")[-1]:
            errores.append(f"Fila {idx + 2}: Correo inv√°lido '{correo}' - {identificador}, no contiene @ ni .")
    
    return errores

def mostrar_stepper(paso_actual):
    """Muestra el indicador de progreso visual"""
    pasos = [
        {"num": 0, "titulo": "Nombre del Colegio", "icono": "üè´"},
        {"num": 1, "titulo": "Archivo 1: N√≥mina", "icono": "üìã"},
        {"num": 2, "titulo": "Archivo 2: Notas", "icono": "üìä"},
        {"num": 3, "titulo": "Descarga Final", "icono": "‚¨áÔ∏è"}
    ]
    
    cols = st.columns(len(pasos))
    for i, paso in enumerate(pasos):
        with cols[i]:
            if paso["num"] < paso_actual:
                st.markdown(f"### ‚úÖ {paso['icono']}")
                st.markdown(f"**{paso['titulo']}**")
                st.markdown("*Completado*")
            elif paso["num"] == paso_actual:
                st.markdown(f"### üîµ {paso['icono']}")
                st.markdown(f"**{paso['titulo']}**")
                st.markdown("*En progreso*")
            else:
                st.markdown(f"### ‚ö™ {paso['icono']}")
                st.markdown(f"<span style='color: gray;'>{paso['titulo']}</span>", unsafe_allow_html=True)
                st.markdown("*Pendiente*")
    
    st.divider()

def crear_archivo_evaluador(df_archivo1, df_archivo2_procesado):
    """
    Crea el archivo evaluador haciendo un full join entre archivo1 y archivo2
    usando IDENTIFICADOR como clave √∫nica. Retorna dos DataFrames separados
    por grado: uno para 1P-3P y otro para 4P-5S.
    
    Args:
        df_archivo1: DataFrame del archivo 1 con todos los alumnos
        df_archivo2_procesado: DataFrame del archivo 2 con notas (puede ser 1P-3P o 4P-5S)
    
    Returns:
        tuple: (df_1p3p, df_4p5s) - DataFrames separados por grado
    """
    df1_base = df_archivo1[[
        "IDENTIFICADOR", 
        "PATERNO", 
        "MATERNO", 
        "NOMBRES", 
        "GRADO", 
        "SECCI√ìN"
    ]].copy()
    
    df2_merge = df_archivo2_procesado.copy()
    df2_merge['_origen'] = 'archivo2'
    df1_base['_origen'] = 'archivo1'
    df_evaluador = pd.merge(
        df2_merge,
        df1_base,
        on="IDENTIFICADOR",
        how="outer",
        suffixes=("", "_archivo1"),
        indicator=True
    )
    
    def asignar_observacion(row):
        if row['_merge'] == 'both':
            return ''
        elif row['_merge'] == 'right_only':
            return 'SN'
        else:  # 'left_only'
            return 'RET'
    
    df_evaluador['OBSERVADOS'] = df_evaluador.apply(asignar_observacion, axis=1)
    df_evaluador = df_evaluador.drop(columns=['_merge', '_origen'], errors='ignore')
    columnas_comunes = ["PATERNO", "MATERNO", "NOMBRES", "GRADO", "SECCI√ìN"]
    
    if "GRADO_archivo1" in df_evaluador.columns:
        df_evaluador["GRADO"] = df_evaluador["GRADO"].fillna(df_evaluador["GRADO_archivo1"])
        mask_vacio = (df_evaluador["GRADO"] == "") | (df_evaluador["GRADO"].isna())
        df_evaluador.loc[mask_vacio, "GRADO"] = df_evaluador.loc[mask_vacio, "GRADO_archivo1"]
    
    df_evaluador = df_evaluador.drop(columns=["GRADO_archivo1"], errors='ignore')
    
    columnas_comunes_restantes = ["PATERNO", "MATERNO", "NOMBRES", "SECCI√ìN"]

    for col in columnas_comunes_restantes:
        col_archivo1 = f"{col}_archivo1"
        if col_archivo1 in df_evaluador.columns:
            df_evaluador[col] = df_evaluador[col].fillna(df_evaluador[col_archivo1])
            mask_vacio = (df_evaluador[col] == "") | (df_evaluador[col].isna())
            df_evaluador.loc[mask_vacio, col] = df_evaluador.loc[mask_vacio, col_archivo1]
            df_evaluador = df_evaluador.drop(columns=[col_archivo1])
    
    if "CURSO" not in df_evaluador.columns:
        df_evaluador["CURSO"] = ""

    grados_presentes = df_evaluador["GRADO"].unique()
    es_1p3p = any(g in ["1P", "2P", "3P"] for g in grados_presentes if pd.notna(g))
    es_4p5s = any(g in ["4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"] for g in grados_presentes if pd.notna(g))

    if es_1p3p and "NOTA VIGESIMAL 100%" not in df_evaluador.columns:
        df_evaluador["NOTA VIGESIMAL 100%"] = ""
    if es_4p5s and "NOTA VIGESIMAL 25%" not in df_evaluador.columns:
        df_evaluador["NOTA VIGESIMAL 25%"] = ""

    df_evaluador = df_evaluador.fillna("")
    
    grados_1p3p = ["1P", "2P", "3P"]
    grados_4p5s = ["4P", "5P", "6P", "1S", "2S", "3S", "4S", "5S"]
    
    df_1p3p = df_evaluador[df_evaluador["GRADO"].isin(grados_1p3p)].copy()
    df_4p5s = df_evaluador[df_evaluador["GRADO"].isin(grados_4p5s)].copy()
    
    columnas_1p3p = [
        "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", 
        "GRADO", "SECCI√ìN", "NOTA VIGESIMAL 100%", "IDENTIFICADOR", "OBSERVADOS"
    ]
    
    columnas_4p5s = [
        "NRO.", "PATERNO", "MATERNO", "NOMBRES", "CURSO", 
        "GRADO", "SECCI√ìN", "NOTA VIGESIMAL 25%", 
        "NOTAS VIGESIMALES 75%", "PROMEDIO", "IDENTIFICADOR", "OBSERVADOS"
    ]
    
    for col in columnas_1p3p:
        if col not in df_1p3p.columns:
            df_1p3p[col] = ""
    df_1p3p = df_1p3p[columnas_1p3p]
    
    for col in columnas_4p5s:
        if col not in df_4p5s.columns:
            df_4p5s[col] = ""
    df_4p5s = df_4p5s[columnas_4p5s]
    
    if len(df_1p3p) > 0:
        df_1p3p["NRO."] = range(1, len(df_1p3p) + 1)
    
    if len(df_4p5s) > 0:
        df_4p5s["NRO."] = range(1, len(df_4p5s) + 1)
    
    return df_1p3p, df_4p5s

def guardar_con_formato_original(df_procesado, archivo_original_bytes, nombre_hoja, fila_cabecera, agregar_columnas_nuevas=False, solo_hoja_especificada=False):
    """
    Preserva el formato del archivo original y actualiza solo los datos procesados
    
    Args:
        df_procesado: DataFrame con los datos procesados
        archivo_original_bytes: Bytes del archivo Excel original
        nombre_hoja: Nombre de la hoja a actualizar (None para usar la primera hoja)
        fila_cabecera: √çndice de la fila donde est√° la cabecera (base 0 de pandas)
        agregar_columnas_nuevas: Si True, agrega columnas nuevas del df_procesado a la cabecera
    
    Returns:
        BytesIO con el archivo actualizado preservando formato
    """

    if 'PATERNO' in df_procesado.columns:
        columnas_orden = ['PATERNO']
        if 'MATERNO' in df_procesado.columns:
            columnas_orden.append('MATERNO')
        if 'NOMBRES' in df_procesado.columns or 'NOMBRE' in df_procesado.columns:
            columnas_orden.append('NOMBRES' if 'NOMBRES' in df_procesado.columns else 'NOMBRE')
        df_procesado = df_procesado.sort_values(columnas_orden).reset_index(drop=True)

    # Renumerar columna NRO. en orden ascendente si existe
    if 'NRO.' in df_procesado.columns:
        df_procesado['NRO.'] = range(1, len(df_procesado) + 1)

    wb = load_workbook(BytesIO(archivo_original_bytes))
    
    if nombre_hoja is None or nombre_hoja not in wb.sheetnames:
        ws = wb.active
    else:
        ws = wb[nombre_hoja]

    if solo_hoja_especificada:
      hoja_a_mantener = ws.title
      hojas_a_eliminar = [sheet for sheet in wb.sheetnames if sheet != hoja_a_mantener]
      for hoja in hojas_a_eliminar:
          wb.remove(wb[hoja])
    
    fila_cabecera_excel = fila_cabecera + 1
    fila_inicio_datos = fila_cabecera_excel + 1
    
    if agregar_columnas_nuevas:
        
        cabecera_actual = []
        ultima_col_con_datos = 0
        for idx, cell in enumerate(ws[fila_cabecera_excel], start=1):
            if cell.value is not None:
                cabecera_actual.append(str(cell.value).upper().strip())
                ultima_col_con_datos = idx
        
        cabecera_df = [str(col).upper().strip() for col in df_procesado.columns]
        columnas_nuevas = [col for col in cabecera_df if col not in cabecera_actual]
        
        if columnas_nuevas:
            celda_referencia = ws.cell(row=fila_cabecera_excel, column=ultima_col_con_datos)
            
            for idx, nueva_col in enumerate(columnas_nuevas, start=1):
                nueva_celda = ws.cell(row=fila_cabecera_excel, column=ultima_col_con_datos + idx)
                nueva_celda.value = nueva_col
                
                if celda_referencia.fill:
                    nueva_celda.fill = PatternFill(
                        start_color=celda_referencia.fill.start_color,
                        end_color=celda_referencia.fill.end_color,
                        fill_type=celda_referencia.fill.fill_type
                    )
                if celda_referencia.font:
                    nueva_celda.font = Font(
                        name=celda_referencia.font.name,
                        size=celda_referencia.font.size,
                        bold=celda_referencia.font.bold,
                        italic=celda_referencia.font.italic,
                        color=celda_referencia.font.color
                    )
                if celda_referencia.alignment:
                    nueva_celda.alignment = Alignment(
                        horizontal=celda_referencia.alignment.horizontal,
                        vertical=celda_referencia.alignment.vertical
                    )

    for col_idx, col_name in enumerate(df_procesado.columns, start=1):
        celda_cabecera = ws.cell(row=fila_cabecera_excel, column=col_idx)
        celda_cabecera.value = str(col_name).upper()

    if ws.max_row >= fila_inicio_datos:
        ws.delete_rows(fila_inicio_datos, ws.max_row - fila_inicio_datos + 1)
    
    for r_idx, row in enumerate(dataframe_to_rows(df_procesado, index=False, header=False), start=fila_inicio_datos):
        for c_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def guardar_evaluador_con_multiples_hojas(archivo_original_bytes, dict_hojas_procesadas, solo_hojas_especificadas=False):
    """
    Guarda un archivo Excel con m√∫ltiples hojas preservando el formato original.
    
    Args:
        archivo_original_bytes: Bytes del archivo Excel original
        dict_hojas_procesadas: Diccionario con formato {
            'nombre_hoja': {
                'df': DataFrame procesado,
                'fila_cabecera': int (√≠ndice de cabecera en base 0)
            }
        }
        solo_hojas_especificadas: Si True, solo incluye las hojas en dict_hojas_procesadas
                                   Si False, mantiene todas las hojas del archivo original
    
    Returns:
        BytesIO con el archivo actualizado preservando formato
    """
    wb = load_workbook(BytesIO(archivo_original_bytes))
    
    if solo_hojas_especificadas:
        hojas_a_mantener = list(dict_hojas_procesadas.keys())
        hojas_a_eliminar = [sheet for sheet in wb.sheetnames if sheet not in hojas_a_mantener]
        for hoja in hojas_a_eliminar:
            wb.remove(wb[hoja])
    
    for nombre_hoja, datos in dict_hojas_procesadas.items():
        df_procesado = datos['df']
        if 'PATERNO' in df_procesado.columns:
            columnas_orden = ['PATERNO']
            if 'MATERNO' in df_procesado.columns:
                columnas_orden.append('MATERNO')
            if 'NOMBRES' in df_procesado.columns or 'NOMBRE' in df_procesado.columns:
                columnas_orden.append('NOMBRES' if 'NOMBRES' in df_procesado.columns else 'NOMBRE')
            df_procesado = df_procesado.sort_values(columnas_orden).reset_index(drop=True)

        # Renumerar columna NRO. en orden ascendente si existe
        if 'NRO.' in df_procesado.columns:
            df_procesado['NRO.'] = range(1, len(df_procesado) + 1)

        fila_cabecera = datos['fila_cabecera']
        
        if nombre_hoja not in wb.sheetnames:
            if len([k for k in dict_hojas_procesadas.keys()]) == 1:
                ws = wb.active
                ws.title = nombre_hoja
            else:
                ws = wb.create_sheet(title=nombre_hoja)
                fila_cabecera = 0
        else:
            ws = wb[nombre_hoja]
        
        fila_cabecera_excel = fila_cabecera + 1
        fila_inicio_datos = fila_cabecera_excel + 1
        
        cabecera_actual = []
        ultima_col_con_datos = 0
        for idx, cell in enumerate(ws[fila_cabecera_excel], start=1):
            if cell.value is not None:
                cabecera_actual.append(str(cell.value).upper().strip())
                ultima_col_con_datos = idx
        
        cabecera_df = [str(col).upper().strip() for col in df_procesado.columns]
        columnas_nuevas = [col for col in cabecera_df if col not in cabecera_actual]
        
        if columnas_nuevas:
            celda_referencia = ws.cell(row=fila_cabecera_excel, column=max(1, ultima_col_con_datos))
            
            for idx, nueva_col in enumerate(columnas_nuevas, start=1):
                nueva_celda = ws.cell(row=fila_cabecera_excel, column=ultima_col_con_datos + idx)
                nueva_celda.value = nueva_col
                
                if celda_referencia.fill:
                    nueva_celda.fill = PatternFill(
                        start_color=celda_referencia.fill.start_color,
                        end_color=celda_referencia.fill.end_color,
                        fill_type=celda_referencia.fill.fill_type
                    )
                if celda_referencia.font:
                    nueva_celda.font = Font(
                        name=celda_referencia.font.name,
                        size=celda_referencia.font.size,
                        bold=celda_referencia.font.bold,
                        italic=celda_referencia.font.italic,
                        color=celda_referencia.font.color
                    )
                if celda_referencia.alignment:
                    nueva_celda.alignment = Alignment(
                        horizontal=celda_referencia.alignment.horizontal,
                        vertical=celda_referencia.alignment.vertical
                    )

        for col_idx, col_name in enumerate(df_procesado.columns, start=1):
            celda_cabecera = ws.cell(row=fila_cabecera_excel, column=col_idx)
            celda_cabecera.value = str(col_name).upper()
        
        if ws.max_row >= fila_inicio_datos:
            ws.delete_rows(fila_inicio_datos, ws.max_row - fila_inicio_datos + 1)
        
        for r_idx, row in enumerate(dataframe_to_rows(df_procesado, index=False, header=False), start=fila_inicio_datos):
            for c_idx, value in enumerate(row, start=1):
                if pd.isna(value):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generar_reportes_pdf(df, nombre_colegio, tipo_archivo):
    """
    Genera reportes PDF agrupados por Grado ‚Üí Secci√≥n ‚Üí Curso
    Con encabezado personalizado: Logo izquierda, paginaci√≥n y fecha derecha
    
    Args:
        df: DataFrame con los datos homologados
        nombre_colegio: Nombre del colegio para el header
        tipo_archivo: '1P-3P' o '4P-5S'
    """

    def encabezado_pie_pagina(canvas, doc):
        """
        Dibuja el encabezado y pie de p√°gina en cada p√°gina
        """
        canvas.saveState()
        ancho, alto = A4
        
        canvas.setFont('Helvetica-Bold', 11)
        canvas.setFillColor(colors.HexColor('#1a5490'))
        
        x_logo = 15 * mm
        y_logo = alto - 12 * mm
        
        canvas.drawString(x_logo, y_logo, "Alianza")
        canvas.drawString(x_logo, y_logo - 4*mm, "Educativa")
        
        try:
            logo_cibertec_path = os.path.join("logos", "logo_cibertec.jpeg")
            
            if os.path.exists(logo_cibertec_path):
                logo_width = 25 * mm 
                logo_height = 10 * mm 
                
                x_logo_cibertec = ancho - 15*mm - logo_width
                y_logo_cibertec = alto - 12*mm - logo_height*0.5
                
                canvas.drawImage(
                    logo_cibertec_path,
                    x_logo_cibertec,
                    y_logo_cibertec,
                    width=logo_width,
                    height=logo_height,
                    preserveAspectRatio=True,
                    mask='auto'
                )
        except Exception as e:
            pass

        canvas.setStrokeColor(colors.HexColor('#1a5490'))
        canvas.setLineWidth(0.5)
        canvas.line(15*mm, alto - 18*mm, ancho - 15*mm, alto - 18*mm)
        
        canvas.restoreState()
    
    with st.spinner("üìù Generando reportes PDF..."):
        zip_buffer = BytesIO()
        
        with ZipFile(zip_buffer, 'w') as zip_file:
            grupos = df.groupby(['GRADO', 'SECCI√ìN', 'CURSO'])
            total_grupos = len(grupos)
            progress_bar = st.progress(0)
            
            for idx, ((grado, seccion, curso), grupo_df) in enumerate(grupos):
                pdf_buffer = BytesIO()
                
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=A4,
                    rightMargin=15*mm,
                    leftMargin=15*mm,
                    topMargin=22*mm,
                    bottomMargin=15*mm
                )
                
                styles = getSampleStyleSheet()
                style_title = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#1a5490'),
                    spaceAfter=3*mm,
                    alignment=TA_CENTER
                )
                style_subtitle = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Normal'],
                    fontSize=11,
                    textColor=colors.HexColor('#2c3e50'),
                    spaceAfter=2*mm,
                    alignment=TA_LEFT
                )
                style_normal = ParagraphStyle(
                    'CustomSubtitle',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#2c3e50'),
                    spaceAfter=2*mm,
                    alignment=TA_LEFT
                )
                
                story = []
                
                story.append(Paragraph("N√ìMINA DE ALUMNOS", style_title))
                story.append(Paragraph(f"<b>Colegio:</b> {nombre_colegio}", style_subtitle))
                story.append(Paragraph(f"<b>Ciclo:</b> {tipo_archivo}", style_subtitle))
                story.append(Paragraph(f"<b>Grado:</b> {grado} | <b>Secci√≥n:</b> {seccion}", style_subtitle))
                story.append(Paragraph(f"<b>Curso:</b> {curso}", style_subtitle))
                story.append(Spacer(1, 5*mm))
                
                grupo_df_sorted = grupo_df.sort_values(['PATERNO', 'MATERNO', 'NOMBRES'])
                
                total_alumnos = len(grupo_df_sorted)
                aprobados = len(grupo_df_sorted[pd.to_numeric(grupo_df_sorted['NOTA FINAL'], errors='coerce') >= 12.5])
                desaprobados = total_alumnos - aprobados
                excelencia = len(grupo_df_sorted[pd.to_numeric(grupo_df_sorted['NOTA FINAL'], errors='coerce') == 20])
                promedio = pd.to_numeric(grupo_df_sorted["NOTA FINAL"], errors="coerce").mean()
                promedio = round(promedio, 2)
                
                datos_tabla = [['Nro.', 'Nombres', 'Apellido Paterno', 'Apellido Materno', 'Nota']]
                
                for i, (_, row) in enumerate(grupo_df_sorted.iterrows(), 1):
                    nota_original = row['NOTA FINAL']
                    try:
                        nota_float = float(nota_original)
                        nota_redondeada = int(round(nota_float))
                    except (ValueError, TypeError):
                        nota_redondeada = nota_original
                    
                    datos_tabla.append([
                        str(i),
                        str(row['NOMBRES']),
                        str(row['PATERNO']),
                        str(row['MATERNO']),
                        str(nota_redondeada)
                    ])
                
                tabla = Table(datos_tabla, colWidths=[15*mm, 50*mm, 40*mm, 40*mm, 20*mm])
                
                tabla.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    
                    # Datos
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 1), (3, -1), 'LEFT'),
                    ('ALIGN', (4, 1), (4, -1), 'CENTER'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    
                    # Bordes
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                story.append(tabla)
                story.append(Spacer(1, 8*mm))
                
                # Footer
                total_alumnos = len(grupo_df_sorted)
                aprobados = len(grupo_df_sorted[pd.to_numeric(grupo_df_sorted['NOTA FINAL'], errors='coerce') >= 12.5])
                desaprobados = total_alumnos - aprobados
                excelencia = len(grupo_df_sorted[pd.to_numeric(grupo_df_sorted['NOTA FINAL'], errors='coerce') == 20])
                promedio = pd.to_numeric(grupo_df_sorted["NOTA FINAL"], errors="coerce").mean()
                promedio = round(promedio, 2)

                story.append(Paragraph(f"<b>Resultados:</b>", style_subtitle))
                story.append(Paragraph(f"<b>Total de alumnos:</b> {total_alumnos}", style_normal))
                story.append(Paragraph(f"<b>Excelencia (nota 20):</b> {excelencia}", style_normal))
                story.append(Paragraph(f"<b>Promedio del Aula:</b> {promedio}", style_normal))
                story.append(Paragraph(f"<b>Aprobados:</b> {aprobados} | <b>Desaprobados:</b> {desaprobados}", style_normal))
                
                doc.build(story, onFirstPage=encabezado_pie_pagina, onLaterPages=encabezado_pie_pagina)
                
                pdf_buffer.seek(0)
                pdf_reader = PyPDF2.PdfReader(pdf_buffer)
                total_paginas = len(pdf_reader.pages)
                
                pdf_final_buffer = BytesIO()
                pdf_writer = PyPDF2.PdfWriter()
                
                for numero_pagina in range(total_paginas):
                    pagina = pdf_reader.pages[numero_pagina]
                    
                    overlay_buffer = BytesIO()
                    overlay_canvas = canvas.Canvas(overlay_buffer, pagesize=A4)
                    
                    ancho, alto = A4
                    margen_izq = 15 * mm
                    margen_der = 15 * mm
                    y_pie = 12 * mm
                    
                    overlay_canvas.setStrokeColor(colors.HexColor('#1a5490'))
                    overlay_canvas.setLineWidth(0.5)
                    overlay_canvas.line(margen_izq, y_pie + 4*mm, ancho - margen_der, y_pie + 4*mm)
                    
                    overlay_canvas.setFont('Helvetica', 9)
                    overlay_canvas.setFillColor(colors.black)
                    
                    fecha_actual = datetime.now().strftime("%d/%m/%Y")
                    texto_fecha = f"Impreso: {fecha_actual}"
                    overlay_canvas.drawString(margen_izq, y_pie, texto_fecha)
                    
                    texto_pagina = f"P√°gina {numero_pagina + 1}/{total_paginas}"
                    overlay_canvas.drawRightString(ancho - margen_der, y_pie, texto_pagina)
                    
                    overlay_canvas.save()
                    
                    overlay_buffer.seek(0)
                    overlay_pdf = PyPDF2.PdfReader(overlay_buffer)
                    overlay_page = overlay_pdf.pages[0]
                    
                    pagina.merge_page(overlay_page)
                    pdf_writer.add_page(pagina)
                
                pdf_writer.write(pdf_final_buffer)
                pdf_final_buffer.seek(0)
                
                pdf_bytes = pdf_final_buffer.getvalue()
                nombre_archivo = f"{grado}_{seccion}_{curso.replace('/', '-')}.pdf"
                zip_file.writestr(nombre_archivo, pdf_bytes)
                
                progress_bar.progress((idx + 1) / total_grupos)
            
            progress_bar.empty()
        
        zip_buffer.seek(0)
        
        st.success(f"üéâ {total_grupos} reportes PDF generados correctamente")
        
        st.download_button(
            label="üì• Descargar Reportes (ZIP)",
            data=zip_buffer,
            file_name=f"Resultados_{nombre_colegio}_{tipo_archivo}.zip",
            mime="application/zip",
            use_container_width=True
        )

def validar_notas_numericas(df):
    """
    Valida que las columnas de notas cumplan con los requisitos:
    - No pueden tener valores negativos
    - Deben ser n√∫meros (enteros o decimales)
    - M√°ximo valor permitido: 20
    
    Args:
        df: DataFrame con las columnas de notas
    
    Returns:
        tuple: (bool, list) - (es_valido, lista_de_errores)
    """
    errores = []
    columnas_notas = ["NOTA VIGESIMAL 25%", "NOTAS VIGESIMALES 75%", "PROMEDIO"]
    
    for col in columnas_notas:
        if col not in df.columns:
            continue
        
        for idx, valor in df[col].items():
            valor_str = str(valor).strip().upper()
            
            if valor_str in ["", "NAN", "NONE", "NP"]:
                continue
            
            try:
                valor_num = float(valor_str)
                
                if valor_num < 0:
                    errores.append({
                        "fila": idx + 2,
                        "columna": col,
                        "valor": valor_str,
                        "error": "Valor negativo no permitido",
                        "paterno": str(df.loc[idx, "PATERNO"]) if "PATERNO" in df.columns else "",
                        "materno": str(df.loc[idx, "MATERNO"]) if "MATERNO" in df.columns else "",
                        "nombres": str(df.loc[idx, "NOMBRES"]) if "NOMBRES" in df.columns else ""
                    })
                
                elif valor_num > 20:
                    errores.append({
                        "fila": idx + 2,
                        "columna": col,
                        "valor": valor_str,
                        "error": "Valor mayor a 20 no permitido",
                        "paterno": str(df.loc[idx, "PATERNO"]) if "PATERNO" in df.columns else "",
                        "materno": str(df.loc[idx, "MATERNO"]) if "MATERNO" in df.columns else "",
                        "nombres": str(df.loc[idx, "NOMBRES"]) if "NOMBRES" in df.columns else ""
                    })
                    
            except ValueError:
                errores.append({
                    "fila": idx + 2,
                    "columna": col,
                    "valor": valor_str,
                    "error": "Valor no num√©rico",
                    "paterno": str(df.loc[idx, "PATERNO"]) if "PATERNO" in df.columns else "",
                    "materno": str(df.loc[idx, "MATERNO"]) if "MATERNO" in df.columns else "",
                    "nombres": str(df.loc[idx, "NOMBRES"]) if "NOMBRES" in df.columns else ""
                })
    
    es_valido = len(errores) == 0
    return es_valido, errores

def leer_archivo_evaluador(archivo_bytes, nombre_hoja=None):
    """
    Lee un archivo evaluador Excel y retorna DataFrame y metadatos.
    Incluye validaci√≥n de notas num√©ricas.
    
    Args:
        archivo_bytes: Bytes del archivo Excel
        nombre_hoja: Nombre de la hoja a leer (opcional)
    
    Returns:
        tuple: (df, error, fila_cabecera, hojas, df_errores)
            - df: DataFrame con los datos (None si hay error)
            - error: Mensaje de error de lectura (None si no hay error)
            - fila_cabecera: √çndice de la fila de cabecera (None si hay error)
            - hojas: Lista de nombres de hojas (None si hay error)
            - df_errores: DataFrame con errores de validaci√≥n (None si no hay errores de validaci√≥n)
    """
    try:
        wb = load_workbook(BytesIO(archivo_bytes), data_only=True)
        
        if nombre_hoja is None:
            nombre_hoja = wb.sheetnames[0]
        
        if nombre_hoja not in wb.sheetnames:
            return None, f"La hoja '{nombre_hoja}' no existe en el archivo", None, None, None
        
        ws = wb[nombre_hoja]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        df = pd.DataFrame(data)
        
        fila_cabecera = detectar_cabecera_automatica(df, COLUMNAS_EVALUADOR)
        
        if fila_cabecera is None:
            return None, "No se pudo detectar la cabecera autom√°ticamente", None, None, None
        
        nombres_columnas_raw = df.iloc[fila_cabecera].tolist()
        
        nombres_columnas = []
        for i, col in enumerate(nombres_columnas_raw):
            if col is None or pd.isna(col) or str(col).strip() == '' or str(col).lower() == 'nan':
                nombres_columnas.append(f"Columna_Extra_{i}")
            else:
                nombres_columnas.append(str(col).strip())
        
        df.columns = nombres_columnas
        df = df.iloc[fila_cabecera + 1:].reset_index(drop=True)
        
        columnas_a_mantener = [col for col in df.columns if col in COLUMNAS_EVALUADOR]
        
        columnas_faltantes = [col for col in COLUMNAS_EVALUADOR if col not in columnas_a_mantener]
        if columnas_faltantes:
            return None, f"No se encontraron las columnas: {', '.join(columnas_faltantes)}. Revisa que los nombres coincidan exactamente.", None, None, None
        
        df = df[columnas_a_mantener]
        
        df = df.dropna(how='all')
        
        es_valido, errores_validacion = validar_notas_numericas(df)
        if not es_valido:
            df_errores = pd.DataFrame(errores_validacion)
            
            df_errores['nombre_completo'] = df_errores['paterno'] + ' ' + df_errores['materno'] + ', ' + df_errores['nombres']
            
            df_errores_display = df_errores[['fila', 'nombre_completo', 'columna', 'valor', 'error']].copy()
            df_errores_display.columns = ['FILA', 'NOMBRE COMPLETO', 'COLUMNA', 'VALOR', 'TIPO DE ERROR']
            
            return None, None, None, None, df_errores_display
        
        return df, None, fila_cabecera, wb.sheetnames, None
        
    except Exception as e:
        return None, f"Error al leer archivo: {str(e)}", None, None, None

def comparar_evaluadores(df_base, df_revisar):
    """
    Compara dos archivos evaluadores.
    - Ambos deben tener las mismas columnas en el mismo orden
    - Todo debe ser id√©ntico EXCEPTO la columna "NOTAS VIGESIMALES 75%"
    - En el archivo BASE: pueden estar vac√≠as "NOTA VIGESIMAL 25%", "NOTAS VIGESIMALES 75%", "PROMEDIO", "OBSERVADOS"
    - En el archivo A REVISAR: "NOTA VIGESIMAL 25%" y "NOTAS VIGESIMALES 75%" deben estar completas
    """
    errores = []
    
    # Normalizar nombres de columnas (eliminar espacios extras pero mantener capitalizaci√≥n)
    df_base.columns = [str(col).strip() for col in df_base.columns]
    df_revisar.columns = [str(col).strip() for col in df_revisar.columns]
    
    columnas_base = list(df_base.columns)
    columnas_revisar = list(df_revisar.columns)
    
    # 1. Verificar que ambos tienen las columnas requeridas
    columnas_faltantes_base = [col for col in COLUMNAS_EVALUADOR if col not in columnas_base]
    columnas_faltantes_revisar = [col for col in COLUMNAS_EVALUADOR if col not in columnas_revisar]
    
    if columnas_faltantes_base:
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": f"Archivo BASE: Faltan columnas requeridas: {', '.join(columnas_faltantes_base)}",
            "archivo": "BASE",
            "fila": None,
            "columna": None,
            "valor_base": None,
            "valor_revisar": None,
            "detalle": f"Columnas actuales: {columnas_base}"
        })
    
    if columnas_faltantes_revisar:
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": f"Archivo A REVISAR: Faltan columnas requeridas: {', '.join(columnas_faltantes_revisar)}",
            "archivo": "A REVISAR",
            "fila": None,
            "columna": None,
            "valor_base": None,
            "valor_revisar": None,
            "detalle": f"Columnas actuales: {columnas_revisar}"
        })
    
    # Si faltan columnas, retornar ahora
    if columnas_faltantes_base or columnas_faltantes_revisar:
        return errores
    
    # 2. Verificar que las columnas coincidan exactamente en orden
    if columnas_base != columnas_revisar:
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": "Las columnas no coinciden entre archivos (orden diferente)",
            "archivo": "Ambos",
            "fila": None,
            "columna": None,
            "valor_base": None,
            "valor_revisar": None,
            "detalle": f"BASE: {columnas_base}\nREVISAR: {columnas_revisar}"
        })
        return errores
    
    # 3. Verificar que el n√∫mero de filas sea el mismo
    if len(df_base) != len(df_revisar):
        errores.append({
            "tipo": "error_estructura",
            "categoria": "ESTRUCTURA",
            "descripcion": f"Diferente n√∫mero de filas",
            "archivo": "Ambos",
            "fila": None,
            "columna": None,
            "valor_base": f"{len(df_base)} filas",
            "valor_revisar": f"{len(df_revisar)} filas",
            "detalle": None
        })
    
    # 4. Comparar todas las columnas EXCEPTO "NOTAS VIGESIMALES 75%"
    columnas_comparar = [col for col in columnas_base if col not in ["NOTAS VIGESIMALES 75%"]]
    
    for col in columnas_comparar:
        # Para estas columnas opcionales, no comparar si est√°n vac√≠as en BASE
        columnas_opcionales_base = ["NOTA VIGESIMAL 25%", "PROMEDIO", "OBSERVADOS"]
        
        for idx in range(min(len(df_base), len(df_revisar))):
            val_base = str(df_base.loc[idx, col]).strip().upper()
            val_revisar = str(df_revisar.loc[idx, col]).strip().upper()
            
            # Normalizar valores vac√≠os
            valores_vacios = ["", "NAN", "NONE", "NP"]
            if val_base in valores_vacios:
                val_base = ""
            if val_revisar in valores_vacios:
                val_revisar = ""
            
            # Si la columna es opcional en BASE y est√° vac√≠a en BASE, no comparar
            if col in columnas_opcionales_base and val_base == "":
                continue
            
            if val_base != val_revisar:
                # Obtener nombres y apellidos para el reporte
                paterno = str(df_base.loc[idx, "PATERNO"]).strip() if "PATERNO" in df_base.columns else ""
                materno = str(df_base.loc[idx, "MATERNO"]).strip() if "MATERNO" in df_base.columns else ""
                nombres = str(df_base.loc[idx, "NOMBRES"]).strip() if "NOMBRES" in df_base.columns else ""
                
                errores.append({
                    "tipo": "diferencia_contenido",
                    "categoria": "CONTENIDO DIFERENTE",
                    "paterno": paterno,
                    "materno": materno,
                    "nombres": nombres,
                    "descripcion": f"Valor diferente en columna '{col}'",
                    "archivo": "Ambos",
                    "fila": idx + 2,  # +2 por cabecera y porque index empieza en 0
                    "columna": col,
                    "valor_base": str(df_base.loc[idx, col]) if str(df_base.loc[idx, col]).strip() not in ["", "nan", "None"] else "(vac√≠o)",
                    "valor_revisar": str(df_revisar.loc[idx, col]) if str(df_revisar.loc[idx, col]).strip() not in ["", "nan", "None"] else "(vac√≠o)",
                    "detalle": None
                })
    
    # 5. Verificar columna "NOTAS VIGESIMALES 75%" en archivo A REVISAR
    if "NOTAS VIGESIMALES 75%" in columnas_revisar:
        for idx in range(len(df_revisar)):
            val_revisar = str(df_revisar.loc[idx, "NOTAS VIGESIMALES 75%"]).strip().upper()
            
            if val_revisar in ["", "NAN", "NONE", "NP"]:
                # Obtener nombres y apellidos para el reporte
                paterno = str(df_revisar.loc[idx, "PATERNO"]).strip() if "PATERNO" in df_revisar.columns else ""
                materno = str(df_revisar.loc[idx, "MATERNO"]).strip() if "MATERNO" in df_revisar.columns else ""
                nombres = str(df_revisar.loc[idx, "NOMBRES"]).strip() if "NOMBRES" in df_revisar.columns else ""
                
                errores.append({
                    "tipo": "campo_vacio_revisar",
                    "categoria": "CAMPO VAC√çO EN REVISAR",
                    "paterno": paterno,
                    "materno": materno,
                    "nombres": nombres,
                    "descripcion": "Campo 'NOTAS VIGESIMALES 75%' vac√≠o o con 'NP'",
                    "archivo": "A REVISAR",
                    "fila": idx + 2,
                    "columna": "NOTAS VIGESIMALES 75%",
                    "valor_base": str(df_base.loc[idx, "NOTAS VIGESIMALES 75%"]) if idx < len(df_base) else "N/A",
                    "valor_revisar": "(vac√≠o)",
                    "detalle": None
                })
    
    # 6. Verificar que "NOTA VIGESIMAL 25%" est√© completa en archivo A REVISAR
    if "NOTA VIGESIMAL 25%" in columnas_revisar:
        for idx in range(len(df_revisar)):
            val_revisar = str(df_revisar.loc[idx, "NOTA VIGESIMAL 25%"]).strip().upper()
            
            if val_revisar in ["", "NAN", "NONE", "NP"]:
                # Obtener nombres y apellidos para el reporte
                paterno = str(df_revisar.loc[idx, "PATERNO"]).strip() if "PATERNO" in df_revisar.columns else ""
                materno = str(df_revisar.loc[idx, "MATERNO"]).strip() if "MATERNO" in df_revisar.columns else ""
                nombres = str(df_revisar.loc[idx, "NOMBRES"]).strip() if "NOMBRES" in df_revisar.columns else ""
                
                errores.append({
                    "tipo": "campo_vacio_revisar",
                    "categoria": "CAMPO VAC√çO EN REVISAR",
                    "paterno": paterno,
                    "materno": materno,
                    "nombres": nombres,
                    "descripcion": "Campo 'NOTA VIGESIMAL 25%' vac√≠o o con 'NP'",
                    "archivo": "A REVISAR",
                    "fila": idx + 2,
                    "columna": "NOTA VIGESIMAL 25%",
                    "valor_base": str(df_base.loc[idx, "NOTA VIGESIMAL 25%"]) if idx < len(df_base) else "N/A",
                    "valor_revisar": "(vac√≠o)",
                    "detalle": None
                })
    
    return errores

def agregar_columna_nro(df):
    """Agrega columna Nro. al DataFrame si no existe, o la recalcula si existe"""
    if "Nro." in df.columns:
        df = df.drop(columns=["Nro."])
    df.insert(0, "Nro.", range(1, len(df) + 1))
    return df

def guardar_certificado_con_encabezado(archivo_original_bytes, dict_hojas_procesadas):
    """
    Guarda archivo de certificado preservando las primeras 7 filas del formato institucional
    y agregando nuestra cabecera personalizada en la fila 8.
    
    Args:
        archivo_original_bytes: Bytes del archivo Excel original
        dict_hojas_procesadas: Diccionario con formato igual a guardar_evaluador_con_multiples_hojas
    
    Returns:
        BytesIO con el archivo de certificado
    """
    
    wb_original = load_workbook(BytesIO(archivo_original_bytes))
    
    wb_nuevo = Workbook()
    wb_nuevo.remove(wb_nuevo.active)
    
    for nombre_hoja, datos in dict_hojas_procesadas.items():
        df_procesado = datos['df']

        if 'PATERNO' in df_procesado.columns:
            columnas_orden = ['PATERNO']
            if 'MATERNO' in df_procesado.columns:
                columnas_orden.append('MATERNO')
            if 'NOMBRES' in df_procesado.columns or 'NOMBRE' in df_procesado.columns:
                columnas_orden.append('NOMBRES' if 'NOMBRES' in df_procesado.columns else 'NOMBRE')
            df_procesado = df_procesado.sort_values(columnas_orden).reset_index(drop=True)

        # Renumerar columna NRO. en orden ascendente si existe
        if 'NRO.' in df_procesado.columns:
            df_procesado['NRO.'] = range(1, len(df_procesado) + 1)

        ws_nueva = wb_nuevo.create_sheet(title=nombre_hoja)

        celda_estilo_referencia = None
        
        if nombre_hoja in wb_original.sheetnames:
            ws_original = wb_original[nombre_hoja]
            
            for fila_idx in range(1, 8):
                for col_idx in range(1, ws_original.max_column + 1):
                    celda_original = ws_original.cell(row=fila_idx, column=col_idx)
                    celda_nueva = ws_nueva.cell(row=fila_idx, column=col_idx)
                    
                    if fila_idx == 5 and col_idx == 1 and celda_estilo_referencia is None:
                        celda_estilo_referencia = celda_original

                    celda_nueva.value = celda_original.value
                    
                    try:
                        # Copiar fill
                        if celda_original.fill and celda_original.fill.start_color:
                            celda_nueva.fill = copy(celda_original.fill)
                    except:
                        pass
                    
                    try:
                        # Copiar font
                        if celda_original.font:
                            celda_nueva.font = copy(celda_original.font)
                    except:
                        pass
                    
                    try:
                        # Copiar alignment
                        if celda_original.alignment:
                            celda_nueva.alignment = copy(celda_original.alignment)
                    except:
                        pass
                    
                    try:
                        # Copiar border
                        if celda_original.border:
                            celda_nueva.border = copy(celda_original.border)
                    except:
                        pass
                    
                    try:
                        # Copiar number format
                        if celda_original.number_format:
                            celda_nueva.number_format = celda_original.number_format
                    except:
                        pass
            
            # Copiar merges de las primeras 6 filas
            try:
                for merged_range in ws_original.merged_cells.ranges:
                    if merged_range.min_row <= 6:
                        ws_nueva.merge_cells(str(merged_range))
            except:
                pass
        
        # Agregar "Nombre del Evaluador:" en fila 7
        fila_nombre_evaluador = 7
        celda_evaluador = ws_nueva.cell(row=fila_nombre_evaluador, column=1)
        celda_evaluador.value = "Nombre del Evaluador:"
        
        # Copiar el estilo de "Nombre del Colegio:" si existe
        if celda_estilo_referencia:
            try:
                if celda_estilo_referencia.fill and celda_estilo_referencia.fill.start_color:
                    celda_evaluador.fill = copy(celda_estilo_referencia.fill)
            except:
                pass
            
            try:
                if celda_estilo_referencia.font:
                    celda_evaluador.font = Font(bold=True, size=16)
            except:
                pass
            
            try:
                if celda_estilo_referencia.alignment:
                    celda_evaluador.alignment = Alignment(horizontal="left", vertical="center")
            except:
                pass
            
            try:
                if celda_estilo_referencia.border:
                    celda_evaluador.border = copy(celda_estilo_referencia.border)
            except:
                pass
        else:
            celda_evaluador.font = Font(bold=True, size=10)
            celda_evaluador.alignment = Alignment(horizontal="left", vertical="center")
        
        # Agregar cabecera personalizada
        fila_cabecera = 9
        
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, columna in enumerate(df_procesado.columns, start=1):
            celda = ws_nueva.cell(row=fila_cabecera, column=col_idx)
            celda.value = str(columna).upper()
            celda.fill = header_fill
            celda.font = header_font
            celda.alignment = header_alignment
        
        fila_inicio_datos = 10
        for row_idx, row in enumerate(dataframe_to_rows(df_procesado, index=False, header=False), start=fila_inicio_datos):
            for col_idx, value in enumerate(row, start=1):
                celda = ws_nueva.cell(row=row_idx, column=col_idx)
                celda.value = value
                celda.alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, columna in enumerate(df_procesado.columns, start=1):
            col_letter = ws_nueva.cell(row=1, column=col_idx).column_letter
            ws_nueva.column_dimensions[col_letter].width = 15
    
    buffer = BytesIO()
    wb_nuevo.save(buffer)
    buffer.seek(0)
    
    return buffer

def register_custom_font():
    """Registra la fuente Trebuchet MS si est√° disponible"""
    font_path = os.path.join("fonts", "trebuchet.ttf")
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('Trebuchet', font_path))
            return True
        except Exception as e:
            st.warning(f"No se pudo cargar la fuente Trebuchet MS: {e}")
            return False
    else:
        st.info("Fuente Trebuchet MS no encontrada. Usando fuente por defecto.")
        return False

TREBUCHET_AVAILABLE = register_custom_font()
styles_config = None

def mes_en_espanol(fecha):
    meses = {
        'January': 'enero',
        'February': 'febrero',
        'March': 'marzo',
        'April': 'abril',
        'May': 'mayo',
        'June': 'junio',
        'July': 'julio',
        'August': 'agosto',
        'September': 'septiembre',
        'October': 'octubre',
        'November': 'noviembre',
        'December': 'diciembre'
    }

    mes_ingles = fecha.strftime('%B')
    mes_espanol = meses.get(mes_ingles, mes_ingles)
    return fecha.strftime(f"%d de {mes_espanol} del %Y")

def agregar_marca_agua(pdf_bytes, watermark_path):
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_bytes)
        watermark_reader = PyPDF2.PdfReader(watermark_path)
        
        watermark_page = watermark_reader.pages[0]
        
        pdf_writer = PyPDF2.PdfWriter()
        
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            
            page_width = float(page.mediabox.width)
            page_height = float(page.mediabox.height)
            is_landscape = page_width > page_height
            
            if is_landscape:
                landscape_watermark_path = os.path.join("watermarks", "marca_agua_landscape.pdf")
                if os.path.exists(landscape_watermark_path):
                    landscape_watermark_reader = PyPDF2.PdfReader(landscape_watermark_path)
                    watermark = landscape_watermark_reader.pages[0]
                else:
                    watermark = watermark_page
            else:
                watermark = watermark_page
            
            page.merge_page(watermark)
            pdf_writer.add_page(page)
        
        result_pdf = BytesIO()
        pdf_writer.write(result_pdf)
        result_pdf.seek(0)
        
        return result_pdf
    except Exception as e:
        st.error(f"Error al aplicar marca de agua: {e}")
        return pdf_bytes

def cargar_plantillas():
    """Carga las plantillas de fondo desde la carpeta plantillas"""
    plantillas = {}
    plantillas_path = "plantillas"

    if not os.path.exists(plantillas_path):
        st.error(f"‚ùå La carpeta '{plantillas_path}' no existe. Cr√©ala y agrega las im√°genes de fondo.")
        return None

    archivos_plantilla = {
        'PROGRESIVO_1P_5S.jpg': 'fondo_1',
        'PARTICIPACION_1P_5S.jpg': 'fondo_2',
        'APROBADO_1P_3P.jpg': 'fondo_3',
        'APROBADO_4P_5S.jpg': 'fondo_4'
    }

    for archivo, clave in archivos_plantilla.items():
        ruta_completa = os.path.join(plantillas_path, archivo)
        if os.path.exists(ruta_completa):
            with open(ruta_completa, 'rb') as f:
                plantillas[clave] = f.read()
        else:
            st.warning(f"‚ö†Ô∏è No se encontr√≥ {archivo} en la carpeta plantillas")

    if len(plantillas) == 4:
        return plantillas
    else:
        st.error(f"‚ùå Se necesitan 4 plantillas, solo se encontraron {len(plantillas)}")
        return None

def clasificar_estudiantes_por_nota(df, tipo_certificado):
    """
    Clasifica estudiantes seg√∫n el tipo de certificado seleccionado.
    
    Args:
        df: DataFrame con los datos de estudiantes
        tipo_certificado: "Progresivo" o "Regular"
    
    Returns:
        dict con grupos de estudiantes clasificados
    """
    grupos = {
        'grupo_1': pd.DataFrame(),  # Progresivo
        'grupo_2': pd.DataFrame(),  # Nota < 12.5 / Participaci√≥n
        'grupo_3': pd.DataFrame(),  # Nota ‚â• 12.5 y Grado = 1P-3P
        'grupo_4': pd.DataFrame()   # Nota ‚â• 12.5 y Grado = 4P-5S
    }

    if 'nota final' not in df.columns:
        st.error("‚ùå No se encontr√≥ la columna 'NOTA FINAL' en el DataFrame")
        return None

    if 'grado' not in df.columns:
        st.error("‚ùå No se encontr√≥ la columna 'GRADO' en el DataFrame")
        return None

    # Verificar si el tipo de certificado es Progresivo (Grupo1)
    if tipo_certificado == "Progresivo":
        grupos['grupo_1'] = df.copy()
        st.info(f"üìã **Modo Progresivo seleccionado**: Todos los certificados usar√°n el formato Progresivo")

    else:  # "Regular"
        df['nota_final_num'] = pd.to_numeric(df['nota final'], errors='coerce')

        # Grupo 2: Nota < 12.5 - Participaci√≥n
        grupos['grupo_2'] = df[df['nota_final_num'] < 12.5].copy()

        # Grupos 3 y 4: Nota ‚â• 12.5
        df_nota_alta = df[df['nota_final_num'] >= 12.5].copy()

        grupos['grupo_3'] = df_nota_alta[df_nota_alta['grado'].str.lower().str.strip().isin(['1p', '2p', '3p'])].copy()
        grupos['grupo_4'] = df_nota_alta[
            df_nota_alta['grado'].str.lower().str.strip().isin(['4p', '5p', '6p', '1s', '2s', '3s', '4s', '5s'])].copy()
        
        st.info(f"üìã **Modo Normal seleccionado**: Certificados seg√∫n nota (aprobado/participaci√≥n)")

    return grupos

def validar_y_mapear_columnas(df, tipo_certificado="Regular"):
    """
    Valida que el DataFrame tenga las columnas esperadas del usuario y las mapea
    a los nombres que espera procesar_excel_inicial.
    
    Args:
        df: DataFrame a validar
        tipo_certificado: "Progresivo" o "Regular". Si es None, no valida HORAS PROGRESIVO.
    
    Retorna: (df_mapeado, exito, mensaje)
    """
    columnas_esperadas = [
        "NRO.", "PATERNO", "MATERNO", "NOMBRE", "GRADO", "SECCI√ìN", "CURSO", 
        "NOTA LABORATORIO", "¬øASISTI√ì?", "P1 4PTOS.", "P2 4PTOS.", "P3 4PTOS.", 
        "P4 4PTOS.", "P5 4PTOS.", "NOTA EVALUADOR", "NOTA FINAL", "OBSERVADOS", 
        "ESTATUS", "NUMERACI√ìN"
    ]
    
    df.columns = df.columns.str.strip()
    columnas_faltantes = [col for col in columnas_esperadas if col not in df.columns]
    requiere_horas_progresivo = tipo_certificado == "Progresivo"
    
    if requiere_horas_progresivo and "HORAS PROGRESIVO" not in df.columns:
        columnas_faltantes.append("HORAS PROGRESIVO")
    
    if columnas_faltantes:
        mensaje_error = f"‚ùå El archivo no tiene las columnas requeridas. Faltan: {', '.join(columnas_faltantes)}"
        if requiere_horas_progresivo and "HORAS PROGRESIVO" in columnas_faltantes:
            mensaje_error += "\n\nüí° Nota: La columna 'HORAS PROGRESIVO' es requerida cuando se selecciona el tipo 'Progresivo'."
        return None, False, mensaje_error
    
    mapeo_columnas = {
        "NRO.": "nro",
        "PATERNO": "paterno",
        "MATERNO": "materno",
        "NOMBRE": "nombre",
        "GRADO": "grado",
        "SECCI√ìN": "secci√≥n",
        "CURSO": "curso",
        "NOTA LABORATORIO": "nota lab",
        "¬øASISTI√ì?": "lista de asistencia",
        "NOTA EVALUADOR": "nota de examen cibertec",
        "NOTA FINAL": "nota final",
        "OBSERVADOS": "observaci√≥n sobre nota desaprobatoria",
        "ESTATUS": "status",
        "NUMERACI√ìN": "numeraci√≥n"
    }

    orden_final = [
        "NRO.", "PATERNO", "MATERNO", "NOMBRE", "GRADO", "SECCI√ìN", "CURSO",
        "NOTA LABORATORIO", "¬øASISTI√ì?", "NOTA EVALUADOR", "NOTA FINAL",
        "OBSERVADOS", "ESTATUS", "NUMERACI√ìN"
    ]

    tiene_horas_progresivo = "HORAS PROGRESIVO" in df.columns
    if tiene_horas_progresivo:
        mapeo_columnas["HORAS PROGRESIVO"] = "horas_progresivo"
        orden_final.append("HORAS PROGRESIVO")
    
    columnas_a_mantener = [col for col in orden_final if col in df.columns]
    df_filtrado = df[columnas_a_mantener].copy()
    
    if not tiene_horas_progresivo and not requiere_horas_progresivo:
        df_filtrado["HORAS PROGRESIVO"] = ""
        orden_final.append("HORAS PROGRESIVO")
    
    df_filtrado = df_filtrado[orden_final]
    filas_vacias = pd.DataFrame(columns=df_filtrado.columns, index=range(10))

    if tiene_horas_progresivo:
        mapeo_columnas["HORAS PROGRESIVO"] = "horas_progresivo"
    elif "HORAS PROGRESIVO" in orden_final:
        mapeo_columnas["HORAS PROGRESIVO"] = "horas_progresivo"
    
    nombres_mapeados = [mapeo_columnas.get(col, col.lower()) for col in orden_final]
    fila_encabezado = pd.DataFrame([nombres_mapeados], columns=df_filtrado.columns)
    
    df_formateado = pd.concat([filas_vacias, fila_encabezado, df_filtrado], ignore_index=True)

    return df_formateado, True, "‚úÖ Columnas validadas y mapeadas correctamente"

def detectar_fila_encabezado(df):
    """
    Detecta la fila que contiene el encabezado buscando palabras clave.
    Retorna el √≠ndice de la fila o None si no la encuentra.
    """
    palabras_clave = ['paterno', 'materno', 'nombre', 'nro', 'grado', 'secci√≥n', 'curso', 'nota final']
    
    for idx, row in df.iterrows():
        row_str = ' '.join(str(val).lower() for val in row if pd.notna(val))
        coincidencias = sum(1 for palabra in palabras_clave if palabra in row_str)
        if coincidencias >= 4:
            return idx
    
    return None

def procesar_excel_inicial(uploaded_file):
    """
    Procesa el archivo Excel eliminando las primeras 9 filas y columnas J-N y desde la T
    """
    try:
        columnas_requeridas = [
            "nro", "paterno", "materno", "nombre", "grado", "secci√≥n", "curso", 
            "nota lab", "lista de asistencia", "nota de examen cibertec", "nota final", 
            "observaci√≥n sobre nota desaprobatoria", "status", "numeraci√≥n", "horas_progresivo"
        ]
        
        df_original = pd.read_excel(uploaded_file)
        df_procesado = df_original.iloc[10:].copy()
        df_procesado = df_procesado.reset_index(drop=True)
        df_procesado.columns = df_procesado.iloc[0].str.lower()
        df_procesado = df_procesado.drop(df_procesado.index[0]).reset_index(drop=True)
        columnas_existentes = [col for col in columnas_requeridas if col in df_procesado.columns]
        df_procesado = df_procesado[columnas_existentes]

        if 'nota final' in df_procesado.columns:
            df_procesado['nota final'] = df_procesado['nota final'].apply(
                lambda x: 0 if isinstance(x, str) and x.strip().upper() == 'NP' else x
            )

        df_procesado['nombre_certificado'] = df_procesado['nombre'].fillna('').str.strip() + ' ' + df_procesado[
            'paterno'].fillna('').str.strip() + ' ' + df_procesado['materno'].fillna('').str.strip()

        if 'nro' in df_procesado.columns:
            columnas = df_procesado.columns.tolist()
            columnas.remove('nombre_certificado')
            posicion_nro = columnas.index('nro')
            columnas.insert(posicion_nro + 1, 'nombre_certificado')
            df_procesado = df_procesado[columnas]

        return df_procesado, True, "Archivo procesado correctamente"

    except Exception as e:
        return None, False, f"Error al procesar el archivo: {str(e)}"

def wrap_text_to_width(canvas, text, font_name, font_size, max_width_mm):
    max_width_points = max_width_mm * 2.83465
    words = text.split()
    lines = []
    current_line = []

    for word in words:
        test_line = current_line + [word]
        test_text = ' '.join(test_line)
        text_width = canvas.stringWidth(test_text, font_name, font_size)

        if text_width <= max_width_points:
            current_line = test_line
        else:
            if current_line:
                lines.append(' '.join(current_line))
                current_line = [word]
            else:
                lines.append(word)

    if current_line:
        lines.append(' '.join(current_line))

    return lines

def draw_multiline_text(canvas, text, style_key, page_width, styles_config, max_width_mm=None):
    style = styles_config[style_key]
    font_name = style['font_family'] if TREBUCHET_AVAILABLE else 'Helvetica'

    is_bold = style.get('bold', False)

    if is_bold:
        try:
            bold_font_name = f"{font_name}-Bold"
            canvas.setFont(bold_font_name, style['font_size'])
            font_name = bold_font_name
        except Exception as e:
            try:
                canvas.setFont(font_name, style['font_size'])
            except Exception as e:
                canvas.setFont('Helvetica-Bold' if is_bold else 'Helvetica', style['font_size'])
                font_name = 'Helvetica-Bold' if is_bold else 'Helvetica'
    else:
        try:
            canvas.setFont(font_name, style['font_size'])
        except Exception as e:
            canvas.setFont('Helvetica', style['font_size'])
            font_name = 'Helvetica'

    canvas.setFillColor(HexColor(style['color']))
    x_points = style['x'] * 2.83465
    y_points = style['y'] * 2.83465

    if max_width_mm is None:
        if style['x'] == 148 or style['x'] == 105:
            text_width = canvas.stringWidth(text, font_name, style['font_size'])
            x_points = (page_width - text_width) / 2
        canvas.drawString(x_points, y_points, text)
        return style['font_size']

    lines = wrap_text_to_width(canvas, text, font_name, style['font_size'], max_width_mm)
    line_height = style['font_size'] * 1.2
    start_y = y_points

    for i, line in enumerate(lines):
        line_y = start_y - (i * line_height)
        if style['x'] == 148 or style['x'] == 105:
            text_width = canvas.stringWidth(line, font_name, style['font_size'])
            line_x = (page_width - text_width) / 2
        else:
            line_x = x_points
        canvas.drawString(line_x, line_y, line)

    return line_height * len(lines)

def generar_certificados_grupo(grupo_df, plantilla_bytes, plantilla_key, nombre_grupo, zip_file, progress_bar,
    estudiantes_base, total_estudiantes, styles_config_by_template):
    certificados_generados = 0

    usar_marca_agua = st.session_state.get('usar_marca_agua_seleccionado', False)
    aplicar_marca_agua = usar_marca_agua and plantilla_key != 'fondo_2'
    
    watermark_path = os.path.join("watermarks", "marca_agua.pdf")
    if aplicar_marca_agua and not os.path.exists(watermark_path):
        st.warning(f"‚ö†Ô∏è No se encontr√≥ el archivo de marca de agua en {watermark_path}. Se generar√°n PDFs sin marca de agua.")
        aplicar_marca_agua = False

    styles_config = styles_config_by_template[plantilla_key]

    if styles_config.get('orientation') == 'portrait':
        page_size = A4
        page_width, page_height = A4
    else:
        page_size = landscape(A4)
        page_width, page_height = landscape(A4)

    for i, row in grupo_df.iterrows():
        try:
            nombre = str(row["nombre_certificado"]).strip().upper()
            curso = str(row["curso"]).strip().upper()

            fecha_seleccionada = st.session_state.get('fecha_certificado_seleccionada', datetime.now().date())
            if isinstance(fecha_seleccionada, datetime):
                fecha_para_certificado = fecha_seleccionada
            else:
                fecha_para_certificado = datetime.combine(fecha_seleccionada, datetime.min.time())
            
            fecha = mes_en_espanol(fecha_para_certificado)
            numero = (
                str(row["numeraci√≥n"]).strip()
                if "numeraci√≥n" in row and pd.notnull(row["numeraci√≥n"])
                else f"GEN-{i + 1:03}"
            )

            horas = "horas_progresivo"
            horas_progresivo = ""
            if plantilla_key == 'fondo_1' and horas in row and pd.notnull(row[horas]):
                horas_progresivo = str(row[horas])

            with NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
                tmp_img.write(plantilla_bytes)
                tmp_img.flush()
                tmp_img_path = tmp_img.name

            pdf_buffer = BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=page_size)
            c.drawImage(tmp_img_path, 0, 0, width=page_width, height=page_height)
            draw_multiline_text(c, nombre, 'nombre', page_width, styles_config, styles_config['nombre']['max_width'])
            draw_multiline_text(c, curso, 'curso', page_width, styles_config, styles_config['curso']['max_width'])
            draw_multiline_text(c, f"Lima, {fecha}", 'fecha', page_width, styles_config)

            if plantilla_key == 'fondo_1' and horas_progresivo:
                draw_multiline_text(c, horas_progresivo, 'horas', page_width, styles_config)
            
            if plantilla_key != 'fondo_2':
                draw_multiline_text(c, f"Certificado N¬∫ {numero}", 'numero', page_width, styles_config)

            c.save()
            pdf_bytes = pdf_buffer.getvalue()

            if aplicar_marca_agua:
                pdf_buffer = agregar_marca_agua(BytesIO(pdf_bytes), watermark_path)
                pdf_bytes = pdf_buffer.getvalue()

            if plantilla_key == 'fondo_2':
                pdf_name = f"Constancias/{nombre.strip().replace(' ', '_') + '_' + curso[0:11].replace(' ', '_')}.pdf"
            else:
                pdf_name = f"{nombre.strip().replace(' ', '_') + '_' + curso[0:11].replace(' ', '_')}.pdf"

            zip_file.writestr(pdf_name, pdf_bytes)

            certificados_generados += 1

            progreso_actual = (estudiantes_base + certificados_generados) / total_estudiantes
            progress_bar.progress(min(progreso_actual, 1.0))

            try:
                if os.path.exists(tmp_img_path):
                    os.unlink(tmp_img_path)
            except:
                pass

        except Exception as e:
            st.error(f"Error generando certificado para {nombre}: {e}")

    return certificados_generados

def generar_todos_certificados():
    if st.session_state.grupos and st.session_state.plantillas:
        st.info("Generando certificados por grupos...")

        total_estudiantes = sum(len(grupo) for grupo in st.session_state.grupos.values() if not grupo.empty)
        progress_bar = st.progress(0)
        estudiantes_procesados = 0

        zip_buffer = BytesIO()

        with ZipFile(zip_buffer, "a") as zip_file:
            zip_file.writestr("Constancias/", "")

            styles_config_by_template = {
            "fondo_1": {
                'curso': {
                    'font_family': 'Trebuchet',
                    'font_size': 32,
                    'color': '#000000',
                    'x': 52,
                    'y': 129,
                    'max_width': 220,
                    'bold': True
                },
                'nombre': {
                    'font_family': 'Trebuchet',
                    'font_size': 25,
                    'color': '#000000',
                    'x': 52,
                    'y': 85,
                    'max_width': 210
                },
                'fecha': {
                    'font_family': 'Trebuchet',
                    'font_size': 18,
                    'color': '#004064',
                    'x': 52,
                    'y': 36,
                    'max_width': None,
                    'bold': True
                },
                'numero': {
                    'font_family': 'Trebuchet',
                    'font_size': 15.5,
                    'color': '#004064',
                    'x': 52,
                    'y': 27,
                    'max_width': None
                },
                'horas': {
                    'font_family': 'Trebuchet',
                    'font_size': 15.5,
                    'color': '#004064',
                    'x': 132.5,
                    'y': 65.2,
                    'max_width': None
                },
                'orientation': 'landscape'
            },
            "fondo_2": {  # Vertical
                'curso': {
                    'font_family': 'Trebuchet',
                    'font_size': 30.5,
                    'color': '#000000',
                    'x': 105,
                    'y': 185,
                    'max_width': 160,
                    'bold': True
                },
                'nombre': {
                    'font_family': 'Trebuchet',
                    'font_size': 29,
                    'color': '#000000',
                    'x': 105,
                    'y': 133,
                    'max_width': 160,
                    'bold': True
                },
                'fecha': {
                    'font_family': 'Trebuchet',
                    'font_size': 18,
                    'color': '#004064',
                    'x': 105,
                    'y': 78,
                    'max_width': None
                },
                # No aparece en el certificado, s√≥lo est√° para evitar errores en f()
                'numero': {
                    'font_family': 'Trebuchet',
                    'font_size': 1,
                    'color': '#ffffff',
                    'x': 0,
                    'y': 0,
                    'max_width': None
                },
                'orientation': 'portrait'
            },
            "fondo_3": {
                'curso': {
                    'font_family': 'Trebuchet',
                    'font_size': 30.5,
                    'color': '#000000',
                    'x': 148,
                    'y': 117,
                    'max_width': 245,
                    'bold': True
                },
                'nombre': {
                    'font_family': 'Trebuchet',
                    'font_size': 29,
                    'color': '#000000',
                    'x': 148,
                    'y': 75,
                    'max_width': 245,
                    'bold': True
                },
                'fecha': {
                    'font_family': 'Trebuchet',
                    'font_size': 18,
                    'color': '#004064',
                    'x': 20,
                    'y': 41,
                    'max_width': None,
                    'bold': True
                },
                'numero': {
                    'font_family': 'Trebuchet',
                    'font_size': 15.5,
                    'color': '#004064',
                    'x': 20,
                    'y': 32,
                    'max_width': None
                },
                'orientation': 'landscape'
            },
            "fondo_4": {
                'curso': {
                    'font_family': 'Trebuchet',
                    'font_size': 30.5,
                    'color': '#000000',
                    'x': 148,
                    'y': 117,
                    'max_width': 245,
                    'bold': True
                },
                'nombre': {
                    'font_family': 'Trebuchet',
                    'font_size': 29,
                    'color': '#000000',
                    'x': 148,
                    'y': 75,
                    'max_width': 245,
                    'bold': True
                },
                'fecha': {
                    'font_family': 'Trebuchet',
                    'font_size': 18,
                    'color': '#004064',
                    'x': 20,
                    'y': 41,
                    'max_width': None,
                    'bold': True
                },
                'numero': {
                    'font_family': 'Trebuchet',
                    'font_size': 15.5,
                    'color': '#004064',
                    'x': 20,
                    'y': 32,
                    'max_width': None
                },
                'orientation': 'landscape'
            }
        }

            mapeo_plantillas = {
                'grupo_1': 'fondo_1',  # Progresiva
                'grupo_2': 'fondo_2',  # Participaci√≥n Nota < 12.5
                'grupo_3': 'fondo_3',  # Base - Nota ‚â• 12.5 y Grado = 1P-3P
                'grupo_4': 'fondo_4'   # Base - Nota ‚â• 12.5 y Grado = 4P-5S
            }

            for grupo_nombre, grupo_df in st.session_state.grupos.items():
                if not grupo_df.empty:
                    plantilla_key = mapeo_plantillas[grupo_nombre]
                    plantilla_bytes = st.session_state.plantillas[plantilla_key]

                    st.write(f"Procesando {grupo_nombre} ({len(grupo_df)} estudiantes) con plantilla {plantilla_key}...")

                    certificados_gen = generar_certificados_grupo(
                        grupo_df,
                        plantilla_bytes,
                        plantilla_key,
                        grupo_nombre,
                        zip_file,
                        progress_bar,
                        estudiantes_procesados,
                        total_estudiantes,
                        styles_config_by_template
                    )

                    estudiantes_procesados += len(grupo_df)

                    st.success(f"‚úÖ {grupo_nombre}: {certificados_gen} certificados generados con estilo {plantilla_key}")

        zip_buffer.seek(0)
        st.success("üéâ Todos los certificados han sido generados correctamente y est√°n listos para su descarga.")
        
        st.session_state.zip_buffer = zip_buffer
        st.session_state.certificados_generados = True
        
        return True
    return False

def ajustar_texto_inteligente(draw, text, font_path, font_size_inicial, max_width, max_height, 
                               min_font_size=20, line_spacing=1.2):
    """
    Ajusta el texto para que quepa en el espacio disponible.
    Intenta primero dividir en dos l√≠neas, si no cabe, reduce el tama√±o de fuente.
    """
    
    def get_text_size(text, font):
        """Obtiene el ancho y alto del texto"""
        bbox = draw.textbbox((0, 0), text, font=font)
        width = bbox[2] - bbox[0]
        height = bbox[3] - bbox[1]
        return width, height
    
    def dividir_texto_en_dos_lineas(text):
        """Divide el texto en dos l√≠neas de manera inteligente"""
        palabras = text.split()
        if len(palabras) <= 2:
            return [text]
        mitad = len(palabras) // 2
        linea1 = ' '.join(palabras[:mitad])
        linea2 = ' '.join(palabras[mitad:])
        return [linea1, linea2]
    
    current_font_size = font_size_inicial
    
    try:
        font = ImageFont.truetype(font_path, current_font_size)
    except:
        font = ImageFont.load_default()
        return [text], font
    
    text_width, text_height = get_text_size(text, font)
    
    # Caso 1: El texto cabe en una sola l√≠nea con el tama√±o original
    if text_width <= max_width and text_height <= max_height:
        return [text], font
    
    # Caso 2: Intentar dividir en dos l√≠neas con el tama√±o original
    lineas = dividir_texto_en_dos_lineas(text)
    
    if len(lineas) == 2:
        ancho_linea1, alto_linea1 = get_text_size(lineas[0], font)
        ancho_linea2, alto_linea2 = get_text_size(lineas[1], font)
        max_ancho_lineas = max(ancho_linea1, ancho_linea2)
        alto_total = (alto_linea1 + alto_linea2) * line_spacing
        
        if max_ancho_lineas <= max_width and alto_total <= max_height:
            return lineas, font
    
    # Caso 3: Reducir tama√±o de fuente gradualmente
    for size in range(current_font_size - 5, min_font_size - 1, -5):
        try:
            font = ImageFont.truetype(font_path, size)
        except:
            continue
        
        lineas = dividir_texto_en_dos_lineas(text)
        
        if len(lineas) == 2:
            ancho_linea1, alto_linea1 = get_text_size(lineas[0], font)
            ancho_linea2, alto_linea2 = get_text_size(lineas[1], font)
            max_ancho_lineas = max(ancho_linea1, ancho_linea2)
            alto_total = (alto_linea1 + alto_linea2) * line_spacing
            
            if max_ancho_lineas <= max_width and alto_total <= max_height:
                return lineas, font
        
        text_width, text_height = get_text_size(text, font)
        if text_width <= max_width and text_height <= max_height:
            return [text], font
    
    # √öltimo recurso: usar tama√±o m√≠nimo
    try:
        font = ImageFont.truetype(font_path, min_font_size)
    except:
        font = ImageFont.load_default()
    
    return dividir_texto_en_dos_lineas(text), font

def draw_centered_text_adaptive(draw, text, x_center, y_center, font_path, 
                                font_size_inicial, max_width, max_height,
                                min_font_size=20, fill="white", line_spacing=1.2):
    """
    Dibuja texto centrado con ajuste autom√°tico de tama√±o y divisi√≥n en l√≠neas.
    """
    
    lineas, fuente_final = ajustar_texto_inteligente(
        draw, text, font_path, font_size_inicial, 
        max_width, max_height, min_font_size, line_spacing
    )
    
    lineas_info = []
    for linea in lineas:
        bbox = draw.textbbox((0, 0), linea, font=fuente_final)
        width = bbox[2] - bbox[0]
        height = bbox[3] - bbox[1]
        lineas_info.append({'text': linea, 'width': width, 'height': height})
    
    if len(lineas_info) == 1:
        total_height = lineas_info[0]['height']
    else:
        total_height = sum(info['height'] for info in lineas_info) * line_spacing
    
    y_actual = y_center - (total_height / 2)
    
    for i, info in enumerate(lineas_info):
        x_pos = x_center - (info['width'] / 2)
        draw.text((x_pos, y_actual), info['text'], fill=fill, font=fuente_final)
        
        if i < len(lineas_info) - 1:
            y_actual += info['height'] * line_spacing
    
    return {
        'lineas': len(lineas),
        'font_size': fuente_final.size if hasattr(fuente_final, 'size') else font_size_inicial
    }

def draw_centered_text(draw, text, x_position, y_position, font, fill="white"):
    """Versi√≥n original - mantenida para compatibilidad"""
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = bbox[2] - bbox[0]
    bboxY = draw.textbbox((0, 0), text, font=font)
    text_height = bboxY[3] - bboxY[1]
    x_position = x_position - (text_width) / 2
    y_position = y_position - (text_height) / 2
    draw.text((x_position, y_position), text, fill=fill, font=font)


st.title("üìä Sistema de Validaci√≥n de Archivos")

tab1, tab2, tab3, tab4, tab5 = st.tabs(["üîç Validador de N√≥minas", "‚öñÔ∏è Validador Evaluaciones de Alumnos", "üìë Generador de Planilla de Resultados", "üéì Generador de Diplomas, Certificados y Constancias", "üèÖ Generador de Insignias (Docente y Alumno)"])

# TAB 1: VALIDADOR GENERAL
with tab1:
    st.markdown("## üîç Validador de N√≥minas")
    st.markdown("### Sistema de Homologaci√≥n de Datos")

    mostrar_stepper(st.session_state.paso_actual)

    if st.session_state.paso_actual == 0:
        st.header("üè´ Paso 1: Informaci√≥n del Colegio")

        st.markdown("""
            <div style='background-color: #78808C; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
                <h4>Bienvenido al sistema de validaci√≥n</h4>
                <p>Para comenzar, ingresa el Nombre del colegio. Este Nombre se usar√° para identificar los archivos descargables.</p>
            </div>
            """, unsafe_allow_html=True)

        col1, col2 = st.columns([3, 1])
        
        with col1:
            index_seleccionado = None

            if st.session_state.nombre_colegio in LISTA_COLEGIOS:
                index_seleccionado = LISTA_COLEGIOS.index(st.session_state.nombre_colegio)

            NOMBRES = st.selectbox(
                "Selecciona el colegio:",
                options=LISTA_COLEGIOS,
                index=index_seleccionado,
                placeholder="Elige un colegio..."
            )
            
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("‚û°Ô∏è Continuar", type="primary", use_container_width=True, disabled = not NOMBRES):
                if NOMBRES:
                    st.session_state.nombre_colegio = NOMBRES.strip()
                    st.session_state.paso_actual = 1
                    st.rerun()
                else:
                    st.error("Por favor, ingresa el Nombre del colegio")

    # N√ìMINA
    elif st.session_state.paso_actual == 1:
        with st.expander("‚úÖ Paso 1 completado: Nombre del Colegio", expanded=False):
            st.info(f"**Colegio:** {st.session_state.nombre_colegio}")
            if st.button("üîÑ Cambiar Nombre", key="cambiar_nombre"):
                st.session_state.paso_actual = 0
                st.rerun()
        
        st.header("üìã Paso 2: Archivo de N√≥mina de Alumnos")
        
        with st.container():
            st.markdown("""
            <div style='background-color: #78808C; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
                <h4>üìÑ Instrucciones</h4>
                <p>Sube el archivo Excel que contiene la n√≥mina de alumnos.</p>
                <p><strong>Columnas requeridas:</strong></p>
                <code>Nro., Paterno, Materno, Nombres, Nacimiento (DD/MM/YYYY), Sexo (M/F), Grado, Secci√≥n, Correo institucional, Neurodiversidad (S√≠/No), DNI</code>
            </div>
            """, unsafe_allow_html=True)
            
            archivo = st.file_uploader(
                "Selecciona el archivo Excel",
                type=["xls", "xlsx"],
                help="El sistema detectar√° autom√°ticamente la fila de cabecera"
            )
            
            if archivo is not None:
                st.session_state.archivo1_bytes = archivo.getvalue()
                with st.spinner("üîç Analizando archivo..."):
                    try:
                        df_original = pd.read_excel(archivo, header=None)
                        fila_detectada = detectar_cabecera_automatica(df_original, COLUMNAS_ARCHIVO1)
                        
                        if fila_detectada is not None:
                            st.session_state.archivo1_fila_cabecera = fila_detectada
                            st.success(f"‚úÖ Cabecera detectada autom√°ticamente en la fila {fila_detectada + 1}")
                            
                            df = pd.read_excel(archivo, header=fila_detectada)
                            columnas_norm = {c.strip().lower(): c for c in df.columns}
                            cols_a_usar = []
                            for col_req in COLUMNAS_ARCHIVO1:
                                col_norm = col_req.strip().lower()
                                if col_norm in columnas_norm:
                                    cols_a_usar.append(columnas_norm[col_norm])
                            
                            df = df[cols_a_usar]
                            df.columns = [col.upper() for col in COLUMNAS_ARCHIVO1]

                            df = limpiar_filas_vacias(df, columnas_clave=["PATERNO", "MATERNO", "NOMBRES"])

                            if df.empty:
                                st.error("‚ùå La hoja seleccionada no contiene datos v√°lidos despu√©s de limpiar filas vac√≠as")
                                st.stop()
                            
                            df = convertir_numericas_a_entero(df, columnas=["GRADO"])
                            df = homologar_dataframe(df)

                            columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES"]
                            filas_vacias = df[df[columnas_obligatorias].isnull().any(axis=1)]

                            if not filas_vacias.empty:
                                st.error("‚ùå Se detectaron campos vac√≠os en nombres o apellidos (Archivo 1 - N√≥mina)")
                                st.dataframe(filas_vacias, use_container_width=True)
                                st.stop()
                            
                            errores_fatales = []
                            alertas = []
                            
                            df, errores_grados = validar_y_mapear_grados(df, "GRADO")
                            errores_fatales.extend(errores_grados)
                            
                            errores_sexo = validar_sexo(df, "SEXO (M/F)")
                            alertas.extend(errores_sexo)
                            
                            errores_secciones = validar_secciones(df, "SECCI√ìN")
                            errores_fatales.extend(errores_secciones)

                            errores_neuro = validar_neurodiversidad(df, "NEURODIVERSIDAD (S√ç/NO)")
                            alertas.extend(errores_neuro)
                            
                            errores_fecha = validar_fecha_nacimiento(df, "NACIMIENTO (DD/MM/YYYY)")
                            alertas.extend(errores_fecha)
                            
                            errores_dni = validar_dni(df, "DNI")
                            alertas.extend(errores_dni)
                            
                            errores_correo = validar_correo(df, "CORREO INSTITUCIONAL")
                            alertas.extend(errores_correo)
                            
                            if errores_fatales:
                                st.error("‚ùå Se encontraron errores de validaci√≥n:")
                                df_errores_fatales = pd.DataFrame(errores_fatales, columns=["Detalle de los errores cr√≠ticos"])
                                    
                                st.dataframe(
                                    df_errores_fatales,
                                    use_container_width=True,
                                    height=220
                                )
                                    
                                st.caption(f"üîé Total de errores: {len(errores_fatales)}")
                                st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                st.stop()
                                
                            else:
                                df["IDENTIFICADOR"] = crear_identificador(df, "PATERNO", "MATERNO", "NOMBRES")
                                st.session_state.archivo1_df = df
                                
                                if alertas:
                                    st.warning("‚ö†Ô∏è Se detectaron advertencias en los datos (no bloquean el proceso):")
                                    with st.expander("Ver alertas detalladas", expanded=True):
                                        df_alertas = pd.DataFrame(alertas, columns=["Detalle de la Alerta"])
                                        
                                        st.dataframe(
                                            df_alertas,
                                            use_container_width=True,
                                            height=220
                                        )
                                        
                                        st.caption(f"üîé Total de alertas: {len(alertas)}")
                                else:
                                    st.success("‚úÖ Todas las validaciones pasaron correctamente")

                            if not errores_fatales:
                                st.markdown("### üìä Vista Previa de Datos")
                                st.info(f"Total de registros: {len(df)}")
                                st.dataframe(df, use_container_width=True, hide_index=True)

                            col1, col2 = st.columns(2)
                            with col1:
                                df_descarga = df.drop(columns=["IDENTIFICADOR", "N¬∫"], errors="ignore")
                                df_descarga = df_descarga.fillna("")
                                df_descarga = df_descarga.replace(["NAN", "nan", "NaN"], "")

                                buffer = guardar_con_formato_original(
                                    df_procesado=df_descarga,
                                    archivo_original_bytes=st.session_state.archivo1_bytes,
                                    nombre_hoja=None,
                                    fila_cabecera=st.session_state.archivo1_fila_cabecera
                                )
                                st.download_button(
                                    label="üíæ Descargar Archivo Homologado",
                                    data=buffer,
                                    file_name=f"{st.session_state.nombre_colegio}_nomina_RV.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            with col2:
                                if st.button("‚û°Ô∏è Continuar al Paso 3", type="primary", use_container_width=True):
                                    st.session_state.paso_actual = 2
                                    st.rerun()
                        
                        else:
                            st.warning("‚ö†Ô∏è No se pudo detectar la cabecera autom√°ticamente")
                            st.markdown("### üîç Detecci√≥n Manual")
                            st.dataframe(df_original.iloc[:15, :15], use_container_width=True)
                            
                            fila_manual = st.number_input(
                                "Indica el n√∫mero de fila que contiene la cabecera:",
                                min_value=1, max_value=15, step=1
                            )
                            
                            if st.button("‚úîÔ∏è Validar Fila Seleccionada", type="primary"):
                                fila_idx = fila_manual - 1
                                fila = df_original.iloc[fila_idx].astype(str).str.strip().str.lower().tolist()
                                columnas_req_norm = [c.lower() for c in COLUMNAS_ARCHIVO1]
                                
                                if all(col in fila for col in columnas_req_norm):
                                    st.success("‚úÖ Cabecera v√°lida")
                                    df = pd.read_excel(archivo, header=fila_idx)
                                    
                                    columnas_norm = {c.strip().lower(): c for c in df.columns}
                                    cols_a_usar = []
                                    for col_req in COLUMNAS_ARCHIVO1:
                                        col_norm = col_req.strip().lower()
                                        if col_norm in columnas_norm:
                                            cols_a_usar.append(columnas_norm[col_norm])
                                    
                                    df = df[cols_a_usar]
                                    df.columns = [col.upper() for col in COLUMNAS_ARCHIVO1]
                                    
                                    df = homologar_dataframe(df)
                                    
                                    columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES"]
                                    filas_vacias = df[df[columnas_obligatorias].isnull().any(axis=1)]

                                    if not filas_vacias.empty:
                                        st.error("‚ùå Se detectaron campos vac√≠os en nombres o apellidos (Archivo 1 - N√≥mina)")
                                        st.dataframe(filas_vacias, use_container_width=True)
                                        st.stop()
                                    
                                    errores_fatales = []
                                    alertas = []
                                    
                                    df, errores_grados = validar_y_mapear_grados(df, "GRADO")
                                    errores_fatales.extend(errores_grados)
                                    
                                    errores_sexo = validar_sexo(df, "SEXO (M/F)")
                                    alertas.extend(errores_sexo)
                                    
                                    errores_secciones = validar_secciones(df, "SECCI√ìN")
                                    errores_fatales.extend(errores_secciones)

                                    errores_neuro = validar_neurodiversidad(df, "NEURODIVERSIDAD (S√ç/NO)")
                                    alertas.extend(errores_neuro)
                                    
                                    errores_fecha = validar_fecha_nacimiento(df, "NACIMIENTO (DD/MM/YYYY)")
                                    alertas.extend(errores_fecha)
                                    
                                    errores_dni = validar_dni(df, "DNI")
                                    alertas.extend(errores_dni)
                                    
                                    errores_correo = validar_correo(df, "CORREO INSTITUCIONAL")
                                    alertas.extend(errores_correo)
                                    
                                    if errores_fatales:
                                        st.error("‚ùå Se encontraron errores de validaci√≥n:")
                                        df_errores_fatales = pd.DataFrame(errores_fatales, columns=["Detalle de la Alerta"])
                                            
                                        st.dataframe(
                                            df_errores_fatales,
                                            use_container_width=True,
                                            height=220
                                        )
                                            
                                        st.caption(f"üîé Total de errores: {len(errores_fatales)}")
                                        st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                        st.stop()

                                    else:
                                        df["IDENTIFICADOR"] = crear_identificador(df, "PATERNO", "MATERNO", "NOMBRES")
                                        st.session_state.archivo1_df = df
                                        
                                        if alertas:
                                            st.warning("‚ö†Ô∏è Se detectaron advertencias en los datos (no bloquean el proceso):")
                                            with st.expander("Ver alertas detalladas", expanded=True):
                                                df_alertas = pd.DataFrame(alertas, columns=["Detalle de la Alerta"])
                                                
                                                st.dataframe(
                                                    df_alertas,
                                                    use_container_width=True,
                                                    height=220
                                                )
                                                
                                                st.caption(f"üîé Total de alertas: {len(alertas)}")
                                        else:
                                            st.success("‚úÖ Todas las validaciones pasaron correctamente")

                                    if not errores_fatales:
                                        st.markdown("### üìä Vista Previa de Datos")
                                        st.info(f"Total de registros: {len(df)}")
                                        st.dataframe(df, use_container_width=True, hide_index=True)
                                    
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        df_descarga = df.drop(columns=["IDENTIFICADOR", "N¬∫"], errors="ignore")
                                        df_descarga = df_descarga.fillna("")
                                        df_descarga = df_descarga.replace(["NAN", "nan", "NaN"], "")

                                        buffer = guardar_con_formato_original(
                                            df_procesado=df_descarga,
                                            archivo_original_bytes=st.session_state.archivo1_bytes,
                                            nombre_hoja=None,
                                            fila_cabecera=st.session_state.archivo1_fila_cabecera
                                        )
                                        st.download_button(
                                            label="üíæ Descargar Archivo Homologado",
                                            data=buffer,
                                            file_name=f"{st.session_state.nombre_colegio}_nomina_RV.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            use_container_width=True
                                        )
                                    with col2:
                                        if st.button("‚û°Ô∏è Continuar al Paso 3", type="primary", use_container_width=True):
                                            st.session_state.paso_actual = 2
                                            st.rerun()
                                else:
                                    st.error("‚ùå La fila seleccionada no contiene todas las columnas requeridas")
                    
                    except Exception as e:
                        st.error(f"‚ùå Error al procesar el archivo: {e}")

    # NOTAS
    elif st.session_state.paso_actual == 2:
        with st.expander("‚úÖ Pasos completados", expanded=False):
            st.success(f"**Colegio:** {st.session_state.nombre_colegio}")
            st.success(f"**Archivo 1:** {len(st.session_state.archivo1_df)} registros cargados")
            if st.button("üîô Volver al Paso 2", key="volver_paso2"):
                st.session_state.paso_actual = 1
                st.rerun()
        
        st.header("üìä Paso 3: Archivo de Notas de Cursos")
        
        with st.expander("‚öôÔ∏è Configuraci√≥n de Cursos Equivalentes", expanded=False):
            st.markdown("""
            <div style='background-color: #78808C; padding: 15px; border-radius: 10px;'>
                <p>Opcionalmente, puedes cargar un archivo .txt con cursos adicionales para reconocimiento autom√°tico.</p>
            </div>
            """, unsafe_allow_html=True)
            
            archivo_txt = st.file_uploader("Archivo de equivalencias (.txt)", type=["txt"])
            if archivo_txt:
                contenido = archivo_txt.getvalue().decode("utf-8", errors="ignore")
                nuevos = [l.strip().upper() for l in contenido.splitlines() if l.strip()]
                st.session_state.cursos_equivalentes = sorted(list(set(st.session_state.cursos_equivalentes + nuevos)))
                st.success(f"‚úÖ {len(nuevos)} cursos agregados. Total: {len(st.session_state.cursos_equivalentes)}")
        
        st.markdown("""
        <div style='background-color: #78808C; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
            <h4>üìÑ Instrucciones</h4>
            <p>Sube el archivo Excel con las notas de los cursos.</p>
            <p>Columnas requeridas para <strong>1P-3P</strong>:</p>
            <code>NRO., PATERNO, MATERNO, NOMBRES, CURSO, GRADO, SECCI√ìN, NOTA VIGESIMAL 100%</code>
            <p></p>
            <p>Columnas requeridas para <strong>4P-5S</strong>:</p>
            <code>NRO., PATERNO, MATERNO, NOMBRES, CURSO, GRADO, SECCI√ìN, NOTA VIGESIMAL 25%</code>

                    
        </div>
        """, unsafe_allow_html=True)
        
        archivo2 = st.file_uploader("Selecciona el archivo Excel de notas", type=["xls", "xlsx"])
        
        if archivo2 is not None:
            st.session_state.archivo2_bytes = archivo2.getvalue()
            with st.spinner("üîç Analizando archivo y hojas disponibles..."):
                try:
                    xls_file = pd.ExcelFile(archivo2)
                    hojas_disponibles = xls_file.sheet_names
                    tiene_1p3p = "1P-3P" in hojas_disponibles
                    tiene_4p5s = "4P-5S" in hojas_disponibles
                    
                    if not tiene_1p3p and not tiene_4p5s:
                        st.error("‚ùå El archivo no contiene ninguna de las hojas requeridas: '1P-3P' o '4P-5S'")
                        st.info(f"Hojas encontradas: {', '.join(hojas_disponibles)}")
                        st.stop()
                    
                    st.success(f"‚úÖ Hojas detectadas en el archivo, √önicas Opciones ('1P-3P' o '4P-5S'):")
                    cols_info = st.columns(2)
                    with cols_info[0]:
                        if tiene_1p3p:
                            st.info("üìò **1P-3P** encontrada")
                    with cols_info[1]:
                        if tiene_4p5s:
                            st.info("üìó **4P-5S** encontrada")
                    
                    st.divider()
                    
                    # PROCESAR HOJA 1P-3P (Solo may√∫sculas)
                    df_1p3p_procesado = None
                    df_vp_1p3p = None

                    if tiene_1p3p:
                        st.markdown("### üìò Procesando Hoja: 1P-3P")
                        
                        df_1p3p_original = pd.read_excel(archivo2, sheet_name="1P-3P", header=None)
                        fila_detectada_1p3p = detectar_cabecera_automatica(df_1p3p_original, COLUMNAS_ARCHIVO2_1P3P)
                        
                        if fila_detectada_1p3p is not None:
                            st.session_state.archivo2_1p3p_fila_cabecera = fila_detectada_1p3p
                            st.success(f"‚úÖ Cabecera detectada en la fila {fila_detectada_1p3p + 1}")
                            
                            df_1p3p = pd.read_excel(archivo2, sheet_name="1P-3P", header=fila_detectada_1p3p)
                            
                            columnas_norm = {c.strip().lower(): c for c in df_1p3p.columns}
                            cols_a_usar = []
                            for col_req in COLUMNAS_ARCHIVO2_1P3P:
                                col_norm = col_req.strip().lower()
                                if col_norm in columnas_norm:
                                    cols_a_usar.append(columnas_norm[col_norm])
                            
                            df_1p3p = df_1p3p[cols_a_usar]
                            df_1p3p.columns = [col.upper() for col in COLUMNAS_ARCHIVO2_1P3P]
                            df_1p3p = limpiar_filas_vacias(df_1p3p, columnas_clave=["PATERNO", "MATERNO", "NOMBRES"])

                            if not df_1p3p.empty:
                                df_1p3p = convertir_numericas_a_entero(df_1p3p, columnas=["GRADO", "NOTA VIGESIMAL 100%"])
                                df_1p3p = homologar_dataframe(df_1p3p)
                                columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES"]
                                filas_vacias = df_1p3p[df_1p3p[columnas_obligatorias].isnull().any(axis=1)]

                                if not filas_vacias.empty:
                                    st.error("‚ùå Se detectaron campos vac√≠os en nombres o apellidos (Hoja 1P-3P)")
                                    st.dataframe(filas_vacias, use_container_width=True)
                                    st.stop()
                                
                                errores_validacion_1p3p = []

                                if "NOTA VIGESIMAL 100%" in df_1p3p.columns:
                                    df_1p3p["NOTA VIGESIMAL 100%"] = df_1p3p["NOTA VIGESIMAL 100%"].fillna("NP").replace("", "NP")

                                df_1p3p, errores_grados = validar_y_mapear_grados(df_1p3p, "GRADO", tipo_validacion="1p3p")
                                errores_validacion_1p3p.extend(errores_grados)
                                errores_secciones = validar_secciones(df_1p3p, "SECCI√ìN")
                                errores_validacion_1p3p.extend(errores_secciones)
                                
                                if errores_validacion_1p3p:
                                    st.error("‚ùå Errores de validaci√≥n en 1P-3P:")
                                    df_errores_fatales_1p3p = pd.DataFrame(errores_validacion_1p3p, columns=["Detalle de los errores cr√≠ticos"])
                                        
                                    st.dataframe(
                                        df_errores_fatales_1p3p,
                                        use_container_width=True,
                                        height=220
                                    )
                                        
                                    st.caption(f"üîé Total de errores: {len(errores_validacion_1p3p)}")
                                    st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                    st.stop()
                                else:
                                    st.success("‚úÖ Validaciones de grados y secciones pasadas (1P-3P)")
                                
                                cursos_invalidos_1p3p = sorted(df_1p3p.loc[~df_1p3p["CURSO"].isin(st.session_state.cursos_equivalentes), "CURSO"].unique())
                                
                                if len(cursos_invalidos_1p3p) > 0 and st.session_state.archivo2_1p3p_df is None:
                                    st.warning(f"‚ö†Ô∏è Se detectaron {len(cursos_invalidos_1p3p)} cursos no reconocidos en 1P-3P")
                                    
                                    with st.form("equivalencias_form_1p3p"):
                                        st.markdown("### üîÑ Homologaci√≥n de Cursos (1P-3P)")
                                        st.info("Selecciona el curso oficial correspondiente para cada curso no reconocido:")
                                        
                                        equivalencias_1p3p = {}
                                        for curso in cursos_invalidos_1p3p:
                                            equivalencias_1p3p[curso] = st.selectbox(
                                                f"üìå **{curso}**",
                                                options=["-- Seleccionar --"] + st.session_state.cursos_equivalentes,
                                                key=f"eq_1p3p_{curso}"
                                            )
                                        
                                        submitted_1p3p = st.form_submit_button("‚úîÔ∏è Aplicar Equivalencias (1P-3P)", type="primary")
                                        
                                        if submitted_1p3p:
                                            if any(v == "-- Seleccionar --" for v in equivalencias_1p3p.values()):
                                                st.error("‚ùå Debes seleccionar un curso para todos los campos")
                                            else:
                                                for curso_err, curso_ok in equivalencias_1p3p.items():
                                                    df_1p3p.loc[df_1p3p["CURSO"] == curso_err, "CURSO"] = curso_ok
                                                
                                                df_1p3p["IDENTIFICADOR"] = crear_identificador(df_1p3p, "PATERNO", "MATERNO", "NOMBRES")
                                                cols_orden = [c for c in df_1p3p.columns if c != "IDENTIFICADOR"]
                                                cols_orden.append("IDENTIFICADOR")
                                                df_1p3p = df_1p3p[cols_orden]
                                                
                                                st.session_state.archivo2_1p3p_df = df_1p3p
                                                st.success("‚úÖ Cursos homologados correctamente en 1P-3P")
                                                st.rerun()
                                
                                else:
                                    if st.session_state.archivo2_1p3p_df is not None:
                                        df_1p3p = st.session_state.archivo2_1p3p_df
                                    else:
                                        df_1p3p["IDENTIFICADOR"] = crear_identificador(df_1p3p, "PATERNO", "MATERNO", "NOMBRES")
                                        cols_orden = [c for c in df_1p3p.columns if c != "IDENTIFICADOR"]
                                        cols_orden.append("IDENTIFICADOR")
                                        df_1p3p = df_1p3p[cols_orden]
                                        
                                        st.session_state.archivo2_1p3p_df = df_1p3p
                                    
                                    df_1p3p_procesado = df_1p3p
                                    df_vp_1p3p = df_1p3p.copy().drop(columns=["Nro."], errors="ignore")

                                    st.dataframe(df_vp_1p3p, use_container_width=True, hide_index=True)
                            
                        else:
                            st.error("‚ùå Error de cabecera en la hoja 1P-3P")
                            st.warning("‚ö†Ô∏è No se pudo detectar cabecera autom√°ticamente en 1P-3P")
                            st.info("Por favor, verifica que la hoja tenga las columnas correctas:")
                            st.code("NRO., PATERNO, MATERNO, NOMBRES, CURSO, GRADO, SECCI√ìN, NOTA VIGESIMAL 100%")
                            st.stop()
                    
                    # PROCESAR HOJA 4P-5S (Homologaci√≥n completa)
                    df_4p5s_procesado = None
                    df_vp_4p5s = None

                    if tiene_4p5s:
                        st.markdown("### üìó Procesando Hoja: 4P-5S")
                        
                        df_original2 = pd.read_excel(archivo2, sheet_name="4P-5S", header=None)
                        fila_detectada2 = detectar_cabecera_automatica(df_original2, COLUMNAS_ARCHIVO2_4P5S)
                        
                        if fila_detectada2 is not None:
                            st.session_state.archivo2_4p5s_fila_cabecera = fila_detectada2
                            st.success(f"‚úÖ Cabecera detectada en la fila {fila_detectada2 + 1}")
                            
                            df2 = pd.read_excel(archivo2, sheet_name="4P-5S", header=fila_detectada2)
                        
                            columnas_norm = {c.strip().lower(): c for c in df2.columns}
                            cols_a_usar = []
                            for col_req in COLUMNAS_ARCHIVO2_4P5S:
                                col_norm = col_req.strip().lower()
                                if col_norm in columnas_norm:
                                    cols_a_usar.append(columnas_norm[col_norm])
                            
                            df2 = df2[cols_a_usar]
                            df2.columns = [col.upper() for col in COLUMNAS_ARCHIVO2_4P5S]
                            df2 = limpiar_filas_vacias(df2, columnas_clave=["PATERNO", "MATERNO", "NOMBRES"])

                            if not df2.empty:
                                df2 = convertir_numericas_a_entero(df2, columnas=["GRADO", "NOTA VIGESIMAL 25%"])
                                df2 = homologar_dataframe(df2)
                                
                                errores_validacion_4p5s = []

                                df2, errores_grados = validar_y_mapear_grados(df2, "GRADO", tipo_validacion="4p5s")
                                errores_validacion_4p5s.extend(errores_grados)

                                errores_secciones = validar_secciones(df2, "SECCI√ìN")
                                errores_validacion_4p5s.extend(errores_secciones)

                                if errores_validacion_4p5s:
                                    st.error("‚ùå Errores de validaci√≥n en 4P-5S:")
                                    df_errores_fatales_4p5s = pd.DataFrame(errores_validacion_4p5s, columns=["Detalle de los errores cr√≠ticos"])
                                            
                                    st.dataframe(
                                        df_errores_fatales_4p5s,
                                        use_container_width=True,
                                        height=220
                                    )
                                            
                                    st.caption(f"üîé Total de errores: {len(errores_validacion_4p5s)}")
                                    st.info("Por favor, corrige estos errores en el archivo y vuelve a cargarlo")
                                    st.stop()

                                else:
                                    st.success("‚úÖ Validaciones de grados y secciones pasadas (4P-5S)")

                                if "NOTA VIGESIMAL 25%" in df2.columns:
                                    df2["NOTA VIGESIMAL 25%"] = df2["NOTA VIGESIMAL 25%"].fillna("NP").replace("", "NP")

                                columnas_oblig = ["PATERNO", "MATERNO", "NOMBRES", "CURSO", "GRADO", "SECCI√ìN", "NOTA VIGESIMAL 25%"]
                                filas_vacias = df2[df2[columnas_oblig].isnull().any(axis=1)]
                                
                                if not filas_vacias.empty:
                                    st.error("‚ùå Se detectaron campos vac√≠os")
                                    st.dataframe(filas_vacias, use_container_width=True)
                                    st.stop()
                                
                                cursos_invalidos = sorted(df2.loc[~df2["CURSO"].isin(st.session_state.cursos_equivalentes), "CURSO"].unique())
                                
                                if len(cursos_invalidos) > 0 and st.session_state.archivo2_4p5s_df is None:
                                    st.warning(f"‚ö†Ô∏è Se detectaron {len(cursos_invalidos)} cursos no reconocidos")
                                    
                                    with st.form("equivalencias_form"):
                                        st.markdown("### üîÑ Homologaci√≥n de Cursos")
                                        st.info("Selecciona el curso oficial correspondiente para cada curso no reconocido:")
                                        
                                        equivalencias = {}
                                        for curso in cursos_invalidos:
                                            equivalencias[curso] = st.selectbox(
                                                f"üìå **{curso}**",
                                                options=["-- Seleccionar --"] + st.session_state.cursos_equivalentes,
                                                key=f"eq_{curso}"
                                            )
                                        
                                        submitted = st.form_submit_button("‚úîÔ∏è Aplicar Equivalencias", type="primary")
                                        
                                        if submitted:
                                            if any(v == "-- Seleccionar --" for v in equivalencias.values()):
                                                st.error("‚ùå Debes seleccionar un curso para todos los campos")
                                            else:
                                                for curso_err, curso_ok in equivalencias.items():
                                                    df2.loc[df2["CURSO"] == curso_err, "CURSO"] = curso_ok
                                                
                                                df2["IDENTIFICADOR"] = crear_identificador(df2, "PATERNO", "MATERNO", "NOMBRES")
                                                df2["NOTAS VIGESIMALES 75%"] = ""
                                                df2["PROMEDIO"] = ""
                                                
                                                cols_orden = [c for c in df2.columns if c != "IDENTIFICADOR"]
                                                cols_orden.append("IDENTIFICADOR")
                                                df2 = df2[cols_orden]
                                                
                                                st.session_state.archivo2_4p5s_df = df2
                                                
                                                st.success("‚úÖ Cursos homologados correctamente")
                                                st.rerun()
                                else:
                                    if st.session_state.archivo2_4p5s_df is not None:
                                        df2 = st.session_state.archivo2_4p5s_df
                                    else:
                                        df2["IDENTIFICADOR"] = crear_identificador(df2, "PATERNO", "MATERNO", "NOMBRES")
                                        df2["NOTAS VIGESIMALES 75%"] = ""
                                        df2["PROMEDIO"] = ""
                                        
                                        cols_orden = [c for c in df2.columns if c != "IDENTIFICADOR"]
                                        cols_orden.append("IDENTIFICADOR")
                                        df2 = df2[cols_orden]
                                        
                                        st.session_state.archivo2_4p5s_df = df2
                                        
                                    df_4p5s_procesado = df2
                                    df_vp_4p5s = df2.copy().drop(columns=["Nro.", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore")
                                    
                                    st.dataframe(df_vp_4p5s, use_container_width=True, hide_index=True)
                        else:
                            st.error("‚ùå Error de cabecera en la hoja 4P-5S")
                            st.warning("‚ö†Ô∏è No se pudo detectar cabecera autom√°ticamente en 4P-5S")
                            st.info("Por favor, verifica que la hoja tenga las columnas correctas:")
                            st.code("NRO., PATERNO, MATERNO, NOMBRES, CURSO, GRADO, SECCI√ìN, NOTA VIGESIMAL 25%")
                            st.stop()

                    # DESCARGA
                    hoja_1p3p_lista = df_1p3p_procesado is not None and st.session_state.archivo2_1p3p_df is not None
                    hoja_4p5s_lista = df_4p5s_procesado is not None and st.session_state.archivo2_4p5s_df is not None

                    if hoja_1p3p_lista or hoja_4p5s_lista:
                        st.divider()
                        st.markdown("### üíæ Archivos Listos para Descargar")
                        
                        # SECCI√ìN 1P-3P
                        if hoja_1p3p_lista:
                            st.markdown("#### üìò Archivos 1P-3P")
                            
                            df_eval_1p3p_completo, _ = crear_archivo_evaluador(
                                st.session_state.archivo1_df,
                                df_1p3p_procesado
                            )
                            
                            dict_hojas_1p3p = {
                                "1P-3P": {
                                    'df': df_eval_1p3p_completo.drop(columns=["IDENTIFICADOR"], errors="ignore"),
                                    'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                }
                            }
                            
                            df_1p3p_actual = df_eval_1p3p_completo.copy()
                            
                            df_1p3p_observados = df_eval_1p3p_completo[
                                df_eval_1p3p_completo["OBSERVADOS"].isin(["RET", "SN"])
                            ].copy()
                            if len(df_1p3p_observados) > 0:
                                df_1p3p_observados = df_1p3p_observados.reset_index(drop=True)
                                if 'NRO.' in df_1p3p_observados.columns:
                                    df_1p3p_observados = df_1p3p_observados.drop(columns=['NRO.'])
                                df_1p3p_observados.insert(0, 'NRO.', range(1, len(df_1p3p_observados) + 1))
                            
                            df_1p3p_ok = df_eval_1p3p_completo[
                                (df_eval_1p3p_completo["OBSERVADOS"].isna()) | 
                                (df_eval_1p3p_completo["OBSERVADOS"] == "") |
                                (df_eval_1p3p_completo["OBSERVADOS"].astype(str).str.strip() == "")
                            ].copy()
                            if len(df_1p3p_ok) > 0:
                                df_1p3p_ok = df_1p3p_ok.reset_index(drop=True)
                                if 'NRO.' in df_1p3p_ok.columns:
                                    df_1p3p_ok = df_1p3p_ok.drop(columns=['NRO.'])
                                df_1p3p_ok.insert(0, 'NRO.', range(1, len(df_1p3p_ok) + 1))
                            
                            col_1p3p_0, col_1p3p_1, col_1p3p_2, col_1p3p_3 = st.columns(4)
                            
                            with col_1p3p_0:
                                # Archivo homologado
                                df_sin_notas_1p3p = df_1p3p_procesado.drop(columns=["IDENTIFICADOR"], errors="ignore")
                                df_sin_notas_1p3p["NOTA VIGESIMAL 100%"] = df_sin_notas_1p3p["NOTA VIGESIMAL 100%"].astype(str).replace('NAN', 'NP')
                                buffer_1p3p = guardar_con_formato_original(
                                    df_procesado=df_sin_notas_1p3p,
                                    archivo_original_bytes=st.session_state.archivo2_bytes,
                                    nombre_hoja="1P-3P",
                                    fila_cabecera=st.session_state.archivo2_1p3p_fila_cabecera,
                                    solo_hoja_especificada=True
                                )
                                st.download_button(
                                    label="üì• 1P-3P Homologado",
                                    data=buffer_1p3p,
                                    file_name=f"{st.session_state.nombre_colegio}_1P-3P_RV.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )

                            with col_1p3p_1:
                                # Actual 1P-3P
                                dict_actual_1p3p = {
                                    "1P-3P": {
                                        'df': df_1p3p_actual.drop(columns=["IDENTIFICADOR"], errors="ignore"),
                                        'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                    }
                                }
                                buffer_actual_1p3p = guardar_evaluador_con_multiples_hojas(
                                    archivo_original_bytes=st.session_state.archivo2_bytes,
                                    dict_hojas_procesadas=dict_actual_1p3p,
                                    solo_hojas_especificadas=True
                                )
                                st.download_button(
                                    label="üì• ACTUAL (1P-3P)",
                                    data=buffer_actual_1p3p,
                                    file_name=f"{st.session_state.nombre_colegio}_1P-3P_ACTUAL.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    help="Todas las filas de 1P-3P"
                                )
                            
                            with col_1p3p_2:
                                # Observados 1P-3P
                                if len(df_1p3p_observados) > 0:
                                    dict_observados_1p3p = {
                                        "1P-3P": {
                                            'df': df_1p3p_observados.drop(columns=["IDENTIFICADOR"], errors="ignore"),
                                            'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                        }
                                    }
                                    buffer_observados_1p3p = guardar_evaluador_con_multiples_hojas(
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        dict_hojas_procesadas=dict_observados_1p3p,
                                        solo_hojas_especificadas=True
                                    )
                                    st.download_button(
                                        label="üì• OBSERVADOS (1P-3P)",
                                        data=buffer_observados_1p3p,
                                        file_name=f"{st.session_state.nombre_colegio}_1P-3P_OBSERVADOS.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                        help="Solo filas con RET o SN"
                                    )
                                else:
                                    st.info("Sin observados en 1P-3P")
                            
                            with col_1p3p_3:
                                # OK 1P-3P
                                if len(df_1p3p_ok) > 0:
                                    
                                    df_1p3p_ok = df_1p3p_ok.reset_index(drop=True)
                                    df_1p3p_ok.columns = df_1p3p_ok.columns.str.strip()
                                    mapeo_columnas = {}

                                    for col in df_1p3p_ok.columns:
                                        col_upper = col.upper().strip()
                                        
                                        if col_upper == 'NRO.' or col_upper == 'NRO' or 'NRO' in col_upper and len(col_upper) <= 5:
                                            mapeo_columnas[col] = 'NRO.'
                                        elif col_upper == 'PATERNO':
                                            mapeo_columnas[col] = 'PATERNO'
                                        elif col_upper == 'MATERNO':
                                            mapeo_columnas[col] = 'MATERNO'
                                        elif col_upper == 'NOMBRES' or col_upper == 'NOMBRE':
                                            mapeo_columnas[col] = 'NOMBRE'
                                        elif col_upper == 'CURSO':
                                            mapeo_columnas[col] = 'CURSO'
                                        elif col_upper == 'GRADO':
                                            mapeo_columnas[col] = 'GRADO'
                                        elif col_upper == 'SECCI√ìN' or col_upper == 'SECCION':
                                            mapeo_columnas[col] = 'SECCI√ìN'
                                        elif col_upper == 'NOTA VIGESIMAL 100%':
                                            mapeo_columnas[col] = 'NOTA LABORATORIO'
                                    
                                    df_1p3p_ok = df_1p3p_ok.rename(columns=mapeo_columnas)
                                    
                                    columnas_a_eliminar = []
                                    for col in df_1p3p_ok.columns:
                                        col_upper = col.upper()
                                        if 'OBSERVADOS' in col_upper or 'OBSERVACION' in col_upper:
                                            columnas_a_eliminar.append(col)
                                    
                                    df_1p3p_ok = df_1p3p_ok.drop(columns=columnas_a_eliminar, errors='ignore')
                                    
                                    nuevas_columnas = [
                                        '¬øASISTI√ì?', 'P1 4PTOS.', 
                                        'P2 4PTOS.', 'P3 4PTOS.', 'P4 4PTOS.', 'P5 4PTOS.',
                                        'NOTA EVALUADOR', 'NOTA FINAL', 
                                        'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
                                    ]
                                    for col in nuevas_columnas:
                                        if col not in df_1p3p_ok.columns:
                                            df_1p3p_ok[col] = ''
                                    
                                    columnas_certificado_1p3p = [
                                        'NRO.', 'PATERNO', 'MATERNO', 'NOMBRE', 'GRADO', 'SECCI√ìN', 'CURSO', 
                                        'NOTA LABORATORIO', '¬øASISTI√ì?', 'P1 4PTOS.', 
                                        'P2 4PTOS.', 'P3 4PTOS.', 'P4 4PTOS.', 'P5 4PTOS.', 'NOTA EVALUADOR', 
                                        'NOTA FINAL', 'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
                                    ]
                                    columnas_existentes = [col for col in columnas_certificado_1p3p if col in df_1p3p_ok.columns]
                                    df_1p3p_ok = df_1p3p_ok[columnas_existentes]

                                    if "NOTA LABORATORIO" in df_1p3p_ok.columns and "NOTA FINAL" in df_1p3p_ok.columns:
                                        df_1p3p_ok["NOTA FINAL"] = pd.to_numeric(df_1p3p_ok["NOTA LABORATORIO"], errors="coerce")

                                    if "ESTATUS" in df_1p3p_ok.columns and "NOTA FINAL" in df_1p3p_ok.columns:
                                        nota_final = pd.to_numeric(df_1p3p_ok["NOTA FINAL"], errors="coerce")
                                        df_1p3p_ok["ESTATUS"] = nota_final.apply(
                                            lambda x: "Aprobado" if pd.notna(x) and x >= 12.5 else "Desaprobado"
                                        )

                                    if 'NRO.' in df_1p3p_ok.columns:
                                        df_1p3p_ok['NRO.'] = range(1, len(df_1p3p_ok) + 1)
                                    
                                    dict_ok_1p3p = {
                                        "1P-3P": {
                                            'df': df_1p3p_ok,
                                            'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                        }
                                    }
                                    buffer_ok_1p3p = guardar_certificado_con_encabezado(
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        dict_hojas_procesadas=dict_ok_1p3p
                                    )
                                    st.download_button(
                                        label="üì• OK (1P-3P)",
                                        data=buffer_ok_1p3p,
                                        file_name=f"{st.session_state.nombre_colegio}_1P-3P_OK.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                        help="Solo aprobados con formato certificado"
                                    )
                                else:
                                    st.info("Sin aprobados en 1P-3P")
                            
                            st.divider()
                        
                        # SECCI√ìN 4P-5S
                        if hoja_4p5s_lista:
                            st.markdown("#### üìó Archivos 4P-5S")
                            
                            _, df_eval_4p5s_completo = crear_archivo_evaluador(
                                st.session_state.archivo1_df,
                                df_4p5s_procesado
                            )
                            
                            df_4p5s_actual = df_eval_4p5s_completo.copy()
                            
                            df_4p5s_observados = df_eval_4p5s_completo[
                                df_eval_4p5s_completo["OBSERVADOS"].isin(["RET", "SN"])
                            ].copy()
                            if len(df_4p5s_observados) > 0:
                                df_4p5s_observados = df_4p5s_observados.reset_index(drop=True)
                                if 'NRO.' in df_4p5s_observados.columns:
                                    df_4p5s_observados = df_4p5s_observados.drop(columns=['NRO.'])
                                df_4p5s_observados.insert(0, 'NRO.', range(1, len(df_4p5s_observados) + 1))
                            
                            df_4p5s_ok = df_eval_4p5s_completo[
                                (df_eval_4p5s_completo["OBSERVADOS"].isna()) | 
                                (df_eval_4p5s_completo["OBSERVADOS"] == "") |
                                (df_eval_4p5s_completo["OBSERVADOS"].astype(str).str.strip() == "")
                            ].copy()
                            
                            if len(df_4p5s_ok) > 0:
                                df_4p5s_ok = df_4p5s_ok.reset_index(drop=True)
                                df_4p5s_ok.columns = df_4p5s_ok.columns.str.strip()
                                
                                mapeo_columnas = {}
                                for col in df_4p5s_ok.columns:
                                    col_upper = col.upper().strip()
                                    
                                    if col_upper == 'NRO.' or col_upper == 'NRO' or 'NRO' in col_upper and len(col_upper) <= 5:
                                        mapeo_columnas[col] = 'NRO.'
                                    elif col_upper == 'PATERNO':
                                        mapeo_columnas[col] = 'PATERNO'
                                    elif col_upper == 'MATERNO':
                                        mapeo_columnas[col] = 'MATERNO'
                                    elif col_upper == 'NOMBRES' or col_upper == 'NOMBRE':
                                        mapeo_columnas[col] = 'NOMBRE'
                                    elif col_upper == 'CURSO':
                                        mapeo_columnas[col] = 'CURSO'
                                    elif col_upper == 'GRADO':
                                        mapeo_columnas[col] = 'GRADO'
                                    elif col_upper == 'SECCI√ìN' or col_upper == 'SECCION':
                                        mapeo_columnas[col] = 'SECCI√ìN'
                                    elif col_upper == 'NOTA VIGESIMAL 25%':
                                        mapeo_columnas[col] = 'NOTA LABORATORIO'
                                
                                df_4p5s_ok = df_4p5s_ok.rename(columns=mapeo_columnas)
                                
                                columnas_a_eliminar = []
                                for col in df_4p5s_ok.columns:
                                    col_upper = col.upper()
                                    if 'PROMEDIO' in col_upper:
                                        columnas_a_eliminar.append(col)
                                    elif 'OBSERVADOS' in col_upper or 'OBSERVACION' in col_upper:
                                        columnas_a_eliminar.append(col)
                                
                                df_4p5s_ok = df_4p5s_ok.drop(columns=columnas_a_eliminar, errors='ignore')
                                
                                nuevas_columnas = [
                                    '¬øASISTI√ì?', 'P1 4PTOS.', 
                                    'P2 4PTOS.', 'P3 4PTOS.', 'P4 4PTOS.', 'P5 4PTOS.',
                                    'NOTA EVALUADOR', 'NOTA FINAL', 
                                    'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
                                ]
                                for col in nuevas_columnas:
                                    if col not in df_4p5s_ok.columns:
                                        df_4p5s_ok[col] = ''
                                
                                columnas_certificado = [
                                    'NRO.', 'PATERNO', 'MATERNO', 'NOMBRE', 'GRADO', 'SECCI√ìN', 'CURSO', 
                                    'NOTA LABORATORIO', '¬øASISTI√ì?', 'P1 4PTOS.', 
                                    'P2 4PTOS.', 'P3 4PTOS.', 'P4 4PTOS.', 'P5 4PTOS.', 'NOTA EVALUADOR', 
                                    'NOTA FINAL', 'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
                                ]
                                columnas_existentes = [col for col in columnas_certificado if col in df_4p5s_ok.columns]
                                df_4p5s_ok = df_4p5s_ok[columnas_existentes]

                                if 'NRO.' in df_4p5s_ok.columns:
                                    df_4p5s_ok['NRO.'] = range(1, len(df_4p5s_ok) + 1)
                            
                            col_1p3p_0, col_4p5s_1, col_4p5s_2, col_4p5s_3 = st.columns(4)
                            
                            with col_1p3p_0:
                                # Archivo homologado
                                df_sin_notas_4p5s = df_4p5s_procesado.drop(columns=["IDENTIFICADOR", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore")
                                df_sin_notas_4p5s["NOTA VIGESIMAL 25%"] = df_sin_notas_4p5s["NOTA VIGESIMAL 25%"].astype(str).replace('NAN', 'NP')
                                buffer_4p5s = guardar_con_formato_original(
                                    df_procesado=df_sin_notas_4p5s,
                                    archivo_original_bytes=st.session_state.archivo2_bytes,
                                    nombre_hoja="4P-5S",
                                    fila_cabecera=st.session_state.archivo2_4p5s_fila_cabecera,
                                    solo_hoja_especificada=True
                                )
                                st.download_button(
                                    label="üì• 4P-5S Homologado",
                                    data=buffer_4p5s,
                                    file_name=f"{st.session_state.nombre_colegio}_4P-5S_RV.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )

                            with col_4p5s_1:
                                # Actual 4P-5S
                                dict_actual_4p5s = {
                                    "4P-5S": {
                                        'df': df_4p5s_actual.drop(columns=["IDENTIFICADOR", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore"),
                                        'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                    }
                                }
                                buffer_actual_4p5s = guardar_evaluador_con_multiples_hojas(
                                    archivo_original_bytes=st.session_state.archivo2_bytes,
                                    dict_hojas_procesadas=dict_actual_4p5s,
                                    solo_hojas_especificadas=True
                                )
                                st.download_button(
                                    label="üì• ACTUAL (4P-5S)",
                                    data=buffer_actual_4p5s,
                                    file_name=f"{st.session_state.nombre_colegio}_4P-5S_ACTUAL.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    help="Todas las filas de 4P-5S"
                                )
                            
                            with col_4p5s_2:
                                # Observados 4P-5S
                                if len(df_4p5s_observados) > 0:
                                    dict_observados_4p5s = {
                                        "4P-5S": {
                                            'df': df_4p5s_observados.drop(columns=["IDENTIFICADOR", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore"),
                                            'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                        }
                                    }
                                    buffer_observados_4p5s = guardar_evaluador_con_multiples_hojas(
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        dict_hojas_procesadas=dict_observados_4p5s,
                                        solo_hojas_especificadas=True
                                    )
                                    st.download_button(
                                        label="üì• OBSERVADOS (4P-5S)",
                                        data=buffer_observados_4p5s,
                                        file_name=f"{st.session_state.nombre_colegio}_4P-5S_OBSERVADOS.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                        help="Solo filas con RET o SN"
                                    )
                                else:
                                    st.info("Sin observados en 4P-5S")
                            
                            with col_4p5s_3:
                                # OK 4P-5S
                                if len(df_4p5s_ok) > 0:
                                    dict_ok_4p5s = {
                                        "4P-5S": {
                                            'df': df_4p5s_ok,
                                            'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                        }
                                    }
                                    buffer_ok_4p5s = guardar_certificado_con_encabezado(
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        dict_hojas_procesadas=dict_ok_4p5s
                                    )
                                    st.download_button(
                                        label="üì• OK (4P-5S)",
                                        data=buffer_ok_4p5s,
                                        file_name=f"{st.session_state.nombre_colegio}_4P-5S_OK.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                        help="Solo aprobados con formato certificado"
                                    )
                                else:
                                    st.info("Sin aprobados en 4P-5S")
                            
                            st.divider()
                        
                        # DESCARGA
                        st.markdown("#### üì¶ Descarga Completa")
                        st.caption("Descarga todos los archivos procesados en un solo ZIP")
                        
                        col_zip1, col_zip2 = st.columns(2)
                        
                        with col_zip1:
                            # ZIP de 1P-3P
                            if hoja_1p3p_lista:
                                from io import BytesIO
                                from zipfile import ZipFile
                                
                                zip_1p3p_buffer = BytesIO()
                                with ZipFile(zip_1p3p_buffer, 'w') as zip_file:
                                    # 1. Homologado
                                    df_sin_notas_1p3p = df_1p3p_procesado.drop(columns=["IDENTIFICADOR"], errors="ignore")
                                    df_sin_notas_1p3p["NOTA VIGESIMAL 100%"] = df_sin_notas_1p3p["NOTA VIGESIMAL 100%"].astype(str).replace('NAN', 'NP')
                                    buffer_homologado_1p3p = guardar_con_formato_original(
                                        df_procesado=df_sin_notas_1p3p,
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        nombre_hoja="1P-3P",
                                        fila_cabecera=st.session_state.archivo2_1p3p_fila_cabecera,
                                        solo_hoja_especificada=True
                                    )
                                    zip_file.writestr(f"{st.session_state.nombre_colegio}_1P-3P_RV.xlsx", buffer_homologado_1p3p.getvalue())
                                    
                                    # 2. ACTUAL
                                    dict_actual_1p3p = {
                                        "1P-3P": {
                                            'df': df_1p3p_actual.drop(columns=["IDENTIFICADOR"], errors="ignore"),
                                            'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                        }
                                    }
                                    buffer_actual_1p3p = guardar_evaluador_con_multiples_hojas(
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        dict_hojas_procesadas=dict_actual_1p3p,
                                        solo_hojas_especificadas=True
                                    )
                                    zip_file.writestr(f"{st.session_state.nombre_colegio}_1P-3P_ACTUAL.xlsx", buffer_actual_1p3p.getvalue())
                                    
                                    # 3. OBSERVADOS (si existen)
                                    if len(df_1p3p_observados) > 0:
                                        dict_observados_1p3p = {
                                            "1P-3P": {
                                                'df': df_1p3p_observados.drop(columns=["IDENTIFICADOR"], errors="ignore"),
                                                'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                            }
                                        }
                                        buffer_observados_1p3p = guardar_evaluador_con_multiples_hojas(
                                            archivo_original_bytes=st.session_state.archivo2_bytes,
                                            dict_hojas_procesadas=dict_observados_1p3p,
                                            solo_hojas_especificadas=True
                                        )
                                        zip_file.writestr(f"{st.session_state.nombre_colegio}_1P-3P_OBSERVADOS.xlsx", buffer_observados_1p3p.getvalue())
                                    
                                    # 4. OK (si existen)
                                    if len(df_1p3p_ok) > 0:
                                        df_1p3p_ok_zip = df_1p3p_ok.copy()
                                        df_1p3p_ok_zip = df_1p3p_ok_zip.reset_index(drop=True)
                                        df_1p3p_ok_zip.columns = df_1p3p_ok_zip.columns.str.strip()
                                        
                                        mapeo_columnas = {}
                                        for col in df_1p3p_ok_zip.columns:
                                            col_upper = col.upper().strip()
                                            
                                            if col_upper == 'NRO.' or col_upper == 'NRO' or 'NRO' in col_upper and len(col_upper) <= 5:
                                                mapeo_columnas[col] = 'NRO.'
                                            elif col_upper == 'PATERNO':
                                                mapeo_columnas[col] = 'PATERNO'
                                            elif col_upper == 'MATERNO':
                                                mapeo_columnas[col] = 'MATERNO'
                                            elif col_upper == 'NOMBRES' or col_upper == 'NOMBRE':
                                                mapeo_columnas[col] = 'NOMBRE'
                                            elif col_upper == 'CURSO':
                                                mapeo_columnas[col] = 'CURSO'
                                            elif col_upper == 'GRADO':
                                                mapeo_columnas[col] = 'GRADO'
                                            elif col_upper == 'SECCI√ìN' or col_upper == 'SECCION':
                                                mapeo_columnas[col] = 'SECCI√ìN'
                                            elif col_upper == 'NOTA VIGESIMAL 100%':
                                                mapeo_columnas[col] = 'NOTA LABORATORIO'
                                        
                                        df_1p3p_ok_zip = df_1p3p_ok_zip.rename(columns=mapeo_columnas)
                                        
                                        columnas_a_eliminar = []
                                        for col in df_1p3p_ok_zip.columns:
                                            col_upper = col.upper()
                                            if 'OBSERVADOS' in col_upper or 'OBSERVACION' in col_upper:
                                                columnas_a_eliminar.append(col)
                                        
                                        df_1p3p_ok_zip = df_1p3p_ok_zip.drop(columns=columnas_a_eliminar, errors='ignore')
                                        
                                        nuevas_columnas = [
                                            '¬øASISTI√ì?', 'P1 4PTOS.', 
                                            'P2 4PTOS.', 'P3 4PTOS.', 'P4 4PTOS.', 'P5 4PTOS.',
                                            'NOTA EVALUADOR', 'NOTA FINAL', 
                                            'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
                                        ]
                                        for col in nuevas_columnas:
                                            if col not in df_1p3p_ok_zip.columns:
                                                df_1p3p_ok_zip[col] = ''
                                        
                                        columnas_certificado_1p3p = [
                                            'NRO.', 'PATERNO', 'MATERNO', 'NOMBRE', 'GRADO', 'SECCI√ìN', 'CURSO', 
                                            'NOTA LABORATORIO', '¬øASISTI√ì?', 'P1 4PTOS.', 
                                            'P2 4PTOS.', 'P3 4PTOS.', 'P4 4PTOS.', 'P5 4PTOS.', 'NOTA EVALUADOR', 
                                            'NOTA FINAL', 'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
                                        ]
                                        columnas_existentes = [col for col in columnas_certificado_1p3p if col in df_1p3p_ok_zip.columns]
                                        df_1p3p_ok_zip = df_1p3p_ok_zip[columnas_existentes]

                                        if "NOTA LABORATORIO" in df_1p3p_ok_zip.columns and "NOTA FINAL" in df_1p3p_ok_zip.columns:
                                            df_1p3p_ok_zip["NOTA FINAL"] = pd.to_numeric(df_1p3p_ok_zip["NOTA LABORATORIO"], errors="coerce")

                                        if "ESTATUS" in df_1p3p_ok_zip.columns and "NOTA FINAL" in df_1p3p_ok_zip.columns:
                                            nota_final = pd.to_numeric(df_1p3p_ok_zip["NOTA FINAL"], errors="coerce")
                                            df_1p3p_ok_zip["ESTATUS"] = nota_final.apply(
                                                lambda x: "Aprobado" if pd.notna(x) and x >= 12.5 else "Desaprobado"
                                            )

                                        if 'NRO.' in df_1p3p_ok_zip.columns:
                                            df_1p3p_ok_zip['NRO.'] = range(1, len(df_1p3p_ok_zip) + 1)
                                        
                                        dict_ok_1p3p = {
                                            "1P-3P": {
                                                'df': df_1p3p_ok_zip,
                                                'fila_cabecera': st.session_state.archivo2_1p3p_fila_cabecera
                                            }
                                        }
                                        buffer_ok_1p3p = guardar_certificado_con_encabezado(
                                            archivo_original_bytes=st.session_state.archivo2_bytes,
                                            dict_hojas_procesadas=dict_ok_1p3p
                                        )
                                        zip_file.writestr(f"{st.session_state.nombre_colegio}_1P-3P_OK.xlsx", buffer_ok_1p3p.getvalue())
                                
                                zip_1p3p_buffer.seek(0)
                                
                                archivos_1p3p = 2
                                if len(df_1p3p_observados) > 0:
                                    archivos_1p3p += 1
                                if len(df_1p3p_ok) > 0:
                                    archivos_1p3p += 1
                                
                                st.download_button(
                                    label=f"üì¶ Descargar Todo 1P-3P ({archivos_1p3p} archivos)",
                                    data=zip_1p3p_buffer,
                                    file_name=f"{st.session_state.nombre_colegio}_1P-3P_COMPLETO.zip",
                                    mime="application/zip",
                                    use_container_width=True,
                                    help=f"Descarga {archivos_1p3p} archivos Excel en un ZIP"
                                )
                            else:
                                st.info("1P-3P no procesado")
                        
                        with col_zip2:
                            # ZIP de 4P-5S
                            if hoja_4p5s_lista:
                                from io import BytesIO
                                from zipfile import ZipFile
                                
                                zip_4p5s_buffer = BytesIO()
                                with ZipFile(zip_4p5s_buffer, 'w') as zip_file:
                                    # 1. Homologado
                                    df_sin_notas_4p5s = df_4p5s_procesado.drop(columns=["IDENTIFICADOR", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore")
                                    df_sin_notas_4p5s["NOTA VIGESIMAL 25%"] = df_sin_notas_4p5s["NOTA VIGESIMAL 25%"].astype(str).replace('NAN', 'NP')
                                    buffer_homologado_4p5s = guardar_con_formato_original(
                                        df_procesado=df_sin_notas_4p5s,
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        nombre_hoja="4P-5S",
                                        fila_cabecera=st.session_state.archivo2_4p5s_fila_cabecera,
                                        solo_hoja_especificada=True
                                    )
                                    zip_file.writestr(f"{st.session_state.nombre_colegio}_4P-5S_RV.xlsx", buffer_homologado_4p5s.getvalue())
                                    
                                    # 2. ACTUAL
                                    dict_actual_4p5s = {
                                        "4P-5S": {
                                            'df': df_4p5s_actual.drop(columns=["IDENTIFICADOR", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore"),
                                            'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                        }
                                    }
                                    buffer_actual_4p5s = guardar_evaluador_con_multiples_hojas(
                                        archivo_original_bytes=st.session_state.archivo2_bytes,
                                        dict_hojas_procesadas=dict_actual_4p5s,
                                        solo_hojas_especificadas=True
                                    )
                                    zip_file.writestr(f"{st.session_state.nombre_colegio}_4P-5S_ACTUAL.xlsx", buffer_actual_4p5s.getvalue())
                                    
                                    # 3. OBSERVADOS (si existen)
                                    if len(df_4p5s_observados) > 0:
                                        dict_observados_4p5s = {
                                            "4P-5S": {
                                                'df': df_4p5s_observados.drop(columns=["IDENTIFICADOR", "NOTAS VIGESIMALES 75%", "PROMEDIO"], errors="ignore"),
                                                'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                            }
                                        }
                                        buffer_observados_4p5s = guardar_evaluador_con_multiples_hojas(
                                            archivo_original_bytes=st.session_state.archivo2_bytes,
                                            dict_hojas_procesadas=dict_observados_4p5s,
                                            solo_hojas_especificadas=True
                                        )
                                        zip_file.writestr(f"{st.session_state.nombre_colegio}_4P-5S_OBSERVADOS.xlsx", buffer_observados_4p5s.getvalue())
                                    
                                    # 4. OK (si existen)
                                    if len(df_4p5s_ok) > 0:
                                        df_4p5s_ok_zip = df_4p5s_ok.copy()
                                        
                                        dict_ok_4p5s = {
                                            "4P-5S": {
                                                'df': df_4p5s_ok_zip,
                                                'fila_cabecera': st.session_state.archivo2_4p5s_fila_cabecera
                                            }
                                        }
                                        buffer_ok_4p5s = guardar_certificado_con_encabezado(
                                            archivo_original_bytes=st.session_state.archivo2_bytes,
                                            dict_hojas_procesadas=dict_ok_4p5s
                                        )
                                        zip_file.writestr(f"{st.session_state.nombre_colegio}_4P-5S_OK.xlsx", buffer_ok_4p5s.getvalue())
                                
                                zip_4p5s_buffer.seek(0)
                                
                                archivos_4p5s = 2  # Homologado + ACTUAL
                                if len(df_4p5s_observados) > 0:
                                    archivos_4p5s += 1
                                if len(df_4p5s_ok) > 0:
                                    archivos_4p5s += 1
                                
                                st.download_button(
                                    label=f"üì¶ Descargar Todo 4P-5S ({archivos_4p5s} archivos)",
                                    data=zip_4p5s_buffer,
                                    file_name=f"{st.session_state.nombre_colegio}_4P-5S_COMPLETO.zip",
                                    mime="application/zip",
                                    use_container_width=True,
                                    help=f"Descarga {archivos_4p5s} archivos Excel en un ZIP"
                                )
                            else:
                                st.info("4P-5S no procesado")
                        
                        st.divider()

                        col1, col2, col3 = st.columns([1, 1, 2])
                        with col1:
                            if st.button("‚úÖ Finalizar Proceso", type="primary", use_container_width=True):
                                st.session_state.paso_actual = 3
                                st.rerun()

                    else:
                        st.warning("‚ö†Ô∏è Completa el procesamiento de al menos una hoja para descargar archivos")
                    
                except Exception as e:
                    st.error(f"‚ùå Error: {e}")

    # FINALIZACI√ìN
    elif st.session_state.paso_actual == 3:

        st.markdown("""
        <div style='background-color: #78808C; padding: 30px; border-radius: 15px; text-align: center;'>
            <h1>üéâ ¬°Proceso Completado!</h1>
            <p style='font-size: 18px;'>Todos los archivos han sido procesados y homologados correctamente.</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### üìã Resumen del Proceso")
            st.success(f"**Colegio:** {st.session_state.nombre_colegio}")
            st.success(f"**Archivo 1:** {len(st.session_state.archivo1_df)} registros")
            
            if st.session_state.archivo2_1p3p_df is not None:
                st.success(f"**Hoja 1P-3P:** {len(st.session_state.archivo2_1p3p_df)} registros")
            if st.session_state.archivo2_4p5s_df is not None:
                st.success(f"**Hoja 4P-5S:** {len(st.session_state.archivo2_4p5s_df)} registros")
        
        with col2:
            st.markdown("### üîÑ Acciones")
            if st.button("üÜï Procesar Nuevo Colegio", type="primary", use_container_width=True):
                st.session_state.paso_actual = 0
                st.session_state.nombre_colegio = ""
                st.session_state.archivo1_df = None
                st.session_state.archivo2_df = None
                st.session_state.archivo1_bytes = None
                st.session_state.archivo2_bytes = None
                st.session_state.archivo1_fila_cabecera = None
                st.session_state.archivo2_1p3p_fila_cabecera = None
                st.session_state.archivo2_4p5s_fila_cabecera = None
                st.rerun()
            
            if st.button("üîô Volver al Paso 3", use_container_width=True):
                st.session_state.paso_actual = 2
                st.rerun()

# TAB 2: COMPARADOR DE EVALUADORES
with tab2:
    st.markdown("## ‚öñÔ∏è Validador Evaluaciones de Alumnos")
    st.markdown("""
                **DESCRIPCI√ìN:**
                
                Permite validar que el archivo devuelto por el evaluador corresponda exactamente al archivo originalmente enviado, verificando que no se hayan realizado modificaciones al contenido y que la √∫nica informaci√≥n agregada sean las notas u observaciones del proceso de evaluaci√≥n.
                """)
    st.info("""
    üìå **INSTRUCCIONES**
    1.	Formato tipo "{NombreColegio}_4P-5S_OK.xlsx"
    2.	Sube el archivo OK (puede tener campos vac√≠os en: NOTA EVALUADOR, P1-P5 4PTOS., NOTA FINAL)
    3.	Sube el archivo OK_EVALUADOR (debe tener completos: NOTA EVALUADOR y NOTA FINAL)
            
    üìå **VALIDACIONES AUTOM√ÅTICAS**
    1.	NOTA EVALUADOR debe ser la suma de P1 + P2 + P3 + P4 + P5
    2.	NOTA FINAL debe ser: (NOTA LABORATORIO * 0.25) + (NOTA EVALUADOR * 0.75)
    3.	El orden de los alumnos, cursos y notas de laboratorio enviados inicialmente al evaluador deben coincidir exactamente, sino saldr√° una advertencia

    """)
    
    COLUMNAS_CERTIFICADO = [
        'NRO.', 'PATERNO', 'MATERNO', 'NOMBRE', 'GRADO', 'SECCI√ìN', 'CURSO',
        'NOTA LABORATORIO', '¬øASISTI√ì?', 'P1 4PTOS.', 'P2 4PTOS.', 'P3 4PTOS.',
        'P4 4PTOS.', 'P5 4PTOS.', 'NOTA EVALUADOR', 'NOTA FINAL',
        'OBSERVADOS', 'ESTATUS', 'NUMERACI√ìN'
    ]
    
    def leer_archivo_certificado(archivo_bytes, nombre_hoja=None, es_ok_evaluador=False):
        """
        Lee un archivo certificado Excel y retorna DataFrame validado.
        Retorna las FILAS COMPLETAS que tienen errores (sin detallar cada error).
        
        Args:
            archivo_bytes: bytes del archivo Excel
            nombre_hoja: nombre de la hoja a leer
            es_ok_evaluador: True si es archivo OK_EVALUADOR (valida campos vac√≠os), 
                           False si es archivo OK (permite campos vac√≠os)
        
        Returns:
            tuple: (df, error, fila_cabecera, hojas, df_filas_con_errores)
        """
        try:
            wb = load_workbook(BytesIO(archivo_bytes), data_only=True)
            
            if nombre_hoja is None:
                nombre_hoja = wb.sheetnames[0]
            
            if nombre_hoja not in wb.sheetnames:
                return None, f"La hoja '{nombre_hoja}' no existe", None, None, None
            
            ws = wb[nombre_hoja]
            
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            df = pd.DataFrame(data)

            if len(df.columns) > 19:
                df = df.iloc[:, :19]
            
            # Detectar cabecera (debe estar en fila 8 para formato certificado)
            fila_cabecera = None
            for idx in range(min(15, len(df))):
                fila = df.iloc[idx].astype(str).str.strip().str.upper().tolist()
                if 'NOTA LABORATORIO' in fila and 'NOTA EVALUADOR' in fila:
                    fila_cabecera = idx
                    break
            
            if fila_cabecera is None:
                return None, "No se detect√≥ la cabecera del formato certificado", None, None, None
            
            df.columns = df.iloc[fila_cabecera]
            df = df.iloc[fila_cabecera + 1:].reset_index(drop=True)
            df.columns = df.columns.astype(str).str.strip().str.upper()
            
            columnas_faltantes = []
            for col_req in ['PATERNO', 'MATERNO', 'NOMBRE', 'GRADO', 'SECCI√ìN', 'CURSO', 
                           'NOTA LABORATORIO', 'P1 4PTOS.', 'P2 4PTOS.', 'P3 4PTOS.',
                           'P4 4PTOS.', 'P5 4PTOS.', 'NOTA EVALUADOR', 'NOTA FINAL']:
                if col_req not in df.columns:
                    columnas_faltantes.append(col_req)
            
            if columnas_faltantes:
                return None, f"Columnas faltantes: {', '.join(columnas_faltantes)}", None, None, None
            
            df = df.dropna(how='all')
            df = df[df[['PATERNO', 'MATERNO', 'NOMBRE']].notna().all(axis=1)]
            
            if df.empty:
                return None, "No hay datos v√°lidos en el archivo", None, None, None
            
            # Validaci√≥n simple
            columnas_numericas = ['NOTA LABORATORIO', 'P1 4PTOS.', 'P2 4PTOS.', 'P3 4PTOS.',
                                 'P4 4PTOS.', 'P5 4PTOS.', 'NOTA EVALUADOR', 'NOTA FINAL']
            
            filas_con_errores_indices = set()  
            errores_por_fila = {} 
            
            for col in columnas_numericas:
                if col not in df.columns:
                    continue
                
                for idx, valor in df[col].items():
                    valor_str = str(valor).strip().upper()
                    es_vacio = valor_str in ["", "NAN", "NONE", "NAT", "NULL"] or pd.isna(valor)
                    
                    if idx not in errores_por_fila:
                        errores_por_fila[idx] = {
                            'NOTA_EVALUADOR_VACIA': False,
                            'NOTA_FINAL_VACIA': False,
                            'VALOR_NEGATIVO': False,
                            'VALOR_MAYOR_20': False,
                            'P1_P5_MAYOR_4': False,
                            'VALOR_NO_NUMERICO': False
                        }
                    
                    if es_ok_evaluador:
                        if col == 'NOTA EVALUADOR' and es_vacio:
                            filas_con_errores_indices.add(idx)
                            errores_por_fila[idx]['NOTA_EVALUADOR_VACIA'] = True
                            continue
                        
                        if col == 'NOTA FINAL' and es_vacio:
                            filas_con_errores_indices.add(idx)
                            errores_por_fila[idx]['NOTA_FINAL_VACIA'] = True
                            continue
                    
                    if es_vacio:
                        continue
                    
                    try:
                        valor_num = float(valor_str)
                        
                        if valor_num < 0:
                            filas_con_errores_indices.add(idx)
                            errores_por_fila[idx]['VALOR_NEGATIVO'] = True
                        
                        if valor_num > 20:
                            filas_con_errores_indices.add(idx)
                            errores_por_fila[idx]['VALOR_MAYOR_20'] = True
                        
                        if col.startswith('P') and col.endswith('4PTOS.') and valor_num > 4:
                            filas_con_errores_indices.add(idx)
                            errores_por_fila[idx]['P1_P5_MAYOR_4'] = True
                            
                    except ValueError:
                        filas_con_errores_indices.add(idx)
                        errores_por_fila[idx]['VALOR_NO_NUMERICO'] = True
            
            if filas_con_errores_indices:
                df_filas_con_errores = df.loc[list(filas_con_errores_indices)].copy()
                
                for idx in df_filas_con_errores.index:
                    df_filas_con_errores.loc[idx, 'ERROR: NOTA EVALUADOR VAC√çA'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_EVALUADOR_VACIA', False) else ''
                    df_filas_con_errores.loc[idx, 'ERROR: NOTA FINAL VAC√çA'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_FINAL_VACIA', False) else ''
                    df_filas_con_errores.loc[idx, 'ERROR: VALOR NEGATIVO'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('VALOR_NEGATIVO', False) else ''
                    df_filas_con_errores.loc[idx, 'ERROR: VALOR MAYOR A 20'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('VALOR_MAYOR_20', False) else ''
                    df_filas_con_errores.loc[idx, 'ERROR: P1-P5 MAYOR A 4'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('P1_P5_MAYOR_4', False) else ''
                    df_filas_con_errores.loc[idx, 'ERROR: VALOR NO NUM√âRICO'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('VALOR_NO_NUMERICO', False) else ''
                
                df_filas_con_errores = df_filas_con_errores.sort_index()
                return None, None, fila_cabecera, wb.sheetnames, df_filas_con_errores
            
            return df, None, fila_cabecera, wb.sheetnames, None
            
        except Exception as e:
            return None, f"Error al leer archivo: {str(e)}", None, None, None
    
    def comparar_certificados(df_base, df_revisar):
        """
        Compara dos archivos certificados.
        Retorna DataFrame con las FILAS COMPLETAS que tienen errores.
        
        Returns:
            DataFrame con filas que tienen errores o None si no hay errores
        """
        df_base.columns = df_base.columns.str.strip().str.upper()
        df_revisar.columns = df_revisar.columns.str.strip().str.upper()
        
        filas_con_errores_indices = set()
        errores_por_fila = {}
        
        if len(df_base) != len(df_revisar):
            st.error(f"‚ùå Los archivos tienen diferente n√∫mero de filas: BASE={len(df_base)}, REVISAR={len(df_revisar)}")
            return None
        
        columnas_identidad = ['PATERNO', 'MATERNO', 'NOMBRE', 'GRADO', 'SECCI√ìN', 'CURSO', 'NOTA LABORATORIO']
        
        for idx in range(len(df_base)):
            if idx not in errores_por_fila:
                errores_por_fila[idx] = {
                    'NOMBRES_NO_COINCIDEN': False,
                    'GRADO_NO_COINCIDE': False,
                    'SECCION_NO_COINCIDE': False,
                    'CURSO_NO_COINCIDE': False,
                    'NOTA_LAB_NO_COINCIDE': False,
                    'NOTA_EVALUADOR_VACIA': False,
                    'NOTA_FINAL_VACIA': False,
                    'ERROR_SUMA_P1_P5': False,
                    'ERROR_CALCULO_NOTA_FINAL': False
                }
            
            for col in columnas_identidad:
                if col not in df_base.columns or col not in df_revisar.columns:
                    continue
                
                val_base = str(df_base.loc[idx, col]).strip().upper()
                val_revisar = str(df_revisar.loc[idx, col]).strip().upper()
                
                if val_base in ["", "NAN", "NONE"]:
                    val_base = ""
                if val_revisar in ["", "NAN", "NONE"]:
                    val_revisar = ""
                
                if val_base != val_revisar:
                    filas_con_errores_indices.add(idx)
                    
                    if col in ['PATERNO', 'MATERNO', 'NOMBRE']:
                        errores_por_fila[idx]['NOMBRES_NO_COINCIDEN'] = True
                    elif col == 'GRADO':
                        errores_por_fila[idx]['GRADO_NO_COINCIDE'] = True
                    elif col == 'SECCI√ìN':
                        errores_por_fila[idx]['SECCION_NO_COINCIDE'] = True
                    elif col == 'CURSO':
                        errores_por_fila[idx]['CURSO_NO_COINCIDE'] = True
                    elif col == 'NOTA LABORATORIO':
                        errores_por_fila[idx]['NOTA_LAB_NO_COINCIDE'] = True
        
        if filas_con_errores_indices:
            df_errores = df_revisar.loc[list(filas_con_errores_indices)].copy()
            
            for idx in df_errores.index:
                df_errores.loc[idx, 'ERROR: NOMBRES NO COINCIDEN'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOMBRES_NO_COINCIDEN', False) else ''
                df_errores.loc[idx, 'ERROR: GRADO NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('GRADO_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: SECCI√ìN NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('SECCION_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: CURSO NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('CURSO_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: NOTA LAB NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_LAB_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: NOTA EVALUADOR VAC√çA'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_EVALUADOR_VACIA', False) else ''
                df_errores.loc[idx, 'ERROR: NOTA FINAL VAC√çA'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_FINAL_VACIA', False) else ''
                df_errores.loc[idx, 'ERROR: SUMA P1-P5'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('ERROR_SUMA_P1_P5', False) else ''
                df_errores.loc[idx, 'ERROR: C√ÅLCULO NOTA FINAL'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('ERROR_CALCULO_NOTA_FINAL', False) else ''
            
            df_errores = df_errores.sort_index()
            return df_errores
        
        for idx in range(len(df_revisar)):
            if idx not in errores_por_fila:
                errores_por_fila[idx] = {
                    'NOMBRES_NO_COINCIDEN': False,
                    'GRADO_NO_COINCIDE': False,
                    'SECCION_NO_COINCIDE': False,
                    'CURSO_NO_COINCIDE': False,
                    'NOTA_LAB_NO_COINCIDE': False,
                    'NOTA_EVALUADOR_VACIA': False,
                    'NOTA_FINAL_VACIA': False,
                    'ERROR_SUMA_P1_P5': False,
                    'ERROR_CALCULO_NOTA_FINAL': False
                }
            
            # NOTA EVALUADOR debe estar completa
            val_evaluador = str(df_revisar.loc[idx, 'NOTA EVALUADOR']).strip().upper()
            es_vacio_evaluador = (
                val_evaluador in ["", "NAN", "NONE", "NAT", "NULL"] or 
                pd.isna(df_revisar.loc[idx, 'NOTA EVALUADOR']) or
                val_evaluador.replace(" ", "") == ""
            )
            
            if es_vacio_evaluador:
                filas_con_errores_indices.add(idx)
                errores_por_fila[idx]['NOTA_EVALUADOR_VACIA'] = True
            
            # NOTA FINAL debe estar completa
            val_final = str(df_revisar.loc[idx, 'NOTA FINAL']).strip().upper()
            es_vacio_final = (
                val_final in ["", "NAN", "NONE", "NAT", "NULL"] or 
                pd.isna(df_revisar.loc[idx, 'NOTA FINAL']) or
                val_final.replace(" ", "") == ""
            )
            
            if es_vacio_final:
                filas_con_errores_indices.add(idx)
                errores_por_fila[idx]['NOTA_FINAL_VACIA'] = True
        
        for idx in range(len(df_revisar)):
            if idx not in errores_por_fila:
                errores_por_fila[idx] = {
                    'NOMBRES_NO_COINCIDEN': False,
                    'GRADO_NO_COINCIDE': False,
                    'SECCION_NO_COINCIDE': False,
                    'CURSO_NO_COINCIDE': False,
                    'NOTA_LAB_NO_COINCIDE': False,
                    'NOTA_EVALUADOR_VACIA': False,
                    'NOTA_FINAL_VACIA': False,
                    'ERROR_SUMA_P1_P5': False,
                    'ERROR_CALCULO_NOTA_FINAL': False
                }
            
            try:
                p1 = pd.to_numeric(df_revisar.loc[idx, 'P1 4PTOS.'], errors='coerce')
                p2 = pd.to_numeric(df_revisar.loc[idx, 'P2 4PTOS.'], errors='coerce')
                p3 = pd.to_numeric(df_revisar.loc[idx, 'P3 4PTOS.'], errors='coerce')
                p4 = pd.to_numeric(df_revisar.loc[idx, 'P4 4PTOS.'], errors='coerce')
                p5 = pd.to_numeric(df_revisar.loc[idx, 'P5 4PTOS.'], errors='coerce')
                nota_evaluador = pd.to_numeric(df_revisar.loc[idx, 'NOTA EVALUADOR'], errors='coerce')
                nota_lab = pd.to_numeric(df_revisar.loc[idx, 'NOTA LABORATORIO'], errors='coerce')
                nota_final = pd.to_numeric(df_revisar.loc[idx, 'NOTA FINAL'], errors='coerce')
                
                # Reemplazar NaN con 0 para P1-P5
                p1 = 0 if pd.isna(p1) else p1
                p2 = 0 if pd.isna(p2) else p2
                p3 = 0 if pd.isna(p3) else p3
                p4 = 0 if pd.isna(p4) else p4
                p5 = 0 if pd.isna(p5) else p5
                
                # VALIDACI√ìN: NOTA EVALUADOR = P1 + P2 + P3 + P4 + P5
                if not pd.isna(nota_evaluador):
                    suma_esperada = p1 + p2 + p3 + p4 + p5
                    diferencia = abs(nota_evaluador - suma_esperada)
                    
                    if diferencia > 0.01:
                        filas_con_errores_indices.add(idx)
                        errores_por_fila[idx]['ERROR_SUMA_P1_P5'] = True
                
                # VALIDACI√ìN: NOTA FINAL = (NOTA LAB * 0.25) + (NOTA EVAL * 0.75)
                if not pd.isna(nota_final) and not pd.isna(nota_lab) and not pd.isna(nota_evaluador):
                    nota_final_esperada = (nota_lab * 0.25) + (nota_evaluador * 0.75)
                    diferencia = abs(nota_final - nota_final_esperada)
                    
                    if diferencia > 0.01:
                        filas_con_errores_indices.add(idx)
                        errores_por_fila[idx]['ERROR_CALCULO_NOTA_FINAL'] = True
                        
            except Exception as e:
                filas_con_errores_indices.add(idx)
        
        if filas_con_errores_indices:
            df_errores = df_revisar.loc[list(filas_con_errores_indices)].copy()
            
            for idx in df_errores.index:
                df_errores.loc[idx, 'ERROR: NOMBRES NO COINCIDEN'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOMBRES_NO_COINCIDEN', False) else ''
                df_errores.loc[idx, 'ERROR: GRADO NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('GRADO_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: SECCI√ìN NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('SECCION_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: CURSO NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('CURSO_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: NOTA LAB NO COINCIDE'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_LAB_NO_COINCIDE', False) else ''
                df_errores.loc[idx, 'ERROR: NOTA EVALUADOR VAC√çA'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_EVALUADOR_VACIA', False) else ''
                df_errores.loc[idx, 'ERROR: NOTA FINAL VAC√çA'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('NOTA_FINAL_VACIA', False) else ''
                df_errores.loc[idx, 'ERROR: SUMA P1-P5'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('ERROR_SUMA_P1_P5', False) else ''
                df_errores.loc[idx, 'ERROR: C√ÅLCULO NOTA FINAL'] = 'REVISAR' if errores_por_fila.get(idx, {}).get('ERROR_CALCULO_NOTA_FINAL', False) else ''
            
            df_errores = df_errores.sort_index()
            return df_errores
        
        return None
    
    col_izq, col_der = st.columns(2)
    
    # COLUMNA IZQUIERDA: Archivo Base
    with col_izq:
        st.markdown("#### üìÑ Archivo: {NombreColegio}_4P-5S_OK")
        st.caption("Puede tener campos vac√≠os en: NOTA EVALUADOR, P1-P5, NOTA FINAL")
        
        archivo_base = st.file_uploader(
            "Selecciona el archivo certificado OK",
            type=["xlsx"],
            key=f"uploader_base_cert_{st.session_state.comparador_reset_counter}"
        )
        
        if archivo_base:
            archivo_base_bytes = archivo_base.read()
            archivo_base.seek(0)
            
            xls_base = pd.ExcelFile(archivo_base)
            hojas_base = xls_base.sheet_names
            
            hoja_base_seleccionada = st.selectbox(
                "Selecciona la hoja OK:",
                hojas_base,
                key="selector_hoja_base_cert"
            )
            
            if st.button("‚úÖ Cargar Archivo OK", key="btn_cargar_base_cert"):
                df_base, error_base, fila_cab_base, _, df_err = leer_archivo_certificado(
                    archivo_base_bytes,
                    hoja_base_seleccionada,
                    es_ok_evaluador=False 
                )
                
                if error_base:
                    st.error(f"‚ùå {error_base}")
                elif df_err is not None:
                    st.error(f"‚ùå Se encontraron {len(df_err)} filas con errores de validaci√≥n")
                    st.dataframe(df_err, use_container_width=True, hide_index=True)
                else:
                    st.session_state.comparador_archivo_base = {
                        'df': df_base,
                        'nombre_hoja': hoja_base_seleccionada,
                        'fila_cabecera': fila_cab_base
                    }
                    st.session_state.comparador_resultados = None
                    st.session_state.comparador_comparacion_realizada = False
                    st.success(f"‚úÖ Archivo OK cargado ({len(df_base)} registros)")
            
            if st.session_state.comparador_archivo_base:
                with st.expander("Vista previa - OK", expanded=False):
                    st.dataframe(st.session_state.comparador_archivo_base['df'].head(10), hide_index=True)
        else:
            st.info("‚¨ÜÔ∏è Por favor, sube el archivo OK para continuar")
            if st.session_state.comparador_archivo_base is not None:
                st.session_state.comparador_archivo_base = None
                st.session_state.comparador_resultados = None
                st.session_state.comparador_comparacion_realizada = False
    
    # COLUMNA DERECHA: Archivo a Revisar
    with col_der:
        st.markdown("#### üîç Archivo: {NombreColegio}_4P-5S_OK_EVALUADOR")
        st.caption("Debe tener completos: NOTA EVALUADOR y NOTA FINAL")
        
        archivo_revisar = st.file_uploader(
            "Selecciona el archivo certificado OK_EVALUADOR",
            type=["xlsx"],
            key=f"uploader_revisar_cert_{st.session_state.comparador_reset_counter}"
        )

        if archivo_revisar is not None:
            nombre_archivo_evaluador = archivo_revisar.name
            patron_esperado_evaluador = f"_4P-5S_OK_EVALUADOR.xlsx"

            if not nombre_archivo_evaluador.endswith(patron_esperado_evaluador):
                st.error(f"‚ùå Formato de archivo incorrecto")
                st.warning(f"‚ö†Ô∏è El archivo debe terminar en: `{patron_esperado_evaluador}`")
                st.info(f"üìù Ejemplo correcto: `Colegio{patron_esperado_evaluador}`")
                st.info(f"üìù Tu archivo: `{nombre_archivo_evaluador}`")
                st.stop()
            
            nombre_colegio_evaluador = nombre_archivo_evaluador.replace(patron_esperado_evaluador, "")

            if not nombre_colegio_evaluador or nombre_colegio_evaluador.strip() == "":
                st.error("‚ùå No se pudo extraer el nombre del colegio del archivo")
                st.info(f"Archivo recibido: `{nombre_archivo_evaluador}`")
                st.stop()
            
            archivo_revisar_bytes = archivo_revisar.read()
            archivo_revisar.seek(0)
            
            xls_revisar = pd.ExcelFile(archivo_revisar)
            hojas_revisar = xls_revisar.sheet_names
            
            hoja_revisar_seleccionada = st.selectbox(
                "Selecciona la hoja OK_EVALUADOR:",
                hojas_revisar,
                key="selector_hoja_revisar_cert"
            )
            
            if st.button("‚úÖ Cargar Archivo OK_EVALUADOR", key="btn_cargar_revisar_cert"):
                df_revisar, error_revisar, fila_cab_revisar, _, df_err = leer_archivo_certificado(
                    archivo_revisar_bytes,
                    hoja_revisar_seleccionada,
                    es_ok_evaluador=True 
                )
                
                if error_revisar:
                    st.error(f"‚ùå {error_revisar}")
                elif df_err is not None:
                    st.error(f"‚ùå Se encontraron {len(df_err)} filas con errores de validaci√≥n")
                    st.dataframe(df_err, use_container_width=True, hide_index=True)
                    
                    st.divider()
                    csv = df_err.to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        label="üì• Descargar filas con errores (CSV)",
                        data=csv,
                        file_name="filas_con_errores.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                else:
                    st.session_state.comparador_archivo_revisar = {
                        'df': df_revisar,
                        'nombre_hoja': hoja_revisar_seleccionada,
                        'fila_cabecera': fila_cab_revisar,
                        'bytes': archivo_revisar_bytes
                    }
                    st.session_state.comparador_resultados = None
                    st.session_state.comparador_comparacion_realizada = False
                    st.success(f"‚úÖ Archivo OK_EVALUADOR cargado ({len(df_revisar)} registros)")
            
            if st.session_state.comparador_archivo_revisar:
                with st.expander("Vista previa - OK_EVALUADOR", expanded=False):
                    st.dataframe(st.session_state.comparador_archivo_revisar['df'].head(10), hide_index=True)

        else:
            st.info("‚¨ÜÔ∏è Por favor, sube el archivo OK_EVALUADOR para continuar")
            if st.session_state.comparador_archivo_revisar is not None:
                st.session_state.comparador_archivo_revisar = None
                st.session_state.comparador_resultados = None
                st.session_state.comparador_comparacion_realizada = False
    
    st.divider()
    
    if st.session_state.comparador_archivo_base and st.session_state.comparador_archivo_revisar:
        col_comp1, col_comp2, col_comp3 = st.columns([1, 2, 1])
        
        with col_comp2:
            if st.button("üîç COMPARAR ARCHIVOS", type="primary", use_container_width=True):
                with st.spinner("Comparando archivos y validando c√°lculos..."):
                    df_errores = comparar_certificados(
                        st.session_state.comparador_archivo_base['df'].copy(),
                        st.session_state.comparador_archivo_revisar['df'].copy()
                    )
                    st.session_state.comparador_resultados = df_errores
                    st.session_state.comparador_comparacion_realizada = True
        
        # MOSTRAR RESULTADOS
        if st.session_state.comparador_resultados is not None:
            st.divider()
            
            if isinstance(st.session_state.comparador_resultados, pd.DataFrame) and len(st.session_state.comparador_resultados) > 0:
                df_errores = st.session_state.comparador_resultados
                st.error("‚ùå **SE ENCONTRARON ERRORES**")
                st.warning(f"‚ö†Ô∏è Total de filas con errores: **{len(df_errores)}**")
                
                st.markdown("### üìã Filas Completas con Errores")
                st.caption("Estas son las filas que presentan uno o m√°s errores. Revisa cada valor.")
                
                st.dataframe(
                    df_errores,
                    use_container_width=True,
                    height=400
                )
                
                st.divider()
                col_desc1, col_desc2, col_desc3 = st.columns([1, 1, 1])
                with col_desc2:
                    csv = df_errores.to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        label="üì• Descargar Filas con Errores (CSV)",
                        data=csv,
                        file_name="filas_con_errores_completas.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
        
        elif st.session_state.get('comparador_comparacion_realizada', False):
            st.divider()
            st.success("üéâ **¬°VALIDACI√ìN EXITOSA!**")
            st.success("‚úÖ Todos los datos y c√°lculos son correctos")
            
            df_final = st.session_state.comparador_archivo_revisar['df'].copy()
                
            if "ESTATUS" not in df_final.columns:
                df_final["ESTATUS"] = ""
                
            if "NOTA FINAL" in df_final.columns:
                nota_final = pd.to_numeric(df_final["NOTA FINAL"], errors="coerce")
                df_final["ESTATUS"] = nota_final.apply(
                    lambda x: "Aprobado" if pd.notna(x) and x >= 12.5 else "Desaprobado"
                )
                
                total = len(df_final)
                aprobados = (df_final["ESTATUS"] == "Aprobado").sum()
                desaprobados = (df_final["ESTATUS"] == "Desaprobado").sum()
                    
                st.divider()
                st.markdown("### üìä Resumen de Resultados")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total de Estudiantes", total)
                with col2:
                    st.metric("‚úÖ Aprobados", aprobados, delta=f"{aprobados/total*100:.1f}%")
                with col3:
                    st.metric("‚ùå Desaprobados", desaprobados, delta=f"{desaprobados/total*100:.1f}%")
                
            archivo_bytes_original = st.session_state.comparador_archivo_revisar['bytes']

            wb_original = load_workbook(BytesIO(archivo_bytes_original))
            nombre_hoja = st.session_state.comparador_archivo_revisar['nombre_hoja']
            ws_original = wb_original[nombre_hoja]

            fila_cabecera = st.session_state.comparador_archivo_revisar.get('fila_cabecera', 7) 

            col_estatus_idx = None
            for col_idx, cell in enumerate(ws_original[fila_cabecera + 1], start=1):
                if cell.value and str(cell.value).strip().upper() == "ESTATUS":
                    col_estatus_idx = col_idx
                    break

            if col_estatus_idx is None:
                max_col = ws_original.max_column
                for col_idx in range(1, max_col + 1):
                    cell = ws_original.cell(row=fila_cabecera + 1, column=col_idx)
                    if cell.value is None:
                        col_estatus_idx = col_idx
                        break
                else:
                    col_estatus_idx = max_col + 1
                
                ws_original.cell(row=fila_cabecera + 1, column=col_estatus_idx, value="ESTATUS")

            start_row = fila_cabecera + 2
            df_datos = st.session_state.comparador_archivo_revisar['df'].copy()

            estatus_values = []
            for idx in range(len(df_datos)):
                try:
                    if "NOTA FINAL" in df_datos.columns:
                        nota_final = df_datos.loc[idx, "NOTA FINAL"]
                        try:
                            nota_num = float(str(nota_final).strip())
                            estatus = "Aprobado" if nota_num >= 12.5 else "Desaprobado"
                        except:
                            estatus = ""
                    else:
                        estatus = ""
                except:
                    estatus = ""
                estatus_values.append(estatus)

            for i, estatus in enumerate(estatus_values):
                row_idx = start_row + i
                if row_idx <= ws_original.max_row:
                    ws_original.cell(row=row_idx, column=col_estatus_idx, value=estatus)

            output = BytesIO()
            wb_original.save(output)
            excel_data = output.getvalue()

            col_desc1, col_desc2, col_desc3 = st.columns([1, 1, 1])
            with col_desc2:
                st.download_button(
                    label="üì• Descargar {NombreColegio}_OK_EVALUADOR_REV.xlsx",
                    data=excel_data,
                    file_name=f"{nombre_colegio_evaluador}_4P-5S_OK_EVALUADOR_REV.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="Descarga el archivo original con la columna ESTATUS completada seg√∫n la nota final"
                )

    else:
        st.info("üëÜ Carga ambos archivos para comenzar la comparaci√≥n")
    
    st.divider()
    if st.button("üîÑ Limpiar y Nueva Comparaci√≥n", key="btn_reset_comparador_cert"):
        st.session_state.comparador_archivo_base = None
        st.session_state.comparador_archivo_revisar = None
        st.session_state.comparador_resultados = None
        st.session_state.comparador_comparacion_realizada = False
        st.session_state.comparador_reset_counter += 1
        st.rerun()

# TAB 3: Generar Reporte PDF
with tab3:
    st.markdown("## üìë Generador de Planilla de Resultados")
    st.markdown("""
                **DESCRIPCI√ìN:**
                
                Permite generar reportes de manera autom√°tica a partir de la informaci√≥n registrada, organizados por grado, secci√≥n y curso. El proceso crea archivos PDF independientes para cada nivel de corte, los cuales pueden descargarse de manera consolidada en un archivo ZIP, facilitando la revisi√≥n, distribuci√≥n y trazabilidad de la informaci√≥n.
                """)
    st.info("""
    üìå **INSTRUCCIONES**
    1.	Sube un archivo OK con formato:
            
        ‚Ä¢	{NombreColegio}_1P-3P_OK.xlsx
            
        ‚Ä¢	{NombreColegio}_4P-5S_OK_EVALUADOR_REV.xlsx
            
    2.	Se generar√°n PDFs agrupados por: Grado ‚Üí Secci√≥n ‚Üí Curso
    3.	Cada PDF contendr√° la lista completa de estudiantes con sus notas

    üìå **VALIDACIONES AUTOM√ÅTICAS**
    1.	Las columnas PATERNO, MATERNO, NOMBRE, GRADO, SECCI√ìN, CURSO y NOTA FINAL deben estar completas (sin valores vac√≠os)
    """)
    
    tipo_reporte = st.radio(
        "Selecciona el tipo de archivo homologado:",
        ["1P-3P", "4P-5S"],
        horizontal=True,
        key="radio_tipo_reporte"
    )
    
    archivo_reporte = st.file_uploader(
        f"Selecciona el archivo homologado {tipo_reporte}",
        type=["xlsx"],
        key=f"uploader_reporte_{st.session_state.tab3_reset_counter}"
    )
    
    if archivo_reporte:
        nombre_archivo = archivo_reporte.name
        
        if tipo_reporte == "1P-3P":
            patron_esperado = f"_{tipo_reporte}_OK.xlsx"
        else: # "4P-5S"
            patron_esperado = f"_{tipo_reporte}_OK_EVALUADOR_REV.xlsx"
        
        if not nombre_archivo.endswith(patron_esperado):
            st.error(f"‚ùå Formato de archivo incorrecto")
            st.warning(f"‚ö†Ô∏è El archivo debe terminar en: `{patron_esperado}`")
            st.info(f"üìù Ejemplo correcto: `Colegio{patron_esperado}`")
            st.info(f"üìù Tu archivo: `{nombre_archivo}`")
            st.stop()
        
        nombre_colegio_reporte = nombre_archivo.replace(patron_esperado, "")
        
        if not nombre_colegio_reporte or nombre_colegio_reporte.strip() == "":
            st.error("‚ùå No se pudo extraer el nombre del colegio del archivo")
            st.info(f"Archivo recibido: `{nombre_archivo}`")
            st.stop()
        
        st.success(f"üè´ Colegio detectado: **{nombre_colegio_reporte}**")
        
        with st.spinner("üìä Procesando y validando archivo..."):
            try:
                df_temp = pd.read_excel(archivo_reporte, header=None)
                fila_cabecera = detectar_cabecera_automatica(df_temp, COLUMNAS_TAB03)
                
                if fila_cabecera is None:
                    st.error("‚ùå No se pudo detectar la cabecera autom√°ticamente")
                    st.info("Columnas esperadas: NRO., PATERNO, MATERNO, NOMBRE, GRADO, SECCI√ìN, CURSO, NOTA LABORATORIO, ¬øASISTI√ì?, P1 4PTOS., P2 4PTOS., P3 4PTOS., P4 4PTOS., P5 4PTOS., NOTA EVALUADOR, NOTA FINAL, OBSERVADOS, ESTATUS, NUMERACI√ìN")
                    st.stop()
                
                df_reporte = pd.read_excel(archivo_reporte, header=fila_cabecera)
                
                columnas_norm = {c.strip().lower(): c for c in df_reporte.columns}
                cols_requeridas = ["nro.", "paterno", "materno", "nombre", "curso", "grado", "secci√≥n", "nota final"]
                
                cols_a_usar = []
                for col_req in cols_requeridas:
                    col_norm = col_req.strip().lower()
                    if col_norm in columnas_norm:
                        cols_a_usar.append(columnas_norm[col_norm])
                    else:
                        st.error(f"‚ùå Columna no encontrada: '{col_req}'")
                        st.info(f"Columnas disponibles: {list(df_reporte.columns)}")
                        st.stop()
                
                df_reporte = df_reporte[cols_a_usar]
                df_reporte.columns = [
                    "NRO.", "PATERNO", "MATERNO", "NOMBRE", "CURSO", 
                    "GRADO", "SECCI√ìN", "NOTA FINAL"
                ]
                df_reporte = limpiar_filas_vacias(df_reporte, columnas_clave=["PATERNO", "MATERNO", "NOMBRE"])
                
                if df_reporte.empty:
                    st.error("‚ùå No hay datos v√°lidos despu√©s de limpiar filas vac√≠as")
                    st.stop()

                df_reporte = df_reporte.rename(columns={"NOMBRE": "NOMBRES"})
                
                st.markdown("### üîç Validando campos obligatorios...")
                
                columnas_obligatorias = ["PATERNO", "MATERNO", "NOMBRES", "GRADO", "SECCI√ìN", "CURSO", "NOTA FINAL"]
                errores_validacion = []
                
                for col in columnas_obligatorias:
                    if col not in df_reporte.columns:
                        errores_validacion.append(f"Columna '{col}' no encontrada")
                        continue
                    
                    vacios = df_reporte[col].isna() | (df_reporte[col].astype(str).str.strip() == "")
                    num_vacios = vacios.sum()
                    
                    if num_vacios > 0:
                        indices_vacios = df_reporte[vacios].index.tolist()
                        filas_vacias = [idx + fila_cabecera + 2 for idx in indices_vacios]
                        
                        errores_validacion.append({
                            'columna': col,
                            'num_vacios': num_vacios,
                            'filas': filas_vacias[:10]
                        })
                
                if errores_validacion:
                    st.error("‚ùå **VALIDACI√ìN FALLIDA: Existen campos obligatorios vac√≠os**")
                    st.warning("‚ö†Ô∏è Todas las columnas obligatorias deben estar completas antes de generar los reportes PDF")
                    
                    st.markdown("---")
                    st.markdown("### üìã Detalle de Errores Encontrados")
                    
                    for error in errores_validacion:
                        if isinstance(error, dict):
                            st.error(f"**Columna '{error['columna']}'**: {error['num_vacios']} valor(es) vac√≠o(s)")
                            
                            if len(error['filas']) > 0:
                                filas_texto = ", ".join(map(str, error['filas']))
                                if error['num_vacios'] > 10:
                                    st.info(f"üìç Filas afectadas (primeras 10): {filas_texto}... y {error['num_vacios'] - 10} m√°s")
                                else:
                                    st.info(f"üìç Filas afectadas: {filas_texto}")
                            
                            df_problematico = df_reporte[df_reporte[error['columna']].isna() | 
                                                         (df_reporte[error['columna']].astype(str).str.strip() == "")]
                            
                            if not df_problematico.empty:
                                st.markdown(f"**Registros con '{error['columna']}' vac√≠o:**")
                                st.dataframe(df_problematico.head(10), hide_index=True)
                        else:
                            st.error(error)
                    
                    st.markdown("---")
                    st.info("üí° **Soluci√≥n:** Corrige los valores vac√≠os en el archivo Excel y vuelve a subirlo")
                    st.stop()
                
                st.success("‚úÖ **Todas las validaciones pasaron correctamente**")
                st.success(f"‚úÖ Archivo cargado: {len(df_reporte)} registros")
                st.success(f"üìç Cabecera detectada en fila {fila_cabecera + 1}")
                
                df_reporte = homologar_dataframe(df_reporte)
                
                st.session_state.tab3_df_reporte = df_reporte
                st.session_state.tab3_nombre_colegio = nombre_colegio_reporte
                st.session_state.tab3_tipo_archivo = tipo_reporte
                st.session_state.tab3_archivo_procesado = True
                
                st.markdown("---")
                st.markdown("### üìä Vista previa de datos")
                st.dataframe(df_reporte, hide_index=True)
                
                st.markdown("---")
                st.markdown("### üìä Agrupaci√≥n de Datos")
                
                grupos_reportes = df_reporte.groupby(['GRADO', 'SECCI√ìN', 'CURSO'])
                num_grupos = len(grupos_reportes)
                
                col_info1, col_info2, col_info3 = st.columns(3)
                with col_info1:
                    st.metric("Grados", df_reporte['GRADO'].nunique())
                with col_info2:
                    st.metric("Secciones", df_reporte['SECCI√ìN'].nunique())
                with col_info3:
                    st.metric("Reportes a generar", num_grupos)
                
                with st.expander("üìã Ver detalle de grupos", expanded=True):
                    grupos_info = []
                    for (grado, seccion, curso), grupo_df in grupos_reportes:
                        grupos_info.append({
                            'Grado': grado,
                            'Secci√≥n': seccion,
                            'Curso': curso,
                            'Estudiantes': len(grupo_df)
                        })
                    st.dataframe(pd.DataFrame(grupos_info), hide_index=True)
                
                st.markdown("---")
                
                if st.button("üéØ GENERAR REPORTES PDF", type="primary", use_container_width=True):
                    generar_reportes_pdf(
                        df_reporte, 
                        nombre_colegio_reporte, 
                        tipo_reporte
                    )
                
            except Exception as e:
                st.error(f"‚ùå Error al procesar archivo: {str(e)}")
                import traceback
                with st.expander("üîç Ver error detallado"):
                    st.code(traceback.format_exc())
    
    else:
        st.info("üëÜ Sube un archivo para comenzar")
    
    st.markdown("---")
    if st.button("üîÑ Limpiar y empezar de nuevo", use_container_width=True, key="btn_reset_tab3"):
        st.session_state.tab3_archivo_procesado = False
        st.session_state.tab3_df_reporte = None
        st.session_state.tab3_nombre_colegio = ""
        st.session_state.tab3_tipo_archivo = ""
        st.session_state.tab3_reset_counter += 1 
        st.rerun()

# TAB 4: Generador de Certificados
with tab4:
    st.markdown("## üéì Generador de Diplomas, Certificados y Constancias")
    st.markdown("""
                **DESCRIPCI√ìN:**
                
                Permite emitir los diplomas, certificados, certificados progresivos y constancias manera autom√°tica a partir de la informaci√≥n registrada de diplomados, certificados y constancias, seg√∫n el archivo cargado. El proceso crea archivos PDF independientes para cada alumno, los cuales pueden descargarse de manera consolidada en un archivo ZIP, facilitando la revisi√≥n, distribuci√≥n y trazabilidad de la informaci√≥n.
                """)
    st.info("""
    üìå **INSTRUCCIONES:**
    1.	Sube un archivo OK con formato:
            
        ‚Ä¢	{NombreColegio}_1P-3P_OK.xlsx
            
        ‚Ä¢	{NombreColegio}_4P-5S_OK_EVALUADOR_REV.xlsx
            
    2.	Selecciona el tipo de certificado que a generar
    3.	Elige si deseas incluir marca de agua en los certificados
    4.	Se generar√°n archivos comprimidos con todos los certificados correspondientes
            
    ‚ö†Ô∏è **IMPORTANTE**
    1.	**Columnas base requeridas:** NRO., PATERNO, MATERNO, NOMBRE, GRADO, SECCI√ìN, CURSO, NOTA LABORATORIO, ¬øASISTI√ì?, P1 4PTOS., P2 4PTOS., P3 4PTOS., P4 4PTOS., P5 4PTOS., NOTA EVALUADOR, NOTA FINAL, OBSERVADOS, ESTATUS, NUMERACI√ìN
    2.	**Columna HORAS PROGRESIVO:** Solo es REQUERIDA cuando se selecciona el tipo "Progresivo". Para certificados normales, esta columna es opcional.
    3.	La columna "NOTA FINAL" debe estar completa (sin valores vac√≠os)
    """)

    if 'archivo_procesado' not in st.session_state:
        st.session_state.archivo_procesado = False

    st.markdown("### üì§ Subir y procesar archivo Excel")

    if 'tipo_certificado_seleccionado' not in st.session_state:
        st.session_state.tipo_certificado_seleccionado = None
    if 'usar_marca_agua_seleccionado' not in st.session_state:
        st.session_state.usar_marca_agua_seleccionado = False

    uploaded_file = st.file_uploader(
        "Selecciona un archivo Excel", 
        type=["xlsx"],
        key=f"tab4_uploader_{st.session_state.tab4_reset_counter}"
    )

    if uploaded_file:
        st.markdown("### ‚öôÔ∏è Configuraci√≥n de certificados")
        
        deshabilitar_opciones = st.session_state.certificados_generados

        col1, col2 = st.columns(2)
        
        with col1:
            opciones_display = {
                "Regulares (diplomados, certificados y constancias)": "Regular",
                "Progresivo": "Progresivo"
            }
            
            valor_actual = st.session_state.tipo_certificado_seleccionado
            if valor_actual:
                display_actual = [k for k, v in opciones_display.items() if v == valor_actual][0]
            else:
                display_actual = list(opciones_display.keys())[0]
            
            tipo_certificado_display = st.selectbox(
                "üìã Seleccionar la plantilla autom√°tica",
                options=list(opciones_display.keys()),
                index=list(opciones_display.keys()).index(display_actual) if valor_actual else 0,
                help="**Regulares (diplomados, certificados y constancias):** Se genera seg√∫n nota de aprobaci√≥n (‚â•12.5 = aprobado, <12.5 = participaci√≥n)\n\n**Progresivo:** Todos los estudiantes reciben certificado.",
                key="select_tipo_certificado",
                disabled=deshabilitar_opciones
            )
            st.session_state.tipo_certificado_seleccionado = opciones_display[tipo_certificado_display]
        
        with col2:
            st.markdown("""
                <style>
                    div[data-testid="stCheckbox"] {
                        padding-top: 1.3rem;
                    }
                </style>
            """, unsafe_allow_html=True)

            usar_marca_agua = st.checkbox(
                "Incluir marca de agua",
                value=st.session_state.usar_marca_agua_seleccionado,
                help="Agrega marca de agua 'PRELIMINAR' a los certificados generados",
                key="check_marca_agua",
                disabled=deshabilitar_opciones
            )
            st.session_state.usar_marca_agua_seleccionado = usar_marca_agua
        
        st.markdown("### üìÖ Fecha de los certificados")
        col_fecha1, col_fecha2 = st.columns([2, 1])
        
        with col_fecha1:
            fecha_seleccionada = st.date_input(
                "Selecciona la fecha de emisi√≥n",
                value=st.session_state.fecha_certificado_seleccionada,
                help="Esta fecha aparecer√° en todos los certificados generados",
                key="date_certificado",
                disabled=deshabilitar_opciones,
                format="DD/MM/YYYY"
            )
            st.session_state.fecha_certificado_seleccionada = fecha_seleccionada
        
        with col_fecha2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("üîÑ Usar fecha de hoy", disabled=deshabilitar_opciones, use_container_width=True):
                st.session_state.fecha_certificado_seleccionada = datetime.now().date()
                st.rerun()
        
        from datetime import date
        fecha_para_certificado = st.session_state.fecha_certificado_seleccionada
        if isinstance(fecha_para_certificado, date):
            fecha_dt = datetime.combine(fecha_para_certificado, datetime.min.time())
        else:
            fecha_dt = fecha_para_certificado
        fecha_formateada = mes_en_espanol(fecha_dt)
        st.info(f"üìÑ Los certificados mostrar√°n: **Lima, {fecha_formateada}**")
        
        tipo_certificado_actual = st.session_state.tipo_certificado_seleccionado
        if tipo_certificado_actual == "Regular":
            st.info("""
                    ‚ÑπÔ∏è **Modo Regulares (diplomados, certificados y constancias):** 
                    - Se generar√°n certificados de aprobaci√≥n (nota ‚â•12.5) o participaci√≥n (nota <12.5) seg√∫n corresponda.
                    - Los estudiantes de 1P-3P y 4P-5S usar√°n plantillas diferentes.
                    """)
        else:
            st.success("‚ÑπÔ∏è **Modo Progresivo:** Todos los estudiantes recibir√°n certificados progresivos")
        
        if usar_marca_agua:
            st.warning("‚ö†Ô∏è Los certificados incluir√°n la marca de agua 'PRELIMINAR'")
        
        st.markdown("---")

    if uploaded_file and not st.session_state.archivo_procesado:
        if st.button("üöÄ Procesar archivo y generar certificados", type="primary", use_container_width=True, key="btn_procesar_certificados"):
                with st.spinner("Validando y procesando archivo..."):
                    try:
                        uploaded_file.seek(0)
                        df_temp = pd.read_excel(uploaded_file, header=None)
                        fila_encabezado = detectar_fila_encabezado(df_temp)
                        
                        if fila_encabezado is None:
                            st.warning("‚ö†Ô∏è No se pudo detectar autom√°ticamente la fila del encabezado. Usando fila 7 por defecto.")
                            fila_encabezado = 8
                        else:
                            st.info(f"üìç Encabezado detectado en la fila {fila_encabezado + 1}")
                        
                        uploaded_file.seek(0)
                        df_usuario = pd.read_excel(uploaded_file, header=fila_encabezado)

                    except Exception as e:
                        st.error(f"‚ùå Error al leer el archivo: {str(e)}")
                        st.stop()
                    
                    tipo_certificado_para_validar = st.session_state.get('tipo_certificado_seleccionado', 'Regular')
                    df_formateado, exito_mapeo, mensaje_mapeo = validar_y_mapear_columnas(df_usuario, tipo_certificado_para_validar)
                    
                    if not exito_mapeo:
                        st.error(mensaje_mapeo)
                        
                        if tipo_certificado_para_validar == "Progresivo":
                            st.info(""" 
                                    El archivo de Excel debe contener exactamente estas columnas:
                                    - NRO., PATERNO, MATERNO, NOMBRE, GRADO, SECCI√ìN, CURSO, NOTA LABORATORIO, ¬øASISTI√ì?, P1 4PTOS., P2 4PTOS., P3 4PTOS., P4 4PTOS., P5 4PTOS., NOTA EVALUADOR, NOTA FINAL, OBSERVADOS, ESTATUS, NUMERACI√ìN, HORAS PROGRESIVO
                                    
                                    ‚ö†Ô∏è Nota: La columna 'HORAS PROGRESIVO' es obligatoria para certificados Progresivos.
                                    """)
                        else:
                            st.info(""" 
                                    El archivo de Excel debe contener estas columnas base:
                                    - NRO., PATERNO, MATERNO, NOMBRE, GRADO, SECCI√ìN, CURSO, NOTA LABORATORIO, ¬øASISTI√ì?, P1 4PTOS., P2 4PTOS., P3 4PTOS., P4 4PTOS., P5 4PTOS., NOTA EVALUADOR, NOTA FINAL, OBSERVADOS, ESTATUS, NUMERACI√ìN
                                    
                                    ‚ÑπÔ∏è Nota: La columna 'HORAS PROGRESIVO' no es necesaria para certificados Normales.
                                    """)
                        st.stop()
                    
                    st.success(mensaje_mapeo)
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_formateado.to_excel(writer, index=False, sheet_name='Sheet1')
                    output.seek(0)
                    
                    df_procesado, exito, mensaje = procesar_excel_inicial(output)
                    
                    if exito:
                        st.session_state.df_procesado = df_procesado
                        st.session_state.nombre_archivo = uploaded_file.name
                        st.session_state.grupos = None
                        st.session_state.plantillas = None
                        st.session_state.certificados_generados = False
                        st.session_state.zip_buffer = None
                        st.success(mensaje)
                        st.subheader("‚úÖ Archivo procesado - Vista previa de datos limpios")
                        st.write(f"**Dimensiones procesadas:** {df_procesado.shape[0]} filas x {df_procesado.shape[1]} columnas")
                        
                        df_mostrar = df_procesado.copy()
                        df_mostrar.columns = df_mostrar.columns.str.upper()
                        
                        st.dataframe(df_mostrar, hide_index=True)
                        st.session_state.plantillas = cargar_plantillas()
                        
                        tipo_certificado_actual = st.session_state.get('tipo_certificado_seleccionado', 'Regular')
                        st.session_state.grupos = clasificar_estudiantes_por_nota(
                            st.session_state.df_procesado, 
                            tipo_certificado_actual
                        )
                        
                        if st.session_state.grupos:
                            st.markdown("### üìã Preview de certificados a generar")
                            
                            col_prev1, col_prev2, col_prev3, col_prev4 = st.columns(4)
                            
                            with col_prev1:
                                cant_progresivos = len(st.session_state.grupos.get('grupo_1', pd.DataFrame()))
                                if cant_progresivos > 0:
                                    st.metric("üéì Progresivos", cant_progresivos)
                            
                            with col_prev2:
                                cant_participacion = len(st.session_state.grupos.get('grupo_2', pd.DataFrame()))
                                if cant_participacion > 0:
                                    st.metric("üìú Participaci√≥n", cant_participacion)
                            
                            with col_prev3:
                                cant_1p3p = len(st.session_state.grupos.get('grupo_3', pd.DataFrame()))
                                if cant_1p3p > 0:
                                    st.metric("‚úÖ 1P-3P Aprobados", cant_1p3p)
                            
                            with col_prev4:
                                cant_4p5s = len(st.session_state.grupos.get('grupo_4', pd.DataFrame()))
                                if cant_4p5s > 0:
                                    st.metric("‚úÖ 4P-5S Aprobados", cant_4p5s)
                            
                            total_certificados = cant_progresivos + cant_participacion + cant_1p3p + cant_4p5s
                            st.info(f"üéØ **Total de certificados a generar:** {total_certificados}")
                            
                            st.markdown("---")

                        generar_todos_certificados()

                        st.session_state.archivo_procesado = True
                    else:
                        st.error(mensaje)

    elif uploaded_file and st.session_state.archivo_procesado:
        st.success("‚úÖ Archivo ya procesado. Los certificados est√°n listos para descargar.")

    if st.session_state.certificados_generados and st.session_state.zip_buffer:
        nombre_archivo = st.session_state.get('nombre_archivo', '')
        nombre_base = os.path.splitext(nombre_archivo)[0] if nombre_archivo else "CERTIFICADOS"

        import re
        match = re.match(r'(.+?)_(1P-3P|4P-5S)', nombre_base)
        if match:
            nombre_colegio = match.group(1)
            rango_grados = match.group(2)
            nombre_limpio = f"{nombre_colegio}_{rango_grados}"
        else:
            nombre_limpio = nombre_base
        
        tipo_certificado = st.session_state.get('tipo_certificado_seleccionado', 'Regular')
        prefijo_tipo = "Regulares" if tipo_certificado == "Regular" else "Progresivos"
        usar_marca_agua = st.session_state.get('usar_marca_agua_seleccionado', False)

        if usar_marca_agua:
            zip_filename = f"{prefijo_tipo}_{nombre_limpio}_Preliminar.zip"
        else:
            zip_filename = f"{prefijo_tipo}_{nombre_limpio}.zip"
        
        st.download_button(
            label="üì• Descargar todos los certificados (ZIP)",
            data=st.session_state.zip_buffer,
            file_name=zip_filename,
            mime="application/zip"
        )
        
        st.markdown("---")
        if st.button("üîÑ Limpiar y Generar nuevos certificados", use_container_width=True, key="btn_regenerar_certificados"):
            st.session_state.archivo_procesado = False
            st.session_state.df_procesado = None     
            st.session_state.nombre_archivo = None   
            st.session_state.grupos = None
            st.session_state.plantillas = None
            st.session_state.certificados_generados = False
            st.session_state.zip_buffer = None
            st.session_state.tipo_certificado_seleccionado = None 
            st.session_state.usar_marca_agua_seleccionado = False 
            st.session_state.fecha_certificado_seleccionada = datetime.now().date()
            st.session_state.tab4_reset_counter += 1
            st.rerun()
        
    elif not uploaded_file:
        st.info("üëÜ Sube un archivo Excel para generar los certificados autom√°ticamente.")
        st.session_state.archivo_procesado = False

# TAB 5: GENERADOR DE INSIGNIAS
with tab5:
    st.markdown("## üèÖ Generador de Insignias (Docente y Alumno)")
    st.markdown("""
                **DESCRIPCI√ìN:**
                
                Permite emitir las insignias de reconocimiento a los docentes y alumnos manera autom√°tica a partir de la informaci√≥n registrada de diplomados, certificados y constancias, seg√∫n el archivo cargado. El proceso crea archivos PDF independientes para cada alumno o docente, los cuales pueden descargarse de manera consolidada en un archivo ZIP, facilitando la revisi√≥n, distribuci√≥n y trazabilidad de la informaci√≥n.
                """)
    st.info("""
    üìå **INSTRUCCIONES**
    1.	Columnas base requeridas (Columnas con valores completos):
            
        ‚Ä¢	ALUMNO: NOMBRE, PATERNO, MATERNO, CURSO, A√ëO.
            
        ‚Ä¢	DOCENTE: NOMBRE, PATERNO, MATERNO, TIPO DE INSIGNEA, A√ëO.
                    
    ‚ö†Ô∏è **IMPORTANTE**
    1.	Todos los datos se convertir√°n autom√°ticamente a MAY√öSCULAS.
    2.	La cabecera debe estar en la fila 9 del Excel.
    3.	Las columnas deben estar correctamente nombradas y pueden variar de orden.
    4.	El nombre del PDF ser√°: {TIPO DE INSIGNEA}_{NOMBRE COMPLETO}.pdf.
    """)
    
    tipo_insignia = st.radio(
            "Selecciona el tipo de insignia a generar:",
            ["ALUMNO", "DOCENTE"],
            horizontal=True
        )

    uploaded_file = st.file_uploader(
            f"Sube tu archivo Excel (.xlsx) con los datos de {tipo_insignia.lower()}s",
            type=["xlsx"],
            key=f"tab5_uploader_{st.session_state.tab5_reset_counter}",
            help="El archivo debe contener las columnas requeridas seg√∫n el tipo de insignia seleccionado"
        )
    
    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file, header=8)
            df = df.dropna(how='all')
            df = df.fillna('')
            
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.upper()

            columnas_base = ['NOMBRE', 'PATERNO', 'MATERNO', 'A√ëO']

            if tipo_insignia == "ALUMNO":
                columnas_requeridas = columnas_base + ['CURSO']
                mensaje_error = "‚ùå El archivo debe contener las columnas: NOMBRE, PATERNO, MATERNO, CURSO y A√ëO"
            else:  # DOCENTE
                columnas_requeridas = columnas_base + ['TIPO DE INSIGNEA']
                mensaje_error = "‚ùå El archivo debe contener las columnas: NOMBRE, PATERNO, MATERNO, TIPO DE INSIGNEA y A√ëO"
            
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]

            if columnas_faltantes:
                st.error(f"{mensaje_error}")
                st.error(f"üìã Columnas faltantes: {', '.join(columnas_faltantes)}")
                st.info(f"üìã Columnas encontradas en el archivo: {', '.join(df.columns.tolist())}")
                st.stop()

            df['IDENTIFICADOR'] = (
                df['NOMBRE'].astype(str).str.strip() + ' ' + 
                df['PATERNO'].astype(str).str.strip() + ' ' + 
                df['MATERNO'].astype(str).str.strip()
            )
            df['IDENTIFICADOR'] = df['IDENTIFICADOR'].str.replace(r'\s+', ' ', regex=True).str.strip()
            df['IDENTIFICADOR'] = df['IDENTIFICADOR'].str.replace('NAN', '', regex=True).str.strip()

            def es_vacio(valor):
                if pd.isna(valor):
                    return True
                valor_str = str(valor).strip().upper()
                return valor_str == '' or valor_str == 'NAN' or valor_str == 'NONE'
            
            filas_invalidas = []
            
            for idx, row in df.iterrows():
                campos_vacios = []
                
                if es_vacio(row.get('NOMBRE', '')):
                    campos_vacios.append('NOMBRE')
                if es_vacio(row.get('PATERNO', '')):
                    campos_vacios.append('PATERNO')
                if es_vacio(row.get('MATERNO', '')):
                    campos_vacios.append('MATERNO')
                if es_vacio(row.get('A√ëO', '')):
                    campos_vacios.append('A√ëO')
                
                if tipo_insignia == "ALUMNO":
                    if es_vacio(row.get('CURSO', '')):
                        campos_vacios.append('CURSO')
                else:  # DOCENTE
                    if es_vacio(row.get('TIPO DE INSIGNEA', '')):
                        campos_vacios.append('TIPO DE INSIGNEA')
                
                if campos_vacios:
                    filas_invalidas.append({
                        'fila': idx + 1,
                        'identificador': row.get('IDENTIFICADOR', 'N/A'),
                        'campos_vacios': ', '.join(campos_vacios)
                    })
            
            total_filas = len(df)

            if filas_invalidas:
                st.error(f"‚ùå VALIDACI√ìN FALLIDA: Se encontraron {len(filas_invalidas)} registro(s) con datos incompletos de {total_filas} totales")
                
                with st.expander(f"üîç Ver detalles de {len(filas_invalidas)} registro(s) incompleto(s)", expanded=True):
                    st.markdown("**Los siguientes registros tienen campos vac√≠os:**")
                    
                    df_invalidas = pd.DataFrame(filas_invalidas)
                    df_invalidas.columns = ['Fila Excel', 'Identificador', 'Campos Vac√≠os']
                    st.dataframe(df_invalidas, hide_index=True, use_container_width=True)
                
                st.error("üõë **PROCESO DETENIDO**: Todos los registros deben tener datos completos para continuar.")
                st.info("üìù **Instrucciones**: Corrije los registros incompletos en tu archivo Excel y vuelve a subirlo.")
                st.stop()
            else:
                st.success(f"‚úÖ Archivo validado correctamente: {len(df)} registros completos encontrados")
            
            with st.expander("üëÅÔ∏è Vista previa de los datos"):
                df_preview = df.drop(columns=['IDENTIFICADOR'])
                st.dataframe(df_preview, hide_index=True)
            
            if st.button("üé® Generar Insignias PDF", key="generar_insignias", type="primary", use_container_width=True):
                with st.spinner("Generando insignias..."):
                    temp_dir = "temp_insignias"
                    os.makedirs(temp_dir, exist_ok=True)
                    pdf_files = []
                    errores = []
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    total_rows = len(df)

                    for idx, row in df.iterrows():
                        try:
                            status_text.text(f"Procesando insignia {idx + 1} de {total_rows}...")
                            progress_bar.progress((idx + 1) / total_rows)
                            
                            if tipo_insignia == "ALUMNO":
                                imagen_fondo = "plantillas_insignias/ALUMNO.jpg"
                            else:  # DOCENTE
                                tipo_doc = str(row.get("TIPO DE INSIGNEA", "SENIOR")).upper()
                                if "ESPECIALISTA" in tipo_doc:
                                    imagen_fondo = "plantillas_insignias/DOCENTE_ESPECIALISTA.jpg"
                                else:
                                    imagen_fondo = "plantillas_insignias/DOCENTE_SENIOR.jpg"
                            
                            if not os.path.exists(imagen_fondo):
                                errores.append(f"Fila {idx+2}: No se encontr√≥ la imagen {imagen_fondo}")
                                continue
                            
                            img = Image.open(imagen_fondo)
                            draw = ImageDraw.Draw(img)
                            font_path = "fonts/trebuchet.ttf"
                            identificador = str(row.get("IDENTIFICADOR", "")).upper()
                            ano = str(row.get("A√ëO", "")).upper()
                            tipo_doc = str(row.get("TIPO DE INSIGNEA", "ALUMNO")).upper()
                            
                            if tipo_insignia == "ALUMNO":
                                curso = str(row.get("CURSO", "")).upper()

                                config_id = CONFIG_INSIGNIAS['IDENTIFICADOR']
                                draw_centered_text_adaptive(
                                        draw=draw,
                                        text=identificador,
                                        x_center=621,
                                        y_center=435,
                                        font_path=font_path,
                                        font_size_inicial=config_id['font_size_inicial'],
                                        max_width=config_id['max_width'],
                                        max_height=config_id['max_height'],
                                        min_font_size=config_id['min_font_size'],
                                        fill="white"
                                    )
                                    
                                config_curso = CONFIG_INSIGNIAS['CURSO']
                                draw_centered_text_adaptive(
                                        draw=draw,
                                        text=curso,
                                        x_center=621,
                                        y_center=677,
                                        font_path=font_path,
                                        font_size_inicial=config_curso['font_size_inicial'],
                                        max_width=config_curso['max_width'],
                                        max_height=config_curso['max_height'],
                                        min_font_size=config_curso['min_font_size'],
                                        fill="white"
                                    )
                                    
                                config_ano = CONFIG_INSIGNIAS['A√ëO']
                                draw_centered_text_adaptive(
                                        draw=draw,
                                        text=ano,
                                        x_center=621,
                                        y_center=926,
                                        font_path=font_path,
                                        font_size_inicial=config_ano['font_size_inicial'],
                                        max_width=config_ano['max_width'],
                                        max_height=config_ano['max_height'],
                                        min_font_size=config_ano['min_font_size'],
                                        fill="white"
                                    )
                                    
                                pdf_name = f"ALUMNO_{identificador}.pdf"
                                    
                            else:  # DOCENTE
                                config_id = CONFIG_INSIGNIAS['IDENTIFICADOR']
                                draw_centered_text_adaptive(
                                        draw=draw,
                                        text=identificador,
                                        x_center=641,
                                        y_center=536,
                                        font_path=font_path,
                                        font_size_inicial=config_id['font_size_inicial'],
                                        max_width=config_id['max_width'],
                                        max_height=config_id['max_height'],
                                        min_font_size=config_id['min_font_size'],
                                        fill="white"
                                    )
                                    
                                config_ano = CONFIG_INSIGNIAS['A√ëO']
                                draw_centered_text_adaptive(
                                        draw=draw,
                                        text=ano,
                                        x_center=641,
                                        y_center=905,
                                        font_path=font_path,
                                        font_size_inicial=config_ano['font_size_inicial'],
                                        max_width=config_ano['max_width'],
                                        max_height=config_ano['max_height'],
                                        min_font_size=config_ano['min_font_size'],
                                        fill="white"
                                    )
                                    
                                pdf_name = f"{tipo_doc}_{identificador}.pdf"
                            
                            pdf_name = pdf_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
                            pdf_path = os.path.join(temp_dir, pdf_name)
                            img_rgb = img.convert('RGB')
                            img_rgb.save(pdf_path, "PDF", resolution=100.0)
                            pdf_files.append(pdf_path)
                            
                        except Exception as e:
                            errores.append(f"Fila {idx+2}: {str(e)}")
                    
                    progress_bar.empty()
                    status_text.empty()
                    
                    if pdf_files:
                        st.success(f"‚úÖ Se generaron {len(pdf_files)} insignias correctamente")
                        
                        zip_filename = f"insignias_{tipo_insignia.lower()}.zip"
                        zip_path = os.path.join(temp_dir, zip_filename)
                        
                        with zipfile.ZipFile(zip_path, 'w') as zipf:
                            for pdf_file in pdf_files:
                                zipf.write(pdf_file, os.path.basename(pdf_file))
                        
                        with open(zip_path, 'rb') as f:
                            st.download_button(
                                label="üì¶ Descargar ZIP con todas las insignias",
                                data=f.read(),
                                file_name=zip_filename,
                                mime="application/zip",
                                use_container_width=True
                            )
                    
                    if errores:
                        st.warning(f"‚ö†Ô∏è Se encontraron {len(errores)} errores:")
                        with st.expander("Ver errores"):
                            for error in errores:
                                st.text(error)
                    
                    try:
                        import shutil
                        if os.path.exists(temp_dir):
                            shutil.rmtree(temp_dir)
                    except:
                        pass

            st.markdown("---")
            if st.button("üîÑ Limpiar y Generar nuevas Insignias", use_container_width=True, key="btn_regenerar_insigneas"):
                st.session_state.df_procesado = None
                st.session_state.grupos = None     
                st.session_state.plantillas = None   
                st.session_state.certificados_generados = False
                st.session_state.zip_buffer = None
                st.session_state.tab5_reset_counter += 1
                st.rerun()
                        
        except Exception as e:
            st.error(f"‚ùå Error al procesar el archivo: {str(e)}")
            st.exception(e)
    
    else:
        st.info("üëÜ Sube un archivo Excel para comenzar a generar insignias")